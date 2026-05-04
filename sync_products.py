import argparse
import glob
import os
from datetime import datetime
from typing import Dict, List

from sync_customers import read_csv, write_csv


def _pick_latest(paths: List[str]) -> str:
    if not paths:
        return ""
    return max(paths, key=lambda p: os.path.getmtime(p))


def build_product_sync(args: argparse.Namespace) -> int:
    year_month = args.year_month
    base_dir = args.base_dir
    items_master = args.items_master
    items_sync = args.items_sync
    out_path = args.out_path.replace("{year_month}", year_month)

    if not os.path.exists(items_master):
        print(f"ERROR: items master not found: {items_master}")
        return 2
    if not os.path.exists(items_sync):
        print(f"ERROR: products sync not found: {items_sync}")
        return 2

    inv_glob = os.path.join(base_dir, "**", f"{year_month}_invoice_lines.csv")
    cn_glob = os.path.join(base_dir, "**", f"{year_month}_credit_note_lines.csv")
    inv_paths = glob.glob(inv_glob, recursive=True)
    cn_paths = glob.glob(cn_glob, recursive=True)
    if not inv_paths:
        print(f"ERROR: invoice lines not found for {year_month} (search: {inv_glob})")
        return 2
    if not cn_paths:
        print(f"ERROR: credit note lines not found for {year_month} (search: {cn_glob})")
        return 2

    invoice_lines = _pick_latest(inv_paths)
    credit_lines = _pick_latest(cn_paths)
    if len(inv_paths) > 1:
        print(f"WARNING: multiple invoice_lines found, using latest: {invoice_lines}")
    if len(cn_paths) > 1:
        print(f"WARNING: multiple credit_note_lines found, using latest: {credit_lines}")

    # Load invoice + credit note lines
    _, inv_rows = read_csv(invoice_lines)
    _, cn_rows = read_csv(credit_lines)
    item_records = set()
    for row in inv_rows + cn_rows:
        rec = (row.get("ItemRecordNumber") or "").strip()
        if rec:
            item_records.add(rec)
    item_records.discard("0")

    # Load items master + sync
    _, master_rows = read_csv(items_master)
    master_by_record = { (r.get("ItemRecordNumber") or "").strip(): r for r in master_rows }

    _, sync_rows = read_csv(items_sync)
    sync_by_record = { (r.get("ItemRecordNumber") or "").strip(): r for r in sync_rows }

    fieldnames = [
        "ItemRecordNumber",
        "ItemID",
        "ItemDescription",
        "SalesDescription",
        "Barcode",
        "OdooVariantId",
        "OdooName",
        "OdooVariantName",
        "OdooItemCode",
        "OdooColor",
        "Reason",
    ]
    rows_out: List[Dict[str, str]] = []
    for rec in sorted(item_records, key=lambda x: int(x) if x.isdigit() else x):
        sync = sync_by_record.get(rec) or {}
        reason = ""
        if not sync:
            reason = "NO_SYNC"
        else:
            if not (sync.get("OdooVariantId") or "").strip():
                reason = "NO_ODOO"
        if not reason:
            continue
        master = master_by_record.get(rec) or {}
        rows_out.append({
            "ItemRecordNumber": rec,
            "ItemID": (sync.get("ItemID") or master.get("ItemID") or "").strip(),
            "ItemDescription": (sync.get("ItemDescription") or master.get("ItemDescription") or "").strip(),
            "SalesDescription": (master.get("SalesDescription") or "").strip(),
            "Barcode": (master.get("UPC_SKU") or "").strip(),
            "OdooVariantId": (sync.get("OdooVariantId") or "").strip(),
            "OdooName": (sync.get("OdooName") or "").strip(),
            "OdooVariantName": (sync.get("OdooVariantName") or "").strip(),
            "OdooItemCode": (sync.get("OdooItemCode") or "").strip(),
            "OdooColor": (sync.get("OdooColor") or "").strip(),
            "Reason": reason,
        })

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    write_csv(out_path, fieldnames, rows_out)
    print(f"OK: product sync rows: {len(rows_out)} -> {out_path}")
    return 0


def build_items_sync_new(args: argparse.Namespace) -> int:
    items_sync = args.items_sync
    out_path = args.out_path
    barcode_digits = None if args.barcode_digits == 0 else args.barcode_digits
    invoice_base_dir = args.invoice_base_dir

    if not os.path.exists(items_sync):
        print(f"ERROR: products sync not found: {items_sync}")
        return 2

    fields, rows = read_csv(items_sync)
    if not fields:
        print(f"ERROR: products sync has no headers: {items_sync}")
        return 2

    # Build invoiced set for 2026 (Feb + Mar)
    invoiced = set()
    if invoice_base_dir:
        targets = [
            os.path.join(invoice_base_dir, "**", "2026_02_invoice_lines.csv"),
            os.path.join(invoice_base_dir, "**", "2026_03_invoice_lines.csv"),
            os.path.join(invoice_base_dir, "**", "2026_04_invoice_lines.csv"),
        ]
        matched = []
        for pattern in targets:
            matched.extend(glob.glob(pattern, recursive=True))
        if not matched:
            print("WARNING: no invoice_lines found for 2026_02, 2026_03, or 2026_04")
        for path in matched:
            try:
                _, inv_rows = read_csv(path)
            except Exception:
                continue
            for r in inv_rows:
                rec = (r.get("ItemRecordNumber") or "").strip()
                if rec:
                    invoiced.add(rec)

    def is_barcode_ok(value: str) -> bool:
        v = (value or "").strip()
        if not v:
            return False
        if barcode_digits is None:
            return True
        return v.isdigit() and len(v) >= barcode_digits

    def is_excluded_sale_desc(value: str) -> bool:
        v = (value or "").strip().upper()
        if not v:
            return False
        return v.startswith("DERAPAGE") or v.startswith("ECLIPSE") or v.startswith("90 PIECE")

    def is_force_nobarcode_item(r: Dict[str, str]) -> bool:
        return (r.get("ItemID") or "").strip().upper() == "NW77PLAQUE"

    def is_inactive(value: str) -> bool:
        v = (value or "").strip().lower()
        return v in {"1", "true", "yes", "y"}

    filtered = []
    # Sort by Description for Sales so brands cluster together
    rows_sorted = sorted(rows, key=lambda r: (r.get("ItemDescriptionForSale") or "").upper())
    for r in rows_sorted:
        if (r.get("OdooVariantId") or "").strip():
            continue
        if is_force_nobarcode_item(r):
            continue
        if not is_barcode_ok(r.get("Barcode", "")):
            continue
        if is_excluded_sale_desc(r.get("ItemDescriptionForSale", "")):
            continue
        invoiced_2026 = "X" if (r.get("ItemRecordNumber") or "").strip() in invoiced else ""
        r["Invoiced2026"] = invoiced_2026
        # Do not send inactive Sage products unless they were sold in 2026.
        if is_inactive(r.get("ItemIsInactive", "")) and invoiced_2026 != "X":
            continue
        filtered.append(r)

    out_fields = list(fields)
    if "Invoiced2026" not in out_fields:
        if "LastLookupAt" in out_fields:
            idx = out_fields.index("LastLookupAt")
            out_fields.insert(idx, "Invoiced2026")
        else:
            out_fields.append("Invoiced2026")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    write_csv(out_path, out_fields, filtered)
    print(f"OK: items sync NEW rows: {len(filtered)} -> {out_path}")
    return 0


def build_products_sync_nobarcode_new(args: argparse.Namespace) -> int:
    items_sync = args.items_sync
    out_path = args.out_path
    barcode_digits = None if args.barcode_digits == 0 else args.barcode_digits
    invoice_base_dir = args.invoice_base_dir

    if not os.path.exists(items_sync):
        print(f"ERROR: products sync not found: {items_sync}")
        return 2

    fields, rows = read_csv(items_sync)
    if not fields:
        print(f"ERROR: products sync has no headers: {items_sync}")
        return 2

    # Build invoiced set for 2026 (Feb + Mar)
    invoiced = set()
    if invoice_base_dir:
        targets = [
            os.path.join(invoice_base_dir, "**", "2026_02_invoice_lines.csv"),
            os.path.join(invoice_base_dir, "**", "2026_03_invoice_lines.csv"),
            os.path.join(invoice_base_dir, "**", "2026_04_invoice_lines.csv"),
        ]
        matched = []
        for pattern in targets:
            matched.extend(glob.glob(pattern, recursive=True))
        if not matched:
            print("WARNING: no invoice_lines found for 2026_02, 2026_03, or 2026_04")
        for path in matched:
            try:
                _, inv_rows = read_csv(path)
            except Exception:
                continue
            for r in inv_rows:
                rec = (r.get("ItemRecordNumber") or "").strip()
                if rec:
                    invoiced.add(rec)

    def is_barcode_bad(value: str) -> bool:
        v = (value or "").strip()
        if not v:
            return True
        if barcode_digits is None:
            return False
        return (not v.isdigit()) or len(v) < barcode_digits

    def is_excluded_sale_desc(value: str) -> bool:
        v = (value or "").strip().upper()
        if not v:
            return False
        return v.startswith("DERAPAGE") or v.startswith("ECLIPSE") or v.startswith("90 PIECE")

    def is_force_nobarcode_item(r: Dict[str, str]) -> bool:
        return (r.get("ItemID") or "").strip().upper() == "NW77PLAQUE"

    def is_inactive(value: str) -> bool:
        v = (value or "").strip().lower()
        return v in {"1", "true", "yes", "y"}

    filtered = []
    # Sort by Description for Sales so brands cluster together
    rows_sorted = sorted(rows, key=lambda r: (r.get("ItemDescriptionForSale") or "").upper())
    for r in rows_sorted:
        if (r.get("OdooVariantId") or "").strip():
            continue
        if not is_barcode_bad(r.get("Barcode", "")) and not is_force_nobarcode_item(r):
            continue
        if is_excluded_sale_desc(r.get("ItemDescriptionForSale", "")):
            continue
        invoiced_2026 = "X" if (r.get("ItemRecordNumber") or "").strip() in invoiced else ""
        r["Invoiced2026"] = invoiced_2026
        # Do not send inactive Sage products unless they were sold in 2026.
        if is_inactive(r.get("ItemIsInactive", "")) and invoiced_2026 != "X":
            continue
        filtered.append(r)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    out_fields = list(fields)
    if "Invoiced2026" not in out_fields:
        if "LastLookupAt" in out_fields:
            idx = out_fields.index("LastLookupAt")
            out_fields.insert(idx, "Invoiced2026")
        else:
            out_fields.append("Invoiced2026")
    write_csv(out_path, out_fields, filtered)
    print(f"OK: products sync nobarcode NEW rows: {len(filtered)} -> {out_path}")
    return 0


def build_products_import(args: argparse.Namespace) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        load_workbook = None

    if load_workbook is None:
        print("ERROR: openpyxl not available for XLSX export")
        return 2

    sync_path = args.sync_path
    template_path = args.template_path

    if not os.path.exists(sync_path):
        print(f"ERROR: products sync NEW not found: {sync_path}")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2

    master_root = os.path.dirname(sync_path)
    odoo_imports = os.path.join(master_root, "odoo_imports")
    os.makedirs(odoo_imports, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d")
    out_xlsx = os.path.join(odoo_imports, f"{stamp}_products_NEW.xlsx")

    _, rows = read_csv(sync_path)

    wb = load_workbook(template_path)
    ws = wb["Products"] if "Products" in wb.sheetnames else wb.active
    # Keep only the first (active) sheet to avoid carrying template extras.
    for sheet_name in list(wb.sheetnames):
        if wb[sheet_name] is not ws:
            wb.remove(wb[sheet_name])
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header:
            header_map[str(header).strip()] = col_idx

    def set_cell(col_name: str, value) -> None:
        col_idx = header_map.get(col_name)
        if col_idx:
            ws.cell(row=row_idx, column=col_idx, value=value)
    ws.delete_rows(2, ws.max_row)
    row_idx = 2

    rows_sorted = sorted(rows, key=lambda r: (r.get("ItemDescriptionForSale") or "").upper())
    for r in rows_sorted:
        item_id = (r.get("ItemID") or "").strip()
        barcode = (r.get("Barcode") or "").strip()
        desc_sales = (r.get("ItemDescriptionForSale") or "").strip()
        desc_item = (r.get("ItemDescription") or "").strip()

        set_cell("x", "E")
        set_cell("id", item_id)
        set_cell("barcode", barcode)
        set_cell("if_favorite", "FALSE")
        set_cell("is_storable", "TRUE")
        set_cell("Description for Sales", desc_sales)
        set_cell("Item Description", desc_item)

        row_idx += 1

    wb.save(out_xlsx)
    print(f"OK: products NEW import rows: {len(rows)} -> {out_xlsx}")
    return 0


def build_products_nobarcode_import(args: argparse.Namespace) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        load_workbook = None

    if load_workbook is None:
        print("ERROR: openpyxl not available for XLSX export")
        return 2

    sync_path = args.sync_path
    template_path = args.template_path

    if not os.path.exists(sync_path):
        print(f"ERROR: products nobarcode sync not found: {sync_path}")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2

    master_root = os.path.dirname(sync_path)
    odoo_imports = os.path.join(master_root, "odoo_imports")
    os.makedirs(odoo_imports, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d")
    out_xlsx = os.path.join(odoo_imports, f"{stamp}_products_nobarcode_NEW.xlsx")

    _, rows = read_csv(sync_path)

    wb = load_workbook(template_path)
    ws = wb["Products"] if "Products" in wb.sheetnames else wb.active
    for sheet_name in list(wb.sheetnames):
        if wb[sheet_name] is not ws:
            wb.remove(wb[sheet_name])

    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header:
            header_map[str(header).strip()] = col_idx

    def set_cell(col_name: str, value) -> None:
        col_idx = header_map.get(col_name)
        if col_idx:
            ws.cell(row=row_idx, column=col_idx, value=value)

    def is_priority_row(r: Dict[str, str]) -> bool:
        if (r.get("Invoiced2026") or "").strip().upper() == "X":
            return True
        desc = (r.get("ItemDescription") or "").strip().upper()
        return (
            desc.startswith("ERKERS ")
            or desc.startswith("BA&SH ")
            or desc.startswith("NW 77TH ")
            or desc.startswith("MONOQOOL ")
        )

    ws.delete_rows(2, ws.max_row)
    row_idx = 2

    rows_sorted = sorted(rows, key=lambda r: (r.get("ItemDescriptionForSale") or "").upper())
    selected = [r for r in rows_sorted if is_priority_row(r)]
    for r in selected:
        item_id = (r.get("ItemID") or "").strip()
        barcode = (r.get("Barcode") or "").strip()
        desc_sales = (r.get("ItemDescriptionForSale") or "").strip()
        desc_item = (r.get("ItemDescription") or "").strip()

        set_cell("x", "E")
        set_cell("id", item_id)
        set_cell("barcode", barcode)
        set_cell("if_favorite", "FALSE")
        set_cell("is_storable", "TRUE")
        set_cell("Description for Sales", desc_sales)
        set_cell("Item Description", desc_item)

        row_idx += 1

    wb.save(out_xlsx)
    print(f"OK: products nobarcode import rows: {len(selected)} -> {out_xlsx}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Products sync utilities (Sage <-> Odoo templates)")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p6 = sub.add_parser("build_product_sync", help="Build product sync for a given YYYY_MM using invoice + credit note lines")
    p6.add_argument("year_month", help="YYYY_MM (e.g., 2026_03)")
    p6.add_argument("--base-dir", default=r"ENZO-Sage50")
    p6.add_argument("--items-master", default=r"ENZO-Sage50\_master_sage\items.csv")
    p6.add_argument("--items-sync", default=r"ENZO-Sage50\_master\products_sync.csv")
    p6.add_argument("--out-path", default=r"ENZO-Sage50\_master\{year_month}_products_sync.csv")
    p6.set_defaults(func=build_product_sync)

    p7 = sub.add_parser("build_items_sync_new", help="Build products_sync_NEW with filters (no Odoo ID, active, barcode)")
    p7.add_argument("--sync-path", default=r"ENZO-Sage50\_master\products_sync.csv")
    p7.add_argument("--out-path", default=r"ENZO-Sage50\_master\products_sync_NEW.csv")
    p7.add_argument("--base-dir", default=r"ENZO-Sage50")
    p7.add_argument("--barcode-len", type=int, default=12, help="Require barcode to have exactly N digits (default: 12). Use 0 to disable.")
    p7.set_defaults(func=build_items_sync_new)

    p7b = sub.add_parser("build_products_sync_nobarcode_new", help="Build products_sync_nobarcode_NEW (no Odoo ID, empty/short barcode)")
    p7b.add_argument("--sync-path", default=r"ENZO-Sage50\_master\products_sync.csv")
    p7b.add_argument("--out-path", default=r"ENZO-Sage50\_master\products_sync_nobarcode_NEW.csv")
    p7b.add_argument("--barcode-min", type=int, default=12, help="Require barcode to have at least N digits (default: 12). Use 0 to disable.")
    p7b.add_argument("--base-dir", default=r"ENZO-Sage50")
    p7b.set_defaults(func=build_products_sync_nobarcode_new)

    p7c = sub.add_parser("build_products_import", help="Build products import XLSX from products_sync_NEW")
    p7c.add_argument("--sync-path", default=r"ENZO-Sage50\_master\products_sync_NEW.csv")
    p7c.add_argument("--template-path", default=r"ENZO-Sage50\_master\odoo_templates\products.xlsx")
    p7c.set_defaults(func=build_products_import)

    p7d = sub.add_parser("build_products_nobarcode_import", help="Build products import XLSX from products_sync_nobarcode_NEW")
    p7d.add_argument("--sync-path", default=r"ENZO-Sage50\_master\products_sync_nobarcode_NEW.csv")
    p7d.add_argument("--template-path", default=r"ENZO-Sage50\_master\odoo_templates\products.xlsx")
    p7d.set_defaults(func=build_products_nobarcode_import)

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
