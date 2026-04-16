import argparse
import os
from datetime import datetime
from typing import Dict, List
from collections import Counter

from parity_utils import (
    load_country_name_to_code,
    load_country_parity,
    load_state_parity,
    normalize_country,
)
from sync_customers import normalize_name, read_csv, sanitize_external_id, write_csv


def _norm_text(value: str) -> str:
    return normalize_name(value or "")


def _norm_country_compare(value: str) -> str:
    normalized = _norm_text(value)
    aliases = {
        "us": "united states",
        "usa": "united states",
        "united states of america": "united states",
        "ca": "canada",
        "can": "canada",
        "gb": "united kingdom",
        "uk": "united kingdom",
    }
    return aliases.get(normalized, normalized)


def _norm_addr(street: str, street2: str, city: str, state: str, zip_code: str) -> str:
    parts = [street, street2, city, state, zip_code]
    return " ".join([_norm_text(p) for p in parts if p]).strip()


def _load_customers_by_record(customers_master: str) -> Dict[str, Dict[str, str]]:
    if not os.path.exists(customers_master):
        return {}
    _, rows = read_csv(customers_master)
    by_record: Dict[str, Dict[str, str]] = {}
    for r in rows:
        crn = (r.get("CustomerRecordNumber") or "").strip()
        if crn:
            by_record[crn] = r
    return by_record


def _load_addresses_by_record(address_master: str) -> Dict[str, Dict[str, str]]:
    _, rows = read_csv(address_master)
    by_record: Dict[str, Dict[str, str]] = {}
    for r in rows:
        arn = (r.get("AddressRecordNumber") or "").strip()
        if arn:
            by_record[arn] = r
    return by_record


def _load_odoo_delivery_by_parent(odoo_delivery_csv: str) -> Dict[str, List[Dict[str, str]]]:
    existing: Dict[str, List[Dict[str, str]]] = {}
    if not os.path.exists(odoo_delivery_csv):
        return existing
    _, rows = read_csv(odoo_delivery_csv)
    for r in rows:
        parent_id = (r.get("ParentId") or "").strip()
        if not parent_id:
            continue
        existing.setdefault(parent_id, []).append(r)
    return existing


def _normalize_country(
    raw_country: str,
    country_parity: Dict[str, str],
    country_name_to_code: Dict[str, str],
) -> str:
    return normalize_country(raw_country, country_parity, country_name_to_code)


def _match_delivery(existing_rows: List[Dict[str, str]], sage_reference: str, name: str, addr_key: str) -> Dict[str, str]:
    if not existing_rows:
        return {}
    # Best match: exact SAGE reference already stored in Odoo partner.ref.
    ref = (sage_reference or "").strip()
    odoo_refs = [((r.get("OdooRef") or "").strip()) for r in existing_rows]
    has_any_odoo_ref = any(odoo_refs)
    if ref:
        for r in existing_rows:
            odoo_ref = (r.get("OdooRef") or "").strip()
            if odoo_ref and odoo_ref == ref:
                return r
        # Safety: if Odoo has refs for this parent but none match exactly, do not
        # fall back to name/address to avoid cross-linking different delivery records.
        if has_any_odoo_ref:
            return {}
    n_name = _norm_text(name)
    for r in existing_rows:
        r_name = _norm_text(r.get("OdooName", ""))
        if n_name and r_name and n_name == r_name:
            return r
    if addr_key:
        for r in existing_rows:
            r_key = _norm_addr(
                r.get("Street", ""),
                r.get("Street2", ""),
                r.get("City", ""),
                r.get("State", ""),
                r.get("Zip", ""),
            )
            if r_key and r_key == addr_key:
                return r
    return {}


def _field_mismatches(row: Dict[str, str], existing: Dict[str, str]) -> List[str]:
    checks = [
        ("name", row.get("DeliveryName", ""), existing.get("OdooName", "")),
        ("email", row.get("Email", ""), existing.get("OdooEmail", "")),
        ("street", row.get("Street", ""), existing.get("Street", "")),
        ("street2", row.get("Street2", ""), existing.get("Street2", "")),
        ("city", row.get("City", ""), existing.get("City", "")),
        ("zip", row.get("Zip", ""), existing.get("Zip", "")),
        ("state", row.get("OdooState", "") or row.get("State", ""), existing.get("State", "")),
    ]
    mismatches = [name for name, left, right in checks if _norm_text(left) != _norm_text(right)]
    left_phone = "".join(ch for ch in (row.get("Phone", "") or "") if ch.isdigit())
    right_phone = "".join(ch for ch in (existing.get("OdooPhone", "") or "") if ch.isdigit())
    if left_phone != right_phone:
        mismatches.append("phone")
    return mismatches


def build_addresses_sync(args: argparse.Namespace) -> int:
    contacts_master = args.contacts_master
    address_master = args.address_master
    customers_sync = args.customers_sync
    odoo_delivery = args.odoo_delivery
    country_parity_path = args.country_parity
    state_parity_path = args.state_parity
    countries_odoo_path = args.countries_odoo
    out_path = args.out_path

    if not os.path.exists(contacts_master):
        print(f"ERROR: contacts master not found: {contacts_master}")
        return 2
    if not os.path.exists(address_master):
        print(f"ERROR: address master not found: {address_master}")
        return 2

    customers_by_record = _load_customers_by_record(customers_sync)
    addresses_by_record = _load_addresses_by_record(address_master)
    existing_by_parent = _load_odoo_delivery_by_parent(odoo_delivery)
    country_parity = load_country_parity(country_parity_path)
    country_name_to_code = load_country_name_to_code(countries_odoo_path)
    state_parity = load_state_parity(state_parity_path)

    _, contact_rows = read_csv(contacts_master)

    fieldnames = [
        "ContactRecordNumber",
        "CustomerRecordNumber",
        "CustomerID",
        "CustomerName",
        "AddressRecordNumber",
        "DeliveryName",
        "AddressTypeNumber",
        "AddressTypeDesc",
        "Email",
        "Phone",
        "Street",
        "Street2",
        "City",
        "State",
        "Zip",
        "Country",
        "OdooState",
        "OdooCountry",
        "Notes",
        "OdooAddressId",
        "OdooAddressExternalId",
        "OdooParentId",
        "DeliverySyncStatus",
        "DeliveryMismatchFields",
        "LastLookupAt",
    ]

    rows_out: List[Dict[str, str]] = []
    now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for c in contact_rows:
        customer_record = (c.get("CustomerRecord") or "").strip()
        address_record = (c.get("AddressRecordNumber") or "").strip()
        if not customer_record or not address_record:
            continue
        if (c.get("IsPrimaryContact") or "").strip() == "1":
            # Skip primary contact (company main row)
            continue
        delivery_name = (c.get("CompanyName") or "").strip()
        if not delivery_name:
            continue

        addr = addresses_by_record.get(address_record, {})
        cust = customers_by_record.get(customer_record, {})

        addr_key = _norm_addr(
            addr.get("AddressLine1", ""),
            addr.get("AddressLine2", ""),
            addr.get("City", ""),
            addr.get("State", ""),
            addr.get("Zip", ""),
        )

        raw_zip = (addr.get("Zip") or "").strip()
        raw_country = (addr.get("Country") or "").strip()
        # Sage data cleanup: some rows carry ZIP in Country (e.g. "33166") and leave Zip empty.
        if raw_country.isdigit() and len(raw_country) in {4, 5, 6} and not raw_zip:
            raw_zip = raw_country
            raw_country = ""
        mapped_country = _normalize_country(raw_country, country_parity, country_name_to_code)

        raw_state = (addr.get("State") or "").strip()
        state_info = state_parity.get(raw_state, {})
        mapped_state = state_info.get("state_name") or raw_state
        if mapped_state and state_info.get("country_name"):
            inferred = country_name_to_code.get(state_info.get("country_name", ""), "")
            if inferred:
                mapped_state = f"{mapped_state} ({inferred})"
                if not mapped_country:
                    mapped_country = inferred

        odoo_parent_id = ""
        if cust:
            odoo_parent_id = (cust.get("OdooId") or "").strip()

        match_row: Dict[str, str] = {}
        if odoo_parent_id:
            match_row = _match_delivery(
                existing_by_parent.get(odoo_parent_id, []),
                (c.get("RecordNumber") or "").strip(),
                delivery_name,
                addr_key,
            )

        match_id = (match_row.get("OdooId") or "").strip()
        strict_ref_match = False
        if match_row:
            strict_ref_match = ((c.get("RecordNumber") or "").strip() == ((match_row.get("OdooRef") or "").strip()))
        mismatches = []
        if match_row:
            draft_row = {
                "DeliveryName": delivery_name,
                "Email": (c.get("Email") or "").strip(),
                "Phone": (c.get("Telephone1") or "").strip(),
                "Street": (addr.get("AddressLine1") or "").strip(),
                "Street2": (addr.get("AddressLine2") or "").strip(),
                "City": (addr.get("City") or "").strip(),
                "State": raw_state,
                "Zip": raw_zip,
                "Country": raw_country,
                "OdooState": mapped_state,
                "OdooCountry": mapped_country,
            }
            mismatches = _field_mismatches(draft_row, match_row)

        rows_out.append({
            "ContactRecordNumber": (c.get("RecordNumber") or "").strip(),
            "CustomerRecordNumber": customer_record,
            "CustomerID": (cust.get("CustomerID") or "").strip(),
            "CustomerName": (cust.get("Customer_Bill_Name") or "").strip(),
            "AddressRecordNumber": address_record,
            "DeliveryName": delivery_name,
            "AddressTypeNumber": (addr.get("AddressTypeNumber") or "").strip(),
            "AddressTypeDesc": (addr.get("AddressTypeDesc") or "").strip(),
            "Email": (c.get("Email") or "").strip(),
            "Phone": (c.get("Telephone1") or "").strip(),
            "Street": (addr.get("AddressLine1") or "").strip(),
            "Street2": (addr.get("AddressLine2") or "").strip(),
            "City": (addr.get("City") or "").strip(),
            "State": raw_state,
            "Zip": raw_zip,
            "Country": raw_country,
            "OdooState": mapped_state,
            "OdooCountry": mapped_country,
            "Notes": f"{addr.get('AddressTypeNumber','')} | {addr.get('AddressTypeDesc','')}".strip(),
            "OdooAddressId": match_id,
            "OdooAddressExternalId": (match_row.get("OdooExternalId") or "").strip(),
            "OdooParentId": odoo_parent_id,
            "DeliverySyncStatus": (
                "UPDATE"
                if (match_id and mismatches and strict_ref_match)
                else ("MATCH" if match_id and (not mismatches or strict_ref_match) else "NEW")
            ),
            "DeliveryMismatchFields": (
                "|".join(mismatches)
                if strict_ref_match or not (match_id and mismatches)
                else "fallback_non_strict_match_blocked"
            ),
            "LastLookupAt": now_stamp,
        })

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    write_csv(out_path, fieldnames, rows_out)
    print(f"OK: addresses sync rows: {len(rows_out)} -> {out_path}")
    return 0


def build_delivery_import(args: argparse.Namespace) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        load_workbook = None

    if load_workbook is None:
        print("ERROR: openpyxl not available for XLSX export")
        return 2

    sync_path = args.sync_path
    template_path = args.template_path

    if not os.path.exists(sync_path):
        print(f"ERROR: delivery sync file not found: {sync_path}")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2

    master_root = os.path.dirname(sync_path)
    odoo_imports = os.path.join(master_root, "odoo_imports")
    os.makedirs(odoo_imports, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d")
    out_xlsx = os.path.join(odoo_imports, f"{stamp}_customers_delivery_NEW.xlsx")
    out_csv = os.path.join(master_root, "customer_delivery_addresses_sync_NEW.csv")

    _, sync_rows = read_csv(sync_path)

    wb = load_workbook(template_path)
    ws = wb["Partners"] if "Partners" in wb.sheetnames else wb.active
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header:
            header_map[str(header).strip()] = col_idx

    def set_cell(col_name: str, value: str) -> None:
        col_idx = header_map.get(col_name)
        if col_idx:
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.delete_rows(2, ws.max_row)
    row_idx = 2
    emitted = 0
    new_rows: List[Dict[str, str]] = []
    for r in sync_rows:
        if (r.get("OdooAddressId") or "").strip():
            continue
        parent_id = (r.get("OdooParentId") or "").strip()
        if not parent_id:
            continue
        name = (r.get("DeliveryName") or "").strip()
        if not name:
            continue
        new_rows.append(r)

        customer_id = (r.get("CustomerID") or "").strip()
        contact_record = (r.get("ContactRecordNumber") or "").strip()
        raw_ext = f"{customer_id}_{contact_record}" if customer_id and contact_record else ""
        ext_id = sanitize_external_id(raw_ext)

        state = (r.get("OdooState") or "").strip() or (r.get("State") or "").strip()
        country = (r.get("OdooCountry") or "").strip() or (r.get("Country") or "").strip()

        set_cell("External_ID", ext_id)
        set_cell("Parent/Database ID", parent_id)
        set_cell("Reference", (r.get("ContactRecordNumber") or "").strip())
        set_cell("is_company", 0)
        set_cell("type", "delivery")
        set_cell("Name", name)
        set_cell("Email", (r.get("Email") or "").strip())
        set_cell("Phone", (r.get("Phone") or "").strip())
        set_cell("Street", (r.get("Street") or "").strip())
        set_cell("Street2", (r.get("Street2") or "").strip())
        set_cell("City", (r.get("City") or "").strip())
        set_cell("State", state)
        set_cell("ZIP", (r.get("Zip") or "").strip())
        set_cell("Country", country)
        set_cell("Notes", (r.get("Notes") or "").strip())
        row_idx += 1
        emitted += 1

    wb.save(out_xlsx)
    fieldnames = []
    if new_rows:
        fieldnames = list(new_rows[0].keys())
    write_csv(out_csv, fieldnames, new_rows)
    print(f"OK: delivery new rows: {emitted} -> {out_xlsx}")
    print(f"OK: delivery NEW sync -> {out_csv}")
    return 0


def build_delivery_update(args: argparse.Namespace) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        load_workbook = None

    if load_workbook is None:
        print("ERROR: openpyxl not available for XLSX export")
        return 2

    sync_path = args.sync_path
    template_path = args.template_path
    if not os.path.exists(sync_path):
        print(f"ERROR: delivery sync file not found: {sync_path}")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2

    master_root = os.path.dirname(sync_path)
    odoo_update = os.path.join(master_root, "odoo_UPDATE")
    os.makedirs(odoo_update, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d")
    out_xlsx = os.path.join(odoo_update, f"{stamp}_customers_delivery_UPDATE.xlsx")
    out_csv = os.path.join(master_root, "customer_delivery_addresses_sync_UPDATE.csv")
    conflicts_csv = os.path.join(master_root, "customer_delivery_addresses_sync_UPDATE_CONFLICTS.csv")

    _, sync_rows = read_csv(sync_path)
    candidates = [
        r for r in sync_rows
        if (r.get("DeliverySyncStatus") or "").strip().upper() == "UPDATE"
        and (r.get("OdooParentId") or "").strip()
        and (r.get("OdooAddressExternalId") or "").strip()
    ]
    ext_counts = Counter((r.get("OdooAddressExternalId") or "").strip() for r in candidates)
    update_rows = [r for r in candidates if ext_counts[(r.get("OdooAddressExternalId") or "").strip()] == 1]
    conflict_rows = [r for r in candidates if ext_counts[(r.get("OdooAddressExternalId") or "").strip()] > 1]

    wb = load_workbook(template_path)
    ws = wb["Partners"] if "Partners" in wb.sheetnames else wb.active
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header:
            header_map[str(header).strip()] = col_idx

    def set_cell(col_name: str, value: str) -> None:
        col_idx = header_map.get(col_name)
        if col_idx:
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.delete_rows(2, ws.max_row)
    row_idx = 2
    emitted = 0
    for r in update_rows:
        ext_id = (r.get("OdooAddressExternalId") or "").strip()
        if ext_id.startswith("__import__."):
            ext_id = ext_id.split(".", 1)[1]
        parent_id = (r.get("OdooParentId") or "").strip()
        name = (r.get("DeliveryName") or "").strip()
        if not ext_id or not parent_id or not name:
            continue
        state = (r.get("OdooState") or "").strip() or (r.get("State") or "").strip()
        country = (r.get("OdooCountry") or "").strip() or (r.get("Country") or "").strip()

        set_cell("External_ID", ext_id)
        set_cell("Parent/Database ID", parent_id)
        set_cell("Reference", (r.get("ContactRecordNumber") or "").strip())
        set_cell("is_company", 0)
        set_cell("type", "delivery")
        set_cell("Name", name)
        set_cell("Email", (r.get("Email") or "").strip())
        set_cell("Phone", (r.get("Phone") or "").strip())
        set_cell("Street", (r.get("Street") or "").strip())
        set_cell("Street2", (r.get("Street2") or "").strip())
        set_cell("City", (r.get("City") or "").strip())
        set_cell("State", state)
        set_cell("ZIP", (r.get("Zip") or "").strip())
        set_cell("Country", country)
        set_cell("Notes", (r.get("Notes") or "").strip())
        row_idx += 1
        emitted += 1

    wb.save(out_xlsx)
    fieldnames = list(candidates[0].keys()) if candidates else []
    write_csv(out_csv, fieldnames, update_rows)
    write_csv(conflicts_csv, fieldnames, conflict_rows)
    print(f"OK: delivery update rows: {emitted} -> {out_xlsx}")
    print(f"OK: delivery UPDATE sync -> {out_csv}")
    print(f"OK: delivery UPDATE conflicts -> {conflicts_csv} ({len(conflict_rows)} rows)")
    return 0
