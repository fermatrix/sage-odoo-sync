import argparse
import os
from datetime import datetime
from typing import Dict, List

from sync_customers import normalize_name, read_csv, sanitize_external_id, write_csv


def _load_primary_contacts(contacts_master_path: str) -> Dict[str, Dict[str, str]]:
    primary_contact_by_customer_record: Dict[str, Dict[str, str]] = {}
    if os.path.exists(contacts_master_path):
        _, contact_rows = read_csv(contacts_master_path)
        for r in contact_rows:
            crn = (r.get("CustomerRecord") or "").strip()
            if not crn:
                continue
            if (r.get("IsPrimaryContact") or "").strip() != "1":
                continue
            existing = primary_contact_by_customer_record.get(crn)
            if not existing:
                primary_contact_by_customer_record[crn] = r
                continue
            try:
                new_num = int((r.get("RecordNumber") or "0").strip() or 0)
            except ValueError:
                new_num = 0
            try:
                old_num = int((existing.get("RecordNumber") or "0").strip() or 0)
            except ValueError:
                old_num = 0
            if new_num and (old_num == 0 or new_num < old_num):
                primary_contact_by_customer_record[crn] = r
    return primary_contact_by_customer_record


def _load_customers_by_record(customers_sync: str) -> Dict[str, Dict[str, str]]:
    _, sync_rows = read_csv(customers_sync)
    customers_by_record: Dict[str, Dict[str, str]] = {}
    for r in sync_rows:
        crn = (r.get("CustomerRecordNumber") or "").strip()
        if not crn:
            continue
        customers_by_record[crn] = r
    return customers_by_record


def _load_existing_contacts_by_parent_csv(odoo_contacts_path: str) -> Dict[int, List[Dict[str, str]]]:
    existing_by_parent: Dict[int, List[Dict[str, str]]] = {}
    _, rows = read_csv(odoo_contacts_path)
    for r in rows:
        parent_id = (r.get("ParentId") or "").strip()
        contact_id = (r.get("OdooId") or "").strip()
        if not parent_id or not contact_id:
            continue
        try:
            pid = int(parent_id)
        except ValueError:
            continue
        existing_by_parent.setdefault(pid, []).append({
            "id": contact_id,
            "name": (r.get("OdooName") or "").strip(),
            "email": (r.get("OdooEmail") or "").strip(),
            "phone": (r.get("OdooPhone") or "").strip(),
        })
    return existing_by_parent


def _norm_email(value: str) -> str:
    return (value or "").strip().lower()


def _norm_phone(value: str) -> str:
    return "".join([c for c in (value or "") if c.isdigit()])


def _match_contact(parent_rows: List[Dict[str, str]], name: str, email: str, phone: str):
    if not parent_rows:
        return None
    n_name = normalize_name(name)
    n_email = _norm_email(email)
    n_phone = _norm_phone(phone)
    for r in parent_rows:
        r_email = _norm_email(r.get("email", ""))
        r_phone = _norm_phone(r.get("phone", ""))
        r_name = normalize_name(r.get("name", ""))
        if n_email and r_email and n_email == r_email:
            return r.get("id")
        if n_name and r_name and n_name == r_name:
            if n_phone and r_phone and n_phone == r_phone:
                return r.get("id")
            if not n_email and not r_email:
                return r.get("id")
    return None


def build_contacts_sync(args: argparse.Namespace) -> int:
    customers_sync = args.customers_sync
    customers_master = args.customers_master
    contacts_sync = args.contacts_sync
    odoo_contacts = args.odoo_contacts

    if not os.path.exists(customers_sync):
        print(f"ERROR: customers sync file not found: {customers_sync}")
        return 2
    if not os.path.exists(customers_master):
        print(f"ERROR: customers master not found: {customers_master}")
        return 2
    if not os.path.exists(odoo_contacts):
        print(f"ERROR: odoo contacts file not found: {odoo_contacts}")
        print("Run: python sage_odoo_parity.py refresh_odoo")
        return 2

    customers_by_record = _load_customers_by_record(customers_sync)
    contacts_master_path = os.path.join(os.path.dirname(customers_master), "contacts.csv")
    primary_contact_by_customer_record = _load_primary_contacts(contacts_master_path)
    existing_by_parent = _load_existing_contacts_by_parent_csv(odoo_contacts)

    fieldnames = [
        "ContactId",
        "FirstName",
        "LastName",
        "CustomerRecordNumber",
        "CustomerId",
        "ContactExternalId",
        "OdooContactId",
        "OdooParentId",
        "LastLookupAt",
    ]
    rows_out: List[Dict[str, str]] = []
    now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for crn, customer in customers_by_record.items():
        primary = primary_contact_by_customer_record.get(crn)
        if not primary:
            continue
        odoo_id = (customer.get("OdooId") or "").strip()
        parent_id_int = None
        if odoo_id:
            try:
                parent_id_int = int(odoo_id)
            except ValueError:
                parent_id_int = None

        cid = (customer.get("CustomerID") or "").strip()
        first = (primary.get("FirstName") or "").strip()
        last = (primary.get("LastName") or "").strip()
        contact_name = " ".join([p for p in [first, last] if p]).strip()
        if not contact_name:
            contact_name = (primary.get("CompanyName") or "").strip()
        contact_rec = (primary.get("RecordNumber") or "").strip()
        contact_phone = (primary.get("Telephone1") or "").strip()
        contact_email = (primary.get("Email") or "").strip()

        raw_ext = f"{cid}_{contact_rec}" if contact_rec and cid else (f"{cid}_contact" if cid else "")
        ext_id = sanitize_external_id(raw_ext)

        match_id = None
        if parent_id_int:
            match_id = _match_contact(existing_by_parent.get(parent_id_int, []), contact_name, contact_email, contact_phone)

        rows_out.append({
            "ContactId": contact_rec,
            "FirstName": first,
            "LastName": last,
            "CustomerRecordNumber": crn,
            "CustomerId": cid,
            "ContactExternalId": ext_id,
            "OdooContactId": str(match_id or ""),
            "OdooParentId": odoo_id,
            "LastLookupAt": now_stamp,
        })

    write_csv(contacts_sync, fieldnames, rows_out)
    print(f"OK: contacts sync rows: {len(rows_out)} -> {contacts_sync}")
    return 0


def build_contacts_import(args: argparse.Namespace) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        load_workbook = None

    if load_workbook is None:
        print("ERROR: openpyxl not available for XLSX export")
        return 2

    contacts_sync = args.contacts_sync
    template_path = args.template_path

    if not os.path.exists(contacts_sync):
        print(f"ERROR: contacts sync file not found: {contacts_sync}")
        print("Run: python sage_odoo_parity.py build_contacts_sync")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2

    master_root = os.path.dirname(contacts_sync)
    odoo_imports = os.path.join(master_root, "odoo_imports")
    os.makedirs(odoo_imports, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d")
    contacts_out_xlsx = os.path.join(odoo_imports, f"{stamp}_customers_contacts_NEW.xlsx")

    sync_fields, sync_rows = read_csv(contacts_sync)
    # Load Sage contacts master to enrich email/phone/title/notes
    master_root = os.path.dirname(contacts_sync)
    contacts_master_path = os.path.join(os.path.dirname(master_root), "_master_sage", "contacts.csv")
    contacts_by_key = {}
    contacts_by_customer = {}
    if os.path.exists(contacts_master_path):
        _, contact_rows = read_csv(contacts_master_path)
        for r in contact_rows:
            crn = (r.get("CustomerRecord") or "").strip()
            rec = (r.get("RecordNumber") or "").strip()
            if crn and rec:
                contacts_by_key[(crn, rec)] = r
            if crn:
                contacts_by_customer.setdefault(crn, []).append(r)

    wb = load_workbook(template_path)
    ws = wb["Partners"] if "Partners" in wb.sheetnames else wb.active
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header:
            header_map[str(header).strip()] = col_idx

    def set_cell(col_name: str, value: str) -> None:
        col_idx = header_map.get(col_name)
        if col_idx:
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.delete_rows(2, ws.max_row)
    row_idx = 2
    emitted = 0
    new_rows: List[Dict[str, str]] = []
    for r in sync_rows:
        if (r.get("OdooContactId") or "").strip():
            continue
        parent_id = (r.get("OdooParentId") or "").strip()
        if not parent_id:
            continue
        new_rows.append(r)
        first = (r.get("FirstName") or "").strip()
        last = (r.get("LastName") or "").strip()
        contact_name = " ".join([p for p in [first, last] if p]).strip()
        if not contact_name:
            # Skip contacts with no name to avoid generic placeholders
            continue
        crn = (r.get("CustomerRecordNumber") or "").strip()
        contact_id = (r.get("ContactId") or "").strip()
        sage_row = None
        if crn and contact_id:
            sage_row = contacts_by_key.get((crn, contact_id))
        if not sage_row and crn:
            candidates = contacts_by_customer.get(crn, [])
            if len(candidates) == 1:
                sage_row = candidates[0]
        email = (sage_row.get("Email") or "").strip() if sage_row else ""
        phone = (sage_row.get("Telephone1") or "").strip() if sage_row else ""
        job = (sage_row.get("Title") or "").strip() if sage_row else ""
        notes = (sage_row.get("Notes") or "").strip() if sage_row else ""
        set_cell("External_ID", (r.get("ContactExternalId") or "").strip())
        set_cell("Reference", (r.get("ContactId") or "").strip())
        set_cell("Parent/Database ID", parent_id)
        set_cell("is_company", 0)
        set_cell("Name", contact_name)
        set_cell("Email", email)
        set_cell("Phone", phone)
        set_cell("Job Position", job)
        set_cell("Notes", notes)
        set_cell("Language", "English (US)")
        row_idx += 1
        emitted += 1

    wb.save(contacts_out_xlsx)
    if sync_fields:
        master_root = os.path.dirname(contacts_sync)
        new_csv = os.path.join(master_root, "customer_contacts_sync_NEW.csv")
        write_csv(new_csv, sync_fields, new_rows)
    print(f"OK: contacts new rows: {emitted} -> {contacts_out_xlsx}")
    return 0
