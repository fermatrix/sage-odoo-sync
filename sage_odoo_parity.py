import argparse
import csv
import os
from datetime import datetime
from typing import Dict, List, Tuple
from difflib import SequenceMatcher
from collections import defaultdict

from sync_customers import (
    DELIMITER,
    truthy,
    normalize_name,
    sanitize_external_id,
    parse_date,
    read_csv,
    write_csv,
    load_env_file,
    get_env_value,
)
from sync_contacts import build_contacts_sync, build_contacts_import
from sync_addresses import build_addresses_sync, build_delivery_import, build_delivery_update
from sync_billto import build_billto_sync, build_billto_import, build_billto_update
from sync_parity import OdooClient, export_countries
from parity_utils import load_state_parity
from sync_products import (
    build_product_sync,
    build_items_sync_new,
    build_products_sync_nobarcode_new,
    build_products_import,
    build_products_nobarcode_import,
)




def refresh_sage(args: argparse.Namespace) -> int:
    env = load_env_file(".env")
    min_customer_since = get_env_value(env, "CUSTOMER_SINCE_MIN")
    min_last_salesorder = get_env_value(
        env,
        "LAST_SALESORDER_MIN",
        get_env_value(env, "LAST_INVOICE_MIN"),
    )
    min_customer_since_date = parse_date(min_customer_since)
    min_last_salesorder_date = parse_date(min_last_salesorder)

    customers_master = args.customers_master
    items_master = args.items_master
    vendors_master = args.vendors_master

    if not os.path.exists(customers_master):
        print(f"ERROR: customers master not found: {customers_master}")
        return 2
    if not os.path.exists(items_master):
        print(f"ERROR: items master not found: {items_master}")
        return 2
    if not os.path.exists(vendors_master):
        print(f"ERROR: vendors master not found: {vendors_master}")
        return 2

    customer_out = args.customers_out
    items_out = args.items_out
    vendors_out = args.vendors_out

    existing_customers = {}
    if os.path.exists(customer_out):
        _, rows = read_csv(customer_out)
        for row in rows:
            key = row.get("CustomerRecordNumber", "").strip()
            if key:
                existing_customers[key] = row

    existing_items = {}
    if os.path.exists(items_out):
        _, rows = read_csv(items_out)
        for row in rows:
            key = row.get("ItemRecordNumber", "").strip()
            if key:
                existing_items[key] = row

    existing_vendors = {}
    if os.path.exists(vendors_out):
        _, rows = read_csv(vendors_out)
        for row in rows:
            key = row.get("VendorRecordNumber", "").strip()
            if key:
                existing_vendors[key] = row

    _, customer_rows = read_csv(customers_master)
    _, item_rows = read_csv(items_master)
    _, vendor_rows = read_csv(vendors_master)

    # Build latest Sales Order date by Sage CustomerID from downloaded headers.
    last_sales_order_by_customer_id: Dict[str, str] = {}
    sage_root = os.path.dirname(os.path.dirname(customers_master))
    if os.path.isdir(sage_root):
        for root_dir, _, files in os.walk(sage_root):
            for filename in files:
                if not filename.endswith("_sales_orders_headers.csv"):
                    continue
                so_headers_path = os.path.join(root_dir, filename)
                _, so_rows = read_csv(so_headers_path)
                for so in so_rows:
                    customer_id = (so.get("CustVendId") or "").strip()
                    if not customer_id:
                        continue
                    trx_date = (so.get("TransactionDate") or "").strip()
                    if not trx_date:
                        continue
                    prev = last_sales_order_by_customer_id.get(customer_id, "")
                    if not prev or trx_date > prev:
                        last_sales_order_by_customer_id[customer_id] = trx_date

    address_master_path = os.path.join(os.path.dirname(customers_master), "address.csv")
    address_by_customer = {}
    address_by_vendor = {}
    if os.path.exists(address_master_path):
        _, address_rows = read_csv(address_master_path)
        for addr in address_rows:
            if (addr.get("AddressTypeNumber") or "").strip() != "0":
                continue
            customer_record = (addr.get("CustomerRecordNumber") or "").strip()
            if not customer_record or customer_record == "0":
                continue
            existing_addr = address_by_customer.get(customer_record)
            current_num = int(addr.get("AddressRecordNumber") or "999999999")
            existing_num = int((existing_addr or {}).get("AddressRecordNumber") or "999999999")
            if existing_addr is None or current_num < existing_num:
                address_by_customer[customer_record] = addr
            vendor_record = (addr.get("VendorRecordNumber") or "").strip()
            if vendor_record and vendor_record != "0":
                existing_vendor_addr = address_by_vendor.get(vendor_record)
                existing_vendor_num = int((existing_vendor_addr or {}).get("AddressRecordNumber") or "999999999")
                if existing_vendor_addr is None or current_num < existing_vendor_num:
                    address_by_vendor[vendor_record] = addr

    customer_fields = [
        "CustomerRecordNumber",
        "CustomerIsInactive",
        "CustomerSince",
        "LastInvoiceDate",
        "LastSalesOrderDate",
        "PriceLevel",
        "CustomerID",
        "Customer_Bill_Name",
        "Phone",
        "Email",
        "Street",
        "Street2",
        "City",
        "Zip",
        "State",
        "Country",
        "OdooId",
        "OdooName",
        "OdooPricelistId",
        "OdooPricelist",
        "ExpectedOdooPricelistId",
        "ExpectedOdooPricelist",
        "Exclude",
        "CustomerSyncStatus",
        "CustomerMismatchFields",
        "LastLookupAt",
    ]
    item_fields = [
        "ItemRecordNumber",
        "ItemID",
        "ItemDescription",
        "ItemDescriptionForSale",
        "Barcode",
        "ItemIsInactive",
        "OdooVariantId",
        "OdooVariantExternalId",
        "OdooTemplateId",
        "OdooTemplateExternalId",
        "OdooName",
        "OdooColor",
        "Exclude",
        "LastLookupAt",
    ]
    vendor_fields = [
        "VendorRecordNumber",
        "VendorID",
        "Name",
        "Phone",
        "Email",
        "IsInactive",
        "Street",
        "Street2",
        "City",
        "State",
        "Zip",
        "Country",
        "MailToCity",
        "MailToZip",
        "MailToCountry",
        "OdooId",
        "OdooExternalId",
        "OdooName",
        "OdooRef",
        "VendorSyncStatus",
        "VendorMismatchFields",
        "LastLookupAt",
    ]

    out_customers: List[Dict[str, str]] = []
    for row in customer_rows:
        key = (row.get("CustomerRecordNumber") or "").strip()
        if not key:
            continue
        existing = existing_customers.get(key, {})
        addr = address_by_customer.get(key, {})
        out_customers.append({
            "CustomerRecordNumber": key,
            "CustomerIsInactive": (row.get("CustomerIsInactive") or "").strip(),
            "CustomerSince": (row.get("CustomerSince") or "").strip(),
            "LastInvoiceDate": (row.get("LastInvoiceDate") or "").strip(),
            "LastSalesOrderDate": last_sales_order_by_customer_id.get((row.get("CustomerID") or "").strip(), ""),
            "PriceLevel": (row.get("PriceLevel") or "").strip(),
            "CustomerID": (row.get("CustomerID") or "").strip(),
            "Customer_Bill_Name": (row.get("Customer_Bill_Name") or "").strip(),
            "Phone": (row.get("Phone_Number") or row.get("PhoneNumber2") or "").strip(),
            "Email": (row.get("eMail_Address") or "").strip(),
            "Street": (addr.get("AddressLine1") or "").strip(),
            "Street2": (addr.get("AddressLine2") or "").strip(),
            "City": (addr.get("City") or "").strip(),
            "Zip": (addr.get("Zip") or "").strip(),
            "State": (addr.get("State") or "").strip(),
            "Country": (addr.get("Country") or "").strip(),
            "OdooId": (existing.get("OdooId") or "").strip(),
            "OdooName": (existing.get("OdooName") or "").strip(),
            "OdooPricelistId": (existing.get("OdooPricelistId") or "").strip(),
            "OdooPricelist": (existing.get("OdooPricelist") or "").strip(),
            "ExpectedOdooPricelistId": (existing.get("ExpectedOdooPricelistId") or "").strip(),
            "ExpectedOdooPricelist": (existing.get("ExpectedOdooPricelist") or "").strip(),
            "Exclude": (existing.get("Exclude") or "").strip(),
            "CustomerSyncStatus": (existing.get("CustomerSyncStatus") or "").strip(),
            "CustomerMismatchFields": (existing.get("CustomerMismatchFields") or "").strip(),
            "LastLookupAt": (existing.get("LastLookupAt") or "").strip(),
        })

    out_items: List[Dict[str, str]] = []
    for row in item_rows:
        key = (row.get("ItemRecordNumber") or "").strip()
        if not key:
            continue
        existing = existing_items.get(key, {})
        out_items.append({
            "ItemRecordNumber": key,
            "ItemID": (row.get("ItemID") or "").strip(),
            "ItemDescription": (row.get("ItemDescription") or "").strip(),
            "ItemDescriptionForSale": (row.get("SalesDescription") or "").strip(),
            "Barcode": (row.get("UPC_SKU") or "").strip(),
            "ItemIsInactive": (row.get("ItemIsInactive") or "").strip(),
            "OdooColor": (existing.get("OdooColor") or "").strip(),
            "OdooVariantId": (existing.get("OdooVariantId") or "").strip(),
            "OdooVariantExternalId": (existing.get("OdooVariantExternalId") or "").strip(),
            "OdooTemplateId": (existing.get("OdooTemplateId") or "").strip(),
            "OdooTemplateExternalId": (existing.get("OdooTemplateExternalId") or "").strip(),
            "OdooName": (existing.get("OdooName") or "").strip(),
            "Exclude": (existing.get("Exclude") or "").strip(),
            "LastLookupAt": (existing.get("LastLookupAt") or "").strip(),
        })

    out_vendors: List[Dict[str, str]] = []
    for row in vendor_rows:
        key = (row.get("VendorRecordNumber") or "").strip()
        if not key:
            continue
        existing = existing_vendors.get(key, {})
        addr = address_by_vendor.get(key, {})
        out_vendors.append({
            "VendorRecordNumber": key,
            "VendorID": (row.get("VendorID") or "").strip(),
            "Name": (row.get("Name") or "").strip(),
            "Phone": (row.get("PhoneNumber") or row.get("PhoneNumber2") or "").strip(),
            "Email": (row.get("Email") or "").strip(),
            "IsInactive": (row.get("IsInactive") or "").strip(),
            "Street": (addr.get("AddressLine1") or "").strip(),
            "Street2": (addr.get("AddressLine2") or "").strip(),
            "City": (addr.get("City") or "").strip(),
            "State": (addr.get("State") or "").strip(),
            "Zip": (addr.get("Zip") or "").strip(),
            "Country": (addr.get("Country") or "").strip(),
            "MailToCity": (existing.get("MailToCity") or "").strip(),
            "MailToZip": (existing.get("MailToZip") or "").strip(),
            "MailToCountry": (existing.get("MailToCountry") or "").strip(),
            "OdooId": (existing.get("OdooId") or "").strip(),
            "OdooExternalId": (existing.get("OdooExternalId") or "").strip(),
            "OdooName": (existing.get("OdooName") or "").strip(),
            "OdooRef": (existing.get("OdooRef") or "").strip(),
            "VendorSyncStatus": (existing.get("VendorSyncStatus") or "").strip(),
            "VendorMismatchFields": (existing.get("VendorMismatchFields") or "").strip(),
            "LastLookupAt": (existing.get("LastLookupAt") or "").strip(),
        })

    out_customers.sort(key=lambda r: int(r["CustomerRecordNumber"]))
    out_items.sort(key=lambda r: int(r["ItemRecordNumber"]))
    out_vendors.sort(key=lambda r: int(r["VendorRecordNumber"]))
    inactive_items = sum(1 for r in out_items if (r.get("ItemIsInactive") or "").strip() == "1")

    write_csv(customer_out, customer_fields, out_customers)
    write_csv(items_out, item_fields, out_items)
    write_csv(vendors_out, vendor_fields, out_vendors)

    # Build customers_NEW files
    try:
        from openpyxl import load_workbook
    except Exception:
        load_workbook = None

    master_root = os.path.dirname(customer_out)
    odoo_imports = os.path.join(master_root, "odoo_imports")
    os.makedirs(odoo_imports, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d")
    customers_new_xlsx = os.path.join(odoo_imports, f"{stamp}_customers_NEW.xlsx")
    customers_new_min_xlsx = os.path.join(master_root, "customers_NEW.xlsx")
    template_path = os.path.join(master_root, "odoo_templates", "NEW_customers.xlsx")
    if not os.path.exists(template_path):
        template_path = os.path.join(master_root, "odoo_templates", "NEW_customers.xlsx")

    # Filter: active in Sage + no OdooId + optional date thresholds.
    # One-off business override:
    # - CustomerRecordNumber 5554 (CustomerID 98052OCEAN) must be included even
    #   when LastSalesOrderDate/LastInvoiceDate are blank in customers.csv.
    forced_new_customer_records = {"5554"}
    forced_new_customer_ids = {"98052OCEAN"}
    new_customers = [
        c for c in out_customers
        if (c.get("CustomerIsInactive") or "").strip() != "1"
        and not c.get("OdooId")
        and (
            (c.get("CustomerRecordNumber") or "").strip() in forced_new_customer_records
            or (c.get("CustomerID") or "").strip() in forced_new_customer_ids
            or (
                (
                    not min_customer_since_date
                    or (parse_date(c.get("CustomerSince")) or datetime.min.date()) >= min_customer_since_date
                )
                and (
                    not min_last_salesorder_date
                    or (
                        parse_date(c.get("LastSalesOrderDate"))
                        or parse_date(c.get("LastInvoiceDate"))
                        or datetime.min.date()
                    ) >= min_last_salesorder_date
                )
            )
        )
    ]

    if load_workbook and os.path.exists(template_path):
        wb = load_workbook(template_path)
        ws = wb["Partners"] if "Partners" in wb.sheetnames else wb.active
        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header:
                header_map[str(header).strip()] = col_idx
        def set_cell(col_name: str, value: str) -> None:
            col_idx = header_map.get(col_name)
            if col_idx:
                ws.cell(row=row_idx, column=col_idx, value=value)
        ws.delete_rows(2, ws.max_row)
        row_idx = 2
        # Load Sage master for address details
        sage_master_path = customers_master
        sage_by_id = {}
        if os.path.exists(sage_master_path):
            _, sage_rows = read_csv(sage_master_path)
            for r in sage_rows:
                cid = (r.get("CustomerID") or "").strip()
                if cid:
                    sage_by_id[cid] = r
        # Load Address master (preferred over Cardholder_* when available)
        address_master_path = os.path.join(os.path.dirname(customers_master), "address.csv")
        address_by_customer_record = {}
        if os.path.exists(address_master_path):
            _, address_rows = read_csv(address_master_path)
            for r in address_rows:
                crn = (r.get("CustomerRecordNumber") or "").strip()
                if not crn:
                    continue
                addr_type = (r.get("AddressTypeNumber") or "").strip()
                if addr_type != "0":
                    continue
                existing = address_by_customer_record.get(crn)
                if not existing:
                    address_by_customer_record[crn] = r
                    continue
                # If multiple type-0 addresses exist, keep the lowest AddressRecordNumber
                try:
                    new_num = int((r.get("AddressRecordNumber") or "0").strip() or 0)
                except ValueError:
                    new_num = 0
                try:
                    old_num = int((existing.get("AddressRecordNumber") or "0").strip() or 0)
                except ValueError:
                    old_num = 0
                if new_num and (old_num == 0 or new_num < old_num):
                    address_by_customer_record[crn] = r
        # Load Contacts master to attach primary contact (child) rows
        contacts_master_path = os.path.join(os.path.dirname(customers_master), "contacts.csv")
        primary_contact_by_customer_record = {}
        if os.path.exists(contacts_master_path):
            _, contact_rows = read_csv(contacts_master_path)
            for r in contact_rows:
                crn = (r.get("CustomerRecord") or "").strip()
                if not crn:
                    continue
                if (r.get("IsPrimaryContact") or "").strip() != "1":
                    continue
                existing = primary_contact_by_customer_record.get(crn)
                if not existing:
                    primary_contact_by_customer_record[crn] = r
                    continue
                try:
                    new_num = int((r.get("RecordNumber") or "0").strip() or 0)
                except ValueError:
                    new_num = 0
                try:
                    old_num = int((existing.get("RecordNumber") or "0").strip() or 0)
                except ValueError:
                    old_num = 0
                if new_num and (old_num == 0 or new_num < old_num):
                    primary_contact_by_customer_record[crn] = r

        # Optional country parity for Odoo import (applied only to customers_NEW)
        parity_path = os.path.join(master_root, "_parity_country.csv")
        country_parity = {}
        country_name_to_code = {}
        if os.path.exists(parity_path):
            with open(parity_path, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter=DELIMITER)
                for row in reader:
                    raw = (row.get("sage_country_raw") or "").strip()
                    if not raw:
                        continue
                    odoo_code = (row.get("odoo_country_code") or "").strip()
                    if odoo_code:
                        country_parity[raw] = odoo_code
                    odoo_name = (row.get("odoo_country_name") or "").strip()
                    if odoo_name and odoo_code:
                        country_name_to_code[odoo_name] = odoo_code

        # Fallback: load Odoo country list for name->code mapping
        master_base = os.path.dirname(master_root)
        countries_export = os.path.join(master_base, "_master_odoo", "countries_odoo.csv")
        if os.path.exists(countries_export):
            with open(countries_export, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter=DELIMITER)
                for row in reader:
                    name = (row.get("OdooName") or "").strip()
                    code = (row.get("OdooCode") or "").strip()
                    if name and code:
                        country_name_to_code.setdefault(name, code)

        # Optional state parity (state code -> full name + implied country)
        state_parity_path = os.path.join(master_root, "_parity_state.csv")
        state_parity = {}
        if os.path.exists(state_parity_path):
            with open(state_parity_path, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f, delimiter=DELIMITER)
                for row in reader:
                    raw = (row.get("sage_state_raw") or "").strip()
                    if not raw:
                        continue
                    state_name = (row.get("odoo_state_name") or "").strip()
                    country_name = (row.get("odoo_country_name") or "").strip()
                    state_parity[raw] = {
                        "state_name": state_name,
                        "country_name": country_name,
                    }

        for c in new_customers:
            cid = c.get("CustomerID", "")
            name = c.get("Customer_Bill_Name", "")
            sage = sage_by_id.get(cid, {})
            crn = (c.get("CustomerRecordNumber") or "").strip()
            addr = address_by_customer_record.get(crn, {})
            def pick_addr(addr_key: str) -> str:
                return (addr.get(addr_key) or "").strip()
            ext_id = sanitize_external_id(cid)
            if not ext_id:
                ext_id = sanitize_external_id(crn)
            set_cell("External_ID", ext_id)
            set_cell("name", name)
            set_cell("is_company", 1)
            set_cell("company_name", name)
            raw_country = pick_addr("Country")
            mapped_country = country_parity.get(raw_country, "")
            if not mapped_country and raw_country:
                mapped_country = country_name_to_code.get(raw_country, "")
            if not mapped_country:
                mapped_country = raw_country
            raw_state = pick_addr("State")
            state_info = state_parity.get(raw_state, {})
            mapped_state = state_info.get("state_name") or raw_state
            if mapped_state and state_info.get("country_name"):
                country_code = country_name_to_code.get(state_info.get("country_name", ""), "")
                if country_code:
                    mapped_state = f"{mapped_state} ({country_code})"
            set_cell("country_id", mapped_country)
            # If country missing, infer from state (US/Canada/known from Odoo)
            if not mapped_country:
                country_name = (state_info.get("country_name") or "").strip()
                if country_name:
                    # Map country name to Odoo ISO code; do not fall back to Sage names
                    mapped_country = country_name_to_code.get(country_name, "")
                    if mapped_country:
                        set_cell("country_id", mapped_country)
            set_cell("state_id", mapped_state)
            set_cell("zip", pick_addr("Zip"))
            set_cell("city", pick_addr("City"))
            set_cell("street", pick_addr("AddressLine1"))
            set_cell("street2", pick_addr("AddressLine2"))
            set_cell("phone", (sage.get("Phone_Number") or "").strip())
            set_cell("email", (sage.get("eMail_Address") or "").strip())
            credit_msg = (sage.get("CreditStatusMsg") or "").strip()
            if credit_msg and credit_msg != "You have requested to be notified when a transaction is created for this customer.":
                set_cell("Notes", credit_msg)
            set_cell("Reference", cid)
            # Optional pricelist fields (template-dependent)
            set_cell("Pricelist", c.get("ExpectedOdooPricelist", ""))
            set_cell("Language", "English (US)")
            row_idx += 1
        wb.save(customers_new_xlsx)

    # Minimal XLSX in _master
    minimal_fields = [
        "CustomerRecordNumber",
        "CustomerSince",
        "LastInvoiceDate",
        "LastSalesOrderDate",
        "CustomerID",
        "Customer_Bill_Name",
    ]
    if load_workbook:
        from openpyxl import Workbook
        wb_min = Workbook()
        ws_min = wb_min.active
        ws_min.title = "customers_NEW"
        for col_idx, name in enumerate(minimal_fields, start=1):
            ws_min.cell(row=1, column=col_idx, value=name)
        row_idx = 2
        for c in new_customers:
            for col_idx, name in enumerate(minimal_fields, start=1):
                ws_min.cell(row=row_idx, column=col_idx, value=c.get(name, ""))
            row_idx += 1
        wb_min.save(customers_new_min_xlsx)

    print(f"OK: customers sync rows: {len(out_customers)} -> {customer_out}")
    print(f"OK: products sync rows: {len(out_items)} -> {items_out} (inactive from Sage: {inactive_items})")
    print(f"OK: vendors sync rows: {len(out_vendors)} -> {vendors_out}")
    return 0


def refresh_odoo(args: argparse.Namespace) -> int:
    env = load_env_file(args.env_file)
    url = get_env_value(env, "ODOO_STUDIOOPTYX_URL")
    db = get_env_value(env, "ODOO_STUDIOOPTYX_DB")
    user = get_env_value(env, "ODOO_STUDIOOPTYX_USER")
    apikey = get_env_value(env, "ODOO_STUDIOOPTYX_APIKEY")

    if not (url and db and user and apikey):
        print("ERROR: missing Odoo credentials (URL/DB/USER/APIKEY)")
        return 2

    client = OdooClient(url, db, user, apikey)

    customers_out = args.customers_out
    items_out = args.items_out
    os.makedirs(os.path.dirname(customers_out), exist_ok=True)
    os.makedirs(os.path.dirname(items_out), exist_ok=True)
    odoo_root = os.path.dirname(items_out)
    vendors_out = args.vendors_out
    if not os.path.isabs(vendors_out):
        vendors_out = os.path.normpath(vendors_out)
    os.makedirs(os.path.dirname(vendors_out), exist_ok=True)

    batch = args.batch_size
    only_raw = (getattr(args, "only", "") or "").strip()
    only_pos = (getattr(args, "only_pos", "") or "").strip()
    if only_pos:
        only_raw = only_raw or only_pos
    only_targets = {t.strip().lower() for t in only_raw.split(",") if t.strip()}

    chart_out = args.chart_out
    if not os.path.isabs(chart_out):
        chart_out = os.path.normpath(chart_out)
    os.makedirs(os.path.dirname(chart_out), exist_ok=True)

    customer_fields = [
        "OdooId",
        "OdooExternalId",
        "OdooName",
        "OdooRef",
        "Active",
        "ParentId",
        "OdooEmail",
        "OdooPhone",
        "Street",
        "Street2",
        "City",
        "Zip",
        "State",
        "Country",
        "OdooSalespersonId",
        "OdooSalesperson",
        "OdooPricelistId",
        "OdooPricelist",
    ]

    def _load_external_ids(model: str) -> Dict[str, str]:
        external_by_res_id: Dict[str, str] = {}
        offset = 0
        while True:
            rows = client.search_read(
                "ir.model.data",
                [["model", "=", model]],
                ["module", "name", "res_id"],
                limit=batch,
                offset=offset,
            )
            if not rows:
                break
            for row in rows:
                res_id = str(row.get("res_id") or "").strip()
                module = (row.get("module") or "").strip()
                name = (row.get("name") or "").strip()
                if res_id and module and name and res_id not in external_by_res_id:
                    external_by_res_id[res_id] = f"{module}.{name}"
            offset += len(rows)
        return external_by_res_id

    def _export_items_only() -> None:
        item_fields = [
            "OdooVariantId",
            "OdooVariantExternalId",
            "OdooName",
            "OdooVariantName",
            "OdooItemCode",
            "OdooColor",
            "Active",
            "OdooTemplateId",
            "OdooTemplateExternalId",
            "OdooTemplateListPrice",
        ]
        with open(items_out, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=item_fields, delimiter=DELIMITER)
            writer.writeheader()
            variant_external = _load_external_ids("product.product")
            offset = 0
            while True:
                rows = client.search_read(
                    "product.product",
                    [["active", "=", True]],
                    ["id", "name", "display_name", "default_code", "active", "product_template_attribute_value_ids", "product_tmpl_id"],
                    limit=batch,
                    offset=offset,
                )
                if not rows:
                    break
                tmpl_ids = set()
                attr_ids = set()
                for r in rows:
                    for aid in r.get("product_template_attribute_value_ids") or []:
                        attr_ids.add(aid)
                    tmpl = r.get("product_tmpl_id") or []
                    if isinstance(tmpl, list) and tmpl:
                        tmpl_ids.add(tmpl[0])
                attr_map = {}
                if attr_ids:
                    ids = list(attr_ids)
                    chunk_size = 1000
                    for i in range(0, len(ids), chunk_size):
                        chunk = ids[i:i + chunk_size]
                        vals = client.models.execute_kw(
                            client.db,
                            client.uid,
                            client.apikey,
                            "product.template.attribute.value",
                            "read",
                            [chunk],
                            {"fields": ["name", "attribute_id"]},
                        )
                        for v in vals:
                            attr = v.get("attribute_id")
                            attr_name = attr[1] if isinstance(attr, list) and len(attr) > 1 else ""
                            attr_map[v.get("id")] = {
                                "name": v.get("name", ""),
                                "attribute": attr_name,
                            }
                tmpl_external = {}
                tmpl_prices = {}
                if tmpl_ids:
                    ids = list(tmpl_ids)
                    chunk_size = 1000
                    for i in range(0, len(ids), chunk_size):
                        chunk = ids[i:i + chunk_size]
                        tmpl_rows = client.models.execute_kw(
                            client.db,
                            client.uid,
                            client.apikey,
                            "product.template",
                            "read",
                            [chunk],
                            {"fields": ["list_price"]},
                        )
                        for t in tmpl_rows:
                            if t.get("id"):
                                tmpl_prices[t.get("id")] = t.get("list_price", "")
                        data_rows = client.models.execute_kw(
                            client.db,
                            client.uid,
                            client.apikey,
                            "ir.model.data",
                            "search_read",
                            [[("model", "=", "product.template"), ("res_id", "in", chunk)]],
                            {"fields": ["module", "name", "res_id"]},
                        )
                        for d in data_rows:
                            res_id = d.get("res_id")
                            module = d.get("module") or ""
                            name = d.get("name") or ""
                            if res_id and module and name:
                                tmpl_external[res_id] = f"{module}.{name}"
                for r in rows:
                    color_values = []
                    for aid in r.get("product_template_attribute_value_ids") or []:
                        info = attr_map.get(aid)
                        if info and (info.get("attribute") or "").lower() == "color":
                            color_values.append(info.get("name", ""))
                    tmpl = r.get("product_tmpl_id") or []
                    tmpl_id = tmpl[0] if isinstance(tmpl, list) and tmpl else ""
                    writer.writerow({
                        "OdooVariantId": r.get("id", ""),
                        "OdooVariantExternalId": variant_external.get(str(r.get("id", "")), ""),
                        "OdooName": r.get("name", "") or "",
                        "OdooVariantName": r.get("display_name", "") or "",
                        "OdooItemCode": r.get("default_code", "") or "",
                        "OdooColor": " / ".join([c for c in color_values if c]),
                        "Active": r.get("active", ""),
                        "OdooTemplateId": tmpl_id,
                        "OdooTemplateExternalId": tmpl_external.get(tmpl_id, ""),
                        "OdooTemplateListPrice": tmpl_prices.get(tmpl_id, ""),
                    })
                offset += len(rows)

    if only_targets and not only_targets.issubset({"items", "items_odoo"}):
        print("ERROR: --only currently supports: items_odoo")
        return 2
    if only_targets:
        _export_items_only()
        print(f"OK: odoo items exported -> {items_out}")
        return 0

    partner_external_ids = _load_external_ids("res.partner")

    with open(customers_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=customer_fields, delimiter=DELIMITER)
        writer.writeheader()
        offset = 0
        while True:
            rows = client.search_read(
                "res.partner",
                [],
                [
                    "id",
                    "name",
                    "ref",
                    "active",
                    "parent_id",
                    "email",
                    "phone",
                    "street",
                    "street2",
                    "city",
                    "zip",
                    "state_id",
                    "country_id",
                    "user_id",
                    "property_product_pricelist",
                ],
                limit=batch,
                offset=offset,
            )
            if not rows:
                break
            for r in rows:
                parent = r.get("parent_id") or []
                parent_id = parent[0] if isinstance(parent, list) and parent else ""
                salesperson = r.get("user_id") or []
                pricelist = r.get("property_product_pricelist") or []
                state = r.get("state_id") or []
                country = r.get("country_id") or []
                writer.writerow({
                    "OdooId": r.get("id", ""),
                    "OdooExternalId": partner_external_ids.get(str(r.get("id", "")), ""),
                    "OdooName": r.get("name", "") or "",
                    "OdooRef": r.get("ref", "") or "",
                    "Active": r.get("active", ""),
                    "ParentId": parent_id,
                    "OdooEmail": r.get("email", "") or "",
                    "OdooPhone": r.get("phone", "") or "",
                    "Street": r.get("street", "") or "",
                    "Street2": r.get("street2", "") or "",
                    "City": r.get("city", "") or "",
                    "Zip": r.get("zip", "") or "",
                    "State": state[1] if isinstance(state, list) and len(state) > 1 else "",
                    "Country": country[1] if isinstance(country, list) and len(country) > 1 else "",
                    "OdooSalespersonId": salesperson[0] if isinstance(salesperson, list) and salesperson else "",
                    "OdooSalesperson": salesperson[1] if isinstance(salesperson, list) and len(salesperson) > 1 else "",
                    "OdooPricelistId": pricelist[0] if isinstance(pricelist, list) and pricelist else "",
                    "OdooPricelist": pricelist[1] if isinstance(pricelist, list) and len(pricelist) > 1 else "",
                })
            offset += len(rows)

    vendor_fields = [
        "OdooId",
        "OdooExternalId",
        "OdooName",
        "OdooRef",
        "Active",
        "Phone",
        "Email",
        "Street",
        "Street2",
        "City",
        "Zip",
        "State",
        "Country",
        "SupplierRank",
    ]
    with open(vendors_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=vendor_fields, delimiter=DELIMITER)
        writer.writeheader()
        offset = 0
        while True:
            rows = client.search_read(
                "res.partner",
                [["supplier_rank", ">", 0], ["parent_id", "=", False]],
                [
                    "id",
                    "name",
                    "ref",
                    "active",
                    "phone",
                    "email",
                    "street",
                    "street2",
                    "city",
                    "zip",
                    "state_id",
                    "country_id",
                    "supplier_rank",
                ],
                limit=batch,
                offset=offset,
            )
            if not rows:
                break
            for r in rows:
                state = r.get("state_id") or []
                country = r.get("country_id") or []
                writer.writerow({
                    "OdooId": r.get("id", ""),
                    "OdooExternalId": partner_external_ids.get(str(r.get("id", "")), ""),
                    "OdooName": r.get("name", "") or "",
                    "OdooRef": r.get("ref", "") or "",
                    "Active": r.get("active", ""),
                    "Phone": r.get("phone", "") or "",
                    "Email": r.get("email", "") or "",
                    "Street": r.get("street", "") or "",
                    "Street2": r.get("street2", "") or "",
                    "City": r.get("city", "") or "",
                    "Zip": r.get("zip", "") or "",
                    "State": state[1] if isinstance(state, list) and len(state) > 1 else "",
                    "Country": country[1] if isinstance(country, list) and len(country) > 1 else "",
                    "SupplierRank": r.get("supplier_rank", "") or "",
                })
            offset += len(rows)
    print(f"OK: odoo vendors exported -> {vendors_out}")

    # Export Odoo contacts (res.partner with parent_id) for contact matching
    contacts_out = os.path.join(os.path.dirname(customers_out), "customers_contacts.csv")
    contact_fields = ["OdooId", "ParentId", "OdooName", "OdooEmail", "OdooPhone", "Active"]
    with open(contacts_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=contact_fields, delimiter=DELIMITER)
        writer.writeheader()
        offset = 0
        while True:
            rows = client.search_read(
                "res.partner",
                [["parent_id", "!=", False]],
                ["id", "parent_id", "name", "email", "phone", "active"],
                limit=batch,
                offset=offset,
            )
            if not rows:
                break
            for r in rows:
                parent = r.get("parent_id") or []
                parent_id = parent[0] if isinstance(parent, list) and parent else ""
                writer.writerow({
                    "OdooId": r.get("id", ""),
                    "ParentId": parent_id,
                    "OdooName": r.get("name", "") or "",
                    "OdooEmail": r.get("email", "") or "",
                    "OdooPhone": r.get("phone", "") or "",
                    "Active": r.get("active", ""),
                })
            offset += len(rows)

    # Export child partners excluding delivery (contacts + other non-delivery child records)
    children_out = os.path.join(os.path.dirname(customers_out), "customers_child_partners.csv")
    children_fields = [
        "OdooId",
        "OdooExternalId",
        "ParentId",
        "ParentName",
        "OdooName",
        "OdooRef",
        "Type",
        "Street",
        "Street2",
        "City",
        "Zip",
        "State",
        "Country",
        "OdooEmail",
        "OdooPhone",
        "Active",
    ]
    with open(children_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=children_fields, delimiter=DELIMITER)
        writer.writeheader()
        offset = 0
        while True:
            rows = client.search_read(
                "res.partner",
                [["parent_id", "!=", False], ["type", "!=", "delivery"]],
                [
                    "id",
                    "parent_id",
                    "name",
                    "ref",
                    "type",
                    "street",
                    "street2",
                    "city",
                    "zip",
                    "state_id",
                    "country_id",
                    "email",
                    "phone",
                    "active",
                ],
                limit=batch,
                offset=offset,
            )
            if not rows:
                break
            for r in rows:
                parent = r.get("parent_id") or []
                parent_id = parent[0] if isinstance(parent, list) and parent else ""
                parent_name = parent[1] if isinstance(parent, list) and len(parent) > 1 else ""
                state = r.get("state_id") or []
                country = r.get("country_id") or []
                writer.writerow({
                    "OdooId": r.get("id", ""),
                    "OdooExternalId": partner_external_ids.get(str(r.get("id", "")), ""),
                    "ParentId": parent_id,
                    "ParentName": parent_name,
                    "OdooName": r.get("name", "") or "",
                    "OdooRef": r.get("ref", "") or "",
                    "Type": r.get("type", "") or "",
                    "Street": r.get("street", "") or "",
                    "Street2": r.get("street2", "") or "",
                    "City": r.get("city", "") or "",
                    "Zip": r.get("zip", "") or "",
                    "State": state[1] if isinstance(state, list) and len(state) > 1 else "",
                    "Country": country[1] if isinstance(country, list) and len(country) > 1 else "",
                    "OdooEmail": r.get("email", "") or "",
                    "OdooPhone": r.get("phone", "") or "",
                    "Active": r.get("active", ""),
                })
            offset += len(rows)

    # Export all child partners (no type filter) for inspection
    children_all_out = os.path.join(os.path.dirname(customers_out), "customers_child_partners_all.csv")
    with open(children_all_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=children_fields, delimiter=DELIMITER)
        writer.writeheader()
        offset = 0
        while True:
            rows = client.search_read(
                "res.partner",
                [["parent_id", "!=", False]],
                [
                    "id",
                    "parent_id",
                    "name",
                    "ref",
                    "type",
                    "street",
                    "street2",
                    "city",
                    "zip",
                    "state_id",
                    "country_id",
                    "email",
                    "phone",
                    "active",
                ],
                limit=batch,
                offset=offset,
            )
            if not rows:
                break
            for r in rows:
                parent = r.get("parent_id") or []
                parent_id = parent[0] if isinstance(parent, list) and parent else ""
                parent_name = parent[1] if isinstance(parent, list) and len(parent) > 1 else ""
                state = r.get("state_id") or []
                country = r.get("country_id") or []
                writer.writerow({
                    "OdooId": r.get("id", ""),
                    "OdooExternalId": partner_external_ids.get(str(r.get("id", "")), ""),
                    "ParentId": parent_id,
                    "ParentName": parent_name,
                    "OdooName": r.get("name", "") or "",
                    "OdooRef": r.get("ref", "") or "",
                    "Type": r.get("type", "") or "",
                    "Street": r.get("street", "") or "",
                    "Street2": r.get("street2", "") or "",
                    "City": r.get("city", "") or "",
                    "Zip": r.get("zip", "") or "",
                    "State": state[1] if isinstance(state, list) and len(state) > 1 else "",
                    "Country": country[1] if isinstance(country, list) and len(country) > 1 else "",
                    "OdooEmail": r.get("email", "") or "",
                    "OdooPhone": r.get("phone", "") or "",
                    "Active": r.get("active", ""),
                })
            offset += len(rows)

    # Export Odoo delivery addresses (res.partner type=delivery)
    delivery_out = os.path.join(os.path.dirname(customers_out), "customers_delivery_addresses.csv")
    delivery_fields = [
        "OdooId",
        "OdooExternalId",
        "OdooRef",
        "ParentId",
        "ParentName",
        "OdooName",
        "Type",
        "Street",
        "Street2",
        "City",
        "Zip",
        "State",
        "Country",
        "OdooEmail",
        "OdooPhone",
        "Active",
    ]
    with open(delivery_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=delivery_fields, delimiter=DELIMITER)
        writer.writeheader()
        offset = 0
        while True:
            rows = client.search_read(
                "res.partner",
                [["parent_id", "!=", False], ["type", "=", "delivery"]],
                [
                    "id",
                    "parent_id",
                    "name",
                    "ref",
                    "type",
                    "street",
                    "street2",
                    "city",
                    "zip",
                    "state_id",
                    "country_id",
                    "email",
                    "phone",
                    "active",
                ],
                limit=batch,
                offset=offset,
            )
            if not rows:
                break
            for r in rows:
                parent = r.get("parent_id") or []
                parent_id = parent[0] if isinstance(parent, list) and parent else ""
                parent_name = parent[1] if isinstance(parent, list) and len(parent) > 1 else ""
                state = r.get("state_id") or []
                country = r.get("country_id") or []
                writer.writerow({
                    "OdooId": r.get("id", ""),
                    "OdooExternalId": partner_external_ids.get(str(r.get("id", "")), ""),
                    "OdooRef": r.get("ref", "") or "",
                    "ParentId": parent_id,
                    "ParentName": parent_name,
                    "OdooName": r.get("name", "") or "",
                    "Type": r.get("type", "") or "",
                    "Street": r.get("street", "") or "",
                    "Street2": r.get("street2", "") or "",
                    "City": r.get("city", "") or "",
                    "Zip": r.get("zip", "") or "",
                    "State": state[1] if isinstance(state, list) and len(state) > 1 else "",
                    "Country": country[1] if isinstance(country, list) and len(country) > 1 else "",
                    "OdooEmail": r.get("email", "") or "",
                    "OdooPhone": r.get("phone", "") or "",
                    "Active": r.get("active", ""),
                })
            offset += len(rows)

    _export_items_only()

    print(f"OK: odoo customers exported -> {customers_out}")
    print(f"OK: odoo contacts exported -> {contacts_out}")
    print(f"OK: odoo child partners exported -> {children_out}")
    print(f"OK: odoo child partners (all) exported -> {children_all_out}")
    print(f"OK: odoo delivery addresses exported -> {delivery_out}")
    print(f"OK: odoo items exported -> {items_out}")

    # Export Odoo vendor pricelist (product.supplierinfo)
    vendor_pricelist_out = os.path.join(odoo_root, "vendor_pricelist_odoo.csv")
    supplierinfo_external_ids = _load_external_ids("product.supplierinfo")
    tmpl_external_by_id: Dict[str, str] = {}
    if os.path.exists(items_out):
        _, item_rows = read_csv(items_out)
        for it in item_rows:
            tmpl_id = str(it.get("OdooTemplateId") or "").strip()
            tmpl_ext = (it.get("OdooTemplateExternalId") or "").strip()
            if tmpl_id and tmpl_ext and tmpl_id not in tmpl_external_by_id:
                tmpl_external_by_id[tmpl_id] = tmpl_ext
    supplier_fields = [
        "OdooId",
        "OdooExternalId",
        "OdooVendorId",
        "OdooVendor",
        "OdooTemplateId",
        "OdooTemplateExternalId",
        "OdooVariantId",
        "OdooVariantName",
        "OdooUnitPrice",
        "Currency",
        "MinQty",
        "Delay",
        "VendorProductCode",
        "VendorProductName",
    ]
    with open(vendor_pricelist_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=supplier_fields, delimiter=DELIMITER)
        writer.writeheader()
        offset = 0
        while True:
            rows = client.search_read(
                "product.supplierinfo",
                [],
                [
                    "id",
                    "partner_id",
                    "product_tmpl_id",
                    "product_id",
                    "price",
                    "currency_id",
                    "min_qty",
                    "delay",
                    "product_code",
                    "product_name",
                ],
                limit=batch,
                offset=offset,
            )
            if not rows:
                break
            for r in rows:
                partner = r.get("partner_id") or []
                tmpl = r.get("product_tmpl_id") or []
                variant = r.get("product_id") or []
                currency = r.get("currency_id") or []
                tmpl_id = str(tmpl[0]) if isinstance(tmpl, list) and tmpl else ""
                writer.writerow({
                    "OdooId": r.get("id", ""),
                    "OdooExternalId": supplierinfo_external_ids.get(str(r.get("id", "")), ""),
                    "OdooVendorId": partner[0] if isinstance(partner, list) and partner else "",
                    "OdooVendor": partner[1] if isinstance(partner, list) and len(partner) > 1 else "",
                    "OdooTemplateId": tmpl_id,
                    "OdooTemplateExternalId": tmpl_external_by_id.get(tmpl_id, ""),
                    "OdooVariantId": variant[0] if isinstance(variant, list) and variant else "",
                    "OdooVariantName": variant[1] if isinstance(variant, list) and len(variant) > 1 else "",
                    "OdooUnitPrice": r.get("price", ""),
                    "Currency": currency[1] if isinstance(currency, list) and len(currency) > 1 else "",
                    "MinQty": r.get("min_qty", ""),
                    "Delay": r.get("delay", ""),
                    "VendorProductCode": r.get("product_code", "") or "",
                    "VendorProductName": r.get("product_name", "") or "",
                })
            offset += len(rows)
    print(f"OK: odoo vendor pricelist exported -> {vendor_pricelist_out}")

    # Export Odoo users (salespeople)
    users_out = os.path.join(odoo_root, "users_odoo.csv")
    users_fields = ["OdooId", "OdooName", "OdooLogin", "Active"]
    with open(users_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=users_fields, delimiter=DELIMITER)
        writer.writeheader()
        offset = 0
        while True:
            rows = client.models.execute_kw(
                client.db,
                client.uid,
                client.apikey,
                "res.users",
                "search_read",
                [[]],
                {
                    "fields": ["id", "name", "login", "active"],
                    "limit": batch,
                    "offset": offset,
                    "context": {"active_test": False},
                },
            )
            if not rows:
                break
            for r in rows:
                writer.writerow({
                    "OdooId": r.get("id", ""),
                    "OdooName": r.get("name", "") or "",
                    "OdooLogin": r.get("login", "") or "",
                    "Active": r.get("active", ""),
                })
            offset += len(rows)
    print(f"OK: odoo users exported -> {users_out}")
    # Export Odoo pricelists
    pricelists_out = os.path.join(odoo_root, "pricelists_odoo.csv")
    pricelist_fields = ["OdooId", "OdooName", "Active", "CurrencyId", "CurrencyName"]
    with open(pricelists_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=pricelist_fields, delimiter=DELIMITER)
        writer.writeheader()
        offset = 0
        while True:
            rows = client.search_read(
                "product.pricelist",
                [],
                ["id", "name", "active", "currency_id"],
                limit=batch,
                offset=offset,
            )
            if not rows:
                break
            for r in rows:
                currency = r.get("currency_id") or []
                writer.writerow({
                    "OdooId": r.get("id", ""),
                    "OdooName": r.get("name", "") or "",
                    "Active": r.get("active", ""),
                    "CurrencyId": currency[0] if isinstance(currency, list) and currency else "",
                    "CurrencyName": currency[1] if isinstance(currency, list) and len(currency) > 1 else "",
                })
            offset += len(rows)
    print(f"OK: odoo pricelists exported -> {pricelists_out}")
    # Export Odoo currencies
    currencies_out = os.path.join(odoo_root, "currencies_odoo.csv")
    currency_fields = ["OdooId", "OdooName", "Code", "Symbol", "Rounding", "DecimalPlaces", "Active"]
    with open(currencies_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=currency_fields, delimiter=DELIMITER)
        writer.writeheader()
        offset = 0
        while True:
            rows = client.search_read(
                "res.currency",
                [],
                ["id", "name", "symbol", "rounding", "decimal_places", "active"],
                limit=batch,
                offset=offset,
            )
            if not rows:
                break
            for r in rows:
                writer.writerow({
                    "OdooId": r.get("id", ""),
                    "OdooName": r.get("name", "") or "",
                    "Code": r.get("name", "") or "",
                    "Symbol": r.get("symbol", "") or "",
                    "Rounding": r.get("rounding", ""),
                    "DecimalPlaces": r.get("decimal_places", ""),
                    "Active": r.get("active", ""),
                })
            offset += len(rows)
    print(f"OK: odoo currencies exported -> {currencies_out}")

    # Export Odoo chart of accounts (account.account)
    try:
        account_external_ids = _load_external_ids("account.account")
        account_fields_requested = [
            "id",
            "code",
            "name",
            "account_type",
            "deprecated",
            "reconcile",
            "company_id",
            "currency_id",
        ]
        available_fields = client.models.execute_kw(
            client.db,
            client.uid,
            client.apikey,
            "account.account",
            "fields_get",
            [],
            {"attributes": ["string", "type"]},
        ) or {}
        account_fields = [f for f in account_fields_requested if f in available_fields]
        chart_fields = [
            "OdooId",
            "OdooExternalId",
            "Code",
            "Name",
            "AccountType",
            "Deprecated",
            "Reconcile",
            "CompanyId",
            "CompanyName",
            "CurrencyId",
            "CurrencyName",
        ]
        with open(chart_out, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=chart_fields, delimiter=DELIMITER)
            writer.writeheader()
            offset = 0
            while True:
                rows = client.search_read(
                    "account.account",
                    [],
                    account_fields,
                    limit=batch,
                    offset=offset,
                )
                if not rows:
                    break
                for r in rows:
                    company = r.get("company_id") or []
                    currency = r.get("currency_id") or []
                    writer.writerow({
                        "OdooId": r.get("id", ""),
                        "OdooExternalId": account_external_ids.get(str(r.get("id", "")), ""),
                        "Code": r.get("code", "") or "",
                        "Name": r.get("name", "") or "",
                        "AccountType": r.get("account_type", "") or "",
                        "Deprecated": r.get("deprecated", ""),
                        "Reconcile": r.get("reconcile", ""),
                        "CompanyId": company[0] if isinstance(company, list) and company else "",
                        "CompanyName": company[1] if isinstance(company, list) and len(company) > 1 else "",
                        "CurrencyId": currency[0] if isinstance(currency, list) and currency else "",
                        "CurrencyName": currency[1] if isinstance(currency, list) and len(currency) > 1 else "",
                    })
                offset += len(rows)
        print(f"OK: odoo chart of accounts exported -> {chart_out}")
    except Exception as e:
        print(f"WARNING: failed to export Odoo chart of accounts -> {e}")

    # Export Odoo pricelist items (master + per-pricelist)
    pricelist_items_out = os.path.join(odoo_root, "pricelist_items_odoo.csv")
    pricelists_by_id = {}
    if os.path.exists(pricelists_out):
        with open(pricelists_out, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=DELIMITER)
            for r in reader:
                pid = str(r.get("OdooId", "")).strip()
                if pid:
                    pricelists_by_id[pid] = r.get("OdooName", "") or ""

    items_fields = [
        "OdooId",
        "PricelistId",
        "PricelistName",
        "AppliedOn",
        "ProductTemplateId",
        "ProductId",
        "MinQuantity",
        "FixedPrice",
        "PercentPrice",
        "DateStart",
        "DateEnd",
        "ComputePrice",
        "Base",
        "BasePricelistId",
        "CurrencyId",
    ]
    with open(pricelist_items_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=items_fields, delimiter=DELIMITER)
        writer.writeheader()
        offset = 0
        while True:
            rows = client.search_read(
                "product.pricelist.item",
                [],
                [
                    "id",
                    "pricelist_id",
                    "applied_on",
                    "product_tmpl_id",
                    "product_id",
                    "min_quantity",
                    "fixed_price",
                    "percent_price",
                    "date_start",
                    "date_end",
                    "compute_price",
                    "base",
                    "base_pricelist_id",
                    "currency_id",
                ],
                limit=batch,
                offset=offset,
            )
            if not rows:
                break
            for r in rows:
                pricelist = r.get("pricelist_id") or []
                pid = str(pricelist[0]) if isinstance(pricelist, list) and pricelist else ""
                pname = pricelist[1] if isinstance(pricelist, list) and len(pricelist) > 1 else pricelists_by_id.get(pid, "")
                tmpl = r.get("product_tmpl_id") or []
                prod = r.get("product_id") or []
                base_pl = r.get("base_pricelist_id") or []
                currency = r.get("currency_id") or []
                writer.writerow({
                    "OdooId": r.get("id", ""),
                    "PricelistId": pid,
                    "PricelistName": pname,
                    "AppliedOn": r.get("applied_on", "") or "",
                    "ProductTemplateId": tmpl[0] if isinstance(tmpl, list) and tmpl else "",
                    "ProductId": prod[0] if isinstance(prod, list) and prod else "",
                    "MinQuantity": r.get("min_quantity", ""),
                    "FixedPrice": r.get("fixed_price", ""),
                    "PercentPrice": r.get("percent_price", ""),
                    "DateStart": r.get("date_start", "") or "",
                    "DateEnd": r.get("date_end", "") or "",
                    "ComputePrice": r.get("compute_price", "") or "",
                    "Base": r.get("base", "") or "",
                    "BasePricelistId": base_pl[0] if isinstance(base_pl, list) and base_pl else "",
                    "CurrencyId": currency[0] if isinstance(currency, list) and currency else "",
                })
            offset += len(rows)
    print(f"OK: odoo pricelist items exported -> {pricelist_items_out}")

    # Split pricelist items per pricelist
    try:
        pricelist_dir = os.path.join(odoo_root, "pricelists")
        os.makedirs(pricelist_dir, exist_ok=True)
        by_list = {}
        with open(pricelist_items_out, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=DELIMITER)
            for row in reader:
                name = (row.get("PricelistName") or "UNKNOWN").strip()
                by_list.setdefault(name, []).append(row)
        for name, rows in by_list.items():
            safe_name = "".join([c if c.isalnum() or c in ("_", "-", " ") else "_" for c in name]).strip()
            if not safe_name:
                safe_name = "UNKNOWN"
            out_path = os.path.join(pricelist_dir, f"pricelist_{safe_name}.csv")
            with open(out_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=items_fields, delimiter=DELIMITER)
                writer.writeheader()
                writer.writerows(rows)
        print(f"OK: odoo pricelist items split -> {pricelist_dir}")
    except Exception:
        print("WARNING: failed to split pricelist items per list")
    # Export color attribute values
    try:
        attr_rows = client.search_read(
            "product.attribute",
            [],
            ["id", "name"],
            limit=batch,
            offset=0,
        )
        color_attr_ids = [r.get("id") for r in attr_rows if (r.get("name") or "").strip().lower().find("color") >= 0]
        color_out = os.path.join(odoo_root, "atributos_color.csv")
        with open(color_out, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=["OdooId", "OdooName", "AttributeId", "AttributeName"],
                delimiter=DELIMITER,
            )
            writer.writeheader()
            if color_attr_ids:
                offset = 0
                while True:
                    rows = client.search_read(
                        "product.attribute.value",
                        [["attribute_id", "in", color_attr_ids]],
                        ["id", "name", "attribute_id"],
                        limit=batch,
                        offset=offset,
                    )
                    if not rows:
                        break
                    for r in rows:
                        attr = r.get("attribute_id") or []
                        writer.writerow({
                            "OdooId": r.get("id", ""),
                            "OdooName": r.get("name", "") or "",
                            "AttributeId": attr[0] if isinstance(attr, list) and attr else "",
                            "AttributeName": attr[1] if isinstance(attr, list) and len(attr) > 1 else "",
                        })
                    offset += len(rows)
        print(f"OK: odoo color attributes exported -> {color_out}")
    except Exception:
        print("WARNING: failed to export Odoo color attributes")
    return 0


def sync_local(args: argparse.Namespace) -> int:
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    customer_sync = args.customers_sync
    item_sync = args.items_sync
    odoo_customers = args.odoo_customers
    odoo_items = args.odoo_items
    vendors_sync = args.vendors_sync
    odoo_vendors = args.odoo_vendors

    if not os.path.exists(customer_sync):
        print(f"ERROR: customers sync file not found: {customer_sync}")
        return 2
    if not os.path.exists(item_sync):
        print(f"ERROR: products sync file not found: {item_sync}")
        return 2
    if not os.path.exists(odoo_customers):
        print(f"ERROR: odoo customers file not found: {odoo_customers}")
        return 2
    if not os.path.exists(odoo_items):
        print(f"ERROR: odoo items file not found: {odoo_items}")
        return 2
    if not os.path.exists(vendors_sync):
        print(f"ERROR: vendors sync file not found: {vendors_sync}")
        return 2
    if not os.path.exists(odoo_vendors):
        print(f"ERROR: odoo vendors file not found: {odoo_vendors}")
        return 2

    _, odoo_cust_rows = read_csv(odoo_customers)
    _, odoo_item_rows = read_csv(odoo_items)
    _, odoo_vendor_rows = read_csv(odoo_vendors)

    odoo_cust_by_ref: Dict[str, List[Dict[str, str]]] = {}
    odoo_cust_by_name: Dict[str, List[Dict[str, str]]] = {}
    for r in odoo_cust_rows:
        ref = (r.get("OdooRef") or "").strip()
        name = (r.get("OdooName") or "").strip()
        if ref:
            odoo_cust_by_ref.setdefault(ref, []).append(r)
        if name:
            odoo_cust_by_name.setdefault(name, []).append(r)

    odoo_item_by_code: Dict[str, List[Dict[str, str]]] = {}
    odoo_item_by_template_ext_short: Dict[str, List[Dict[str, str]]] = {}
    odoo_item_by_id: Dict[str, Dict[str, str]] = {}
    for r in odoo_item_rows:
        if str(r.get("Active") or "").strip().lower() not in {"true", "1", "yes"}:
            continue
        code = (r.get("OdooItemCode") or "").strip()
        if code:
            odoo_item_by_code.setdefault(code, []).append(r)
        tmpl_ext = (r.get("OdooTemplateExternalId") or "").strip()
        if tmpl_ext.startswith("__import__."):
            tmpl_ext = tmpl_ext.split(".", 1)[1]
        if tmpl_ext:
            odoo_item_by_template_ext_short.setdefault(tmpl_ext, []).append(r)
        oid = str(r.get("OdooVariantId") or "").strip()
        if oid:
            odoo_item_by_id[oid] = r

    odoo_vendor_by_ref: Dict[str, List[Dict[str, str]]] = {}
    odoo_vendor_by_name: Dict[str, List[Dict[str, str]]] = {}
    odoo_vendor_by_id: Dict[str, Dict[str, str]] = {}
    for r in odoo_vendor_rows:
        ref = (r.get("OdooRef") or "").strip()
        name = (r.get("OdooName") or "").strip()
        oid = str(r.get("OdooId") or "").strip()
        if ref:
            odoo_vendor_by_ref.setdefault(ref, []).append(r)
        if name:
            odoo_vendor_by_name.setdefault(name, []).append(r)
        if oid:
            odoo_vendor_by_id[oid] = r

    customer_fields, customer_rows = read_csv(customer_sync)
    item_fields, item_rows = read_csv(item_sync)
    vendor_fields, vendor_rows = read_csv(vendors_sync)

    if "LastLookupAt" not in customer_fields:
        customer_fields.append("LastLookupAt")
    if "CustomerSyncStatus" not in customer_fields:
        customer_fields.append("CustomerSyncStatus")
    if "CustomerMismatchFields" not in customer_fields:
        customer_fields.append("CustomerMismatchFields")
    if "OdooExternalId" not in customer_fields:
        customer_fields.append("OdooExternalId")
    if "PriceLevel" not in customer_fields:
        customer_fields.append("PriceLevel")
    if "OdooPricelistId" not in customer_fields:
        customer_fields.append("OdooPricelistId")
    if "OdooPricelist" not in customer_fields:
        customer_fields.append("OdooPricelist")
    if "ExpectedOdooPricelistId" not in customer_fields:
        customer_fields.append("ExpectedOdooPricelistId")
    if "ExpectedOdooPricelist" not in customer_fields:
        customer_fields.append("ExpectedOdooPricelist")
    if "LastLookupAt" not in item_fields:
        item_fields.append("LastLookupAt")
    if "OdooTemplateId" not in item_fields:
        item_fields.append("OdooTemplateId")
    if "OdooVariantExternalId" not in item_fields:
        item_fields.append("OdooVariantExternalId")
    if "OdooTemplateExternalId" not in item_fields:
        item_fields.append("OdooTemplateExternalId")
    if "OdooExternalId" not in vendor_fields:
        vendor_fields.append("OdooExternalId")
    if "LastLookupAt" not in vendor_fields:
        vendor_fields.append("LastLookupAt")
    if "VendorSyncStatus" not in vendor_fields:
        vendor_fields.append("VendorSyncStatus")
    if "VendorMismatchFields" not in vendor_fields:
        vendor_fields.append("VendorMismatchFields")
    if "OdooId" not in vendor_fields:
        vendor_fields.append("OdooId")
    if "OdooName" not in vendor_fields:
        vendor_fields.append("OdooName")
    if "OdooRef" not in vendor_fields:
        vendor_fields.append("OdooRef")

    updated_customers = 0

    # Sage PriceLevel -> Odoo pricelist parity
    pricelist_parity = {}
    parity_path = os.path.join(os.path.dirname(customer_sync), "_parity_customer_pricelist.csv")
    if not os.path.exists(parity_path):
        parity_path = os.path.join(os.path.dirname(customer_sync), "_parity_pricelist.csv")
    if os.path.exists(parity_path):
        _, parity_rows = read_csv(parity_path)
        for p in parity_rows:
            level = (p.get("sage_price_level") or p.get("\ufeffsage_price_level") or "").strip()
            if not level:
                continue
            pname = (p.get("odoo_pricelist_name") or "").strip()
            if "(" in pname:
                pname = pname.split("(", 1)[0].strip()
            pricelist_parity[level] = {
                "id": (p.get("odoo_pricelist_id") or "").strip(),
                "name": pname,
            }

    def _resolve_price_level_for_parity(raw_level: str) -> Tuple[str, bool]:
        level = (raw_level or "").strip()
        if not level:
            return "", False
        if level in pricelist_parity:
            return level, True
        # Sage customers store PriceLevel as 0-based index (0..9),
        # while parity file is maintained as UI levels (1..10).
        if level.isdigit():
            shifted = str(int(level) + 1)
            if shifted in pricelist_parity:
                return shifted, True
        return level, False

    def _norm_text(value: str) -> str:
        return " ".join((value or "").strip().upper().split())

    def _norm_phone(value: str) -> str:
        return "".join(ch for ch in (value or "") if ch.isdigit())

    def _normalize_pricelist_name(value: str) -> str:
        raw = (value or "").strip()
        if not raw:
            return ""
        # "UK (GBP)" -> "UK", keep "Price Level 2" as-is.
        if "(" in raw:
            raw = raw.split("(", 1)[0].strip()
        return _norm_text(raw)

    def _short_external_name(value: str) -> str:
        raw = (value or "").strip()
        if not raw:
            return ""
        if "." in raw:
            raw = raw.split(".", 1)[1]
        return _norm_text(raw)

    def _customer_mismatches(row: Dict[str, str], record: Dict[str, str]) -> List[str]:
        # Inactive Sage customers are excluded from UPDATE exports;
        # avoid flagging them as mismatches to keep sync output actionable.
        if (row.get("CustomerIsInactive") or "").strip() == "1":
            return []
        checks = [
            ("name", row.get("Customer_Bill_Name", ""), record.get("OdooName", "")),
            ("reference", row.get("CustomerID", ""), record.get("OdooRef", "")),
            ("email", row.get("Email", ""), record.get("OdooEmail", "")),
            ("street", row.get("Street", ""), record.get("Street", "")),
            ("street2", row.get("Street2", ""), record.get("Street2", "")),
            ("city", row.get("City", ""), record.get("City", "")),
            ("zip", row.get("Zip", ""), record.get("Zip", "")),
        ]
        mismatches = [name for name, left, right in checks if _norm_text(left) != _norm_text(right)]
        if _norm_phone(row.get("Phone", "")) != _norm_phone(record.get("OdooPhone", "")):
            mismatches.append("phone")
        expected_pricelist = _normalize_pricelist_name(row.get("ExpectedOdooPricelist", ""))
        odoo_pricelist = _normalize_pricelist_name(record.get("OdooPricelist", ""))
        level_parity_found = (row.get("_PriceLevelParityFound") or "").strip() == "1"
        if level_parity_found:
            # Exact parity expected (including explicit "no pricelist").
            if expected_pricelist != odoo_pricelist:
                mismatches.append("pricelist")
        elif expected_pricelist:
            if not odoo_pricelist or expected_pricelist != odoo_pricelist:
                mismatches.append("pricelist")
        return mismatches

    for row in customer_rows:
        if truthy(row.get("Exclude")):
            continue
        customer_id = (row.get("CustomerID") or "").strip()
        customer_name = (row.get("Customer_Bill_Name") or "").strip()
        record = None
        match_source = ""
        if row.get("OdooId"):
            record = next((r for r in odoo_cust_rows if str(r.get("OdooId") or "").strip() == str(row.get("OdooId")).strip()), None)
            if record is not None:
                match_source = "id"
        if record is None and customer_id:
            matches = odoo_cust_by_ref.get(customer_id, [])
            if len(matches) == 1:
                record = matches[0]
                match_source = "ref"
            elif len(matches) > 1:
                target_name = _norm_text(customer_name)
                by_name = [m for m in matches if _norm_text(m.get("OdooName", "")) == target_name]
                if len(by_name) == 1:
                    record = by_name[0]
                    match_source = "ref_name"
                else:
                    target_ref = _norm_text(customer_id)
                    by_external = [
                        m for m in matches
                        if _short_external_name(m.get("OdooExternalId", "")) == target_ref
                    ]
                    if len(by_external) == 1:
                        record = by_external[0]
                        match_source = "ref_external"
        if record is None and args.customer_match_name and customer_name:
            matches = odoo_cust_by_name.get(customer_name, [])
            if len(matches) == 1:
                record = matches[0]
                match_source = "name"
        row["LastLookupAt"] = now
        level_key, level_parity_found = _resolve_price_level_for_parity(row.get("PriceLevel", ""))
        expected = pricelist_parity.get(level_key, {})
        row["ExpectedOdooPricelistId"] = expected.get("id", "")
        row["ExpectedOdooPricelist"] = expected.get("name", "")
        row["_PriceLevelParityFound"] = "1" if level_parity_found else "0"
        if record:
            strict_match = match_source in {"id", "ref", "ref_name", "ref_external"}
            if strict_match:
                row["OdooId"] = str(record.get("OdooId", ""))
                row["OdooExternalId"] = record.get("OdooExternalId", "") or ""
                row["OdooName"] = record.get("OdooName", "") or ""
                row["OdooPricelistId"] = str(record.get("OdooPricelistId", "") or "")
                row["OdooPricelist"] = record.get("OdooPricelist", "") or ""
                mismatches = _customer_mismatches(row, record)
                row["CustomerSyncStatus"] = "UPDATE" if mismatches else "MATCH"
                row["CustomerMismatchFields"] = "|".join(mismatches)
                updated_customers += 1
            else:
                # Name fallback can help discovery, but never for UPDATE safety.
                row["CustomerSyncStatus"] = "NEW"
                row["CustomerMismatchFields"] = "fallback_name_only"
        else:
            row["CustomerSyncStatus"] = "NEW"
            row["CustomerMismatchFields"] = ""

    updated_items = 0
    for row in item_rows:
        if row.get("OdooVariantId"):
            existing = odoo_item_by_id.get(str(row.get("OdooVariantId")).strip())
            if existing:
                # Validate stale variant links: variant id can still exist in Odoo but now
                # belong to a different SKU after product maintenance/reimports.
                item_id = (row.get("ItemID") or "").strip().upper()
                existing_code = (existing.get("OdooItemCode") or "").strip().upper()
                existing_template_short = _short_external_name(existing.get("OdooTemplateExternalId", ""))
                code_matches = bool(item_id) and bool(existing_code) and item_id == existing_code
                template_matches = bool(item_id) and item_id == existing_template_short
                if item_id and not (code_matches or template_matches):
                    row["OdooVariantId"] = ""
                    row["OdooVariantExternalId"] = ""
                    row["OdooTemplateId"] = ""
                    row["OdooTemplateExternalId"] = ""
                    row["OdooName"] = ""
                    row["OdooColor"] = ""
                    row["OdooItemCode"] = ""
                    # continue to rematch by code/template below
                else:
                    if not row.get("OdooColor"):
                        row["OdooColor"] = existing.get("OdooColor", "") or ""
                    if not row.get("OdooName"):
                        row["OdooName"] = existing.get("OdooName", "") or ""
                    if not row.get("OdooTemplateId"):
                        row["OdooTemplateId"] = existing.get("OdooTemplateId", "") or ""
                    if not row.get("OdooVariantExternalId"):
                        row["OdooVariantExternalId"] = existing.get("OdooVariantExternalId", "") or ""
                    if not row.get("OdooTemplateExternalId"):
                        row["OdooTemplateExternalId"] = existing.get("OdooTemplateExternalId", "") or ""
                    if not row.get("OdooItemCode"):
                        row["OdooItemCode"] = existing.get("OdooItemCode", "") or ""
                    continue
            # Stale OdooVariantId (record deleted/recreated in Odoo): clear and rematch by code/template.
            row["OdooVariantId"] = ""
            row["OdooVariantExternalId"] = ""
        if truthy(row.get("Exclude")):
            continue
        item_id = (row.get("ItemID") or "").strip()
        if not item_id:
            continue
        matches = odoo_item_by_code.get(item_id, [])
        if len(matches) != 1:
            # Fallback for single-variant templates imported without variant default_code.
            matches = odoo_item_by_template_ext_short.get(item_id, [])
        row["LastLookupAt"] = now
        if len(matches) == 1:
            record = matches[0]
            row["OdooVariantId"] = str(record.get("OdooVariantId", ""))
            row["OdooItemCode"] = record.get("OdooItemCode", "") or ""
            row["OdooName"] = record.get("OdooName", "") or ""
            row["OdooColor"] = record.get("OdooColor", "") or ""
            row["OdooTemplateId"] = record.get("OdooTemplateId", "") or ""
            row["OdooVariantExternalId"] = record.get("OdooVariantExternalId", "") or ""
            row["OdooTemplateExternalId"] = record.get("OdooTemplateExternalId", "") or ""
            updated_items += 1

    def _vendor_mismatches(row: Dict[str, str], record: Dict[str, str]) -> List[str]:
        checks = [
            ("name", row.get("Name", ""), record.get("OdooName", "")),
            ("reference", row.get("VendorID", ""), record.get("OdooRef", "")),
            ("email", row.get("Email", ""), record.get("Email", "")),
            ("street", row.get("Street", ""), record.get("Street", "")),
            ("street2", row.get("Street2", ""), record.get("Street2", "")),
            ("city", row.get("City", ""), record.get("City", "")),
            ("zip", row.get("Zip", ""), record.get("Zip", "")),
            ("country", row.get("Country", ""), record.get("Country", "")),
        ]
        mismatches = [name for name, left, right in checks if _norm_text(left) != _norm_text(right)]
        if _norm_phone(row.get("Phone", "")) != _norm_phone(record.get("Phone", "")):
            mismatches.append("phone")
        return mismatches

    updated_vendors = 0
    for row in vendor_rows:
        if truthy(row.get("IsInactive")):
            continue
        vendor_id = (row.get("VendorID") or "").strip()
        vendor_name = (row.get("Name") or "").strip()
        record = None
        match_source = ""
        existing_odoo_id = (row.get("OdooId") or "").strip()
        if existing_odoo_id:
            record = odoo_vendor_by_id.get(existing_odoo_id)
            if record is not None:
                match_source = "id"
        if record is None and vendor_id:
            matches = odoo_vendor_by_ref.get(vendor_id, [])
            if len(matches) == 1:
                record = matches[0]
                match_source = "ref"
        if record is None and args.vendor_match_name and vendor_name:
            matches = odoo_vendor_by_name.get(vendor_name, [])
            if len(matches) == 1:
                record = matches[0]
                match_source = "name"
        row["LastLookupAt"] = now
        if record:
            strict_match = match_source in {"id", "ref"}
            if strict_match:
                row["OdooId"] = str(record.get("OdooId", "") or "")
                row["OdooExternalId"] = record.get("OdooExternalId", "") or ""
                row["OdooName"] = record.get("OdooName", "") or ""
                row["OdooRef"] = record.get("OdooRef", "") or ""
                mismatches = _vendor_mismatches(row, record)
                row["VendorSyncStatus"] = "UPDATE" if mismatches else "MATCH"
                row["VendorMismatchFields"] = "|".join(mismatches)
                updated_vendors += 1
            else:
                row["VendorSyncStatus"] = "NEW"
                row["VendorMismatchFields"] = "fallback_name_only"
        else:
            row["VendorSyncStatus"] = "NEW"
            row["VendorMismatchFields"] = ""

    customer_rows.sort(key=lambda r: int(r["CustomerRecordNumber"]))
    item_rows.sort(key=lambda r: int(r["ItemRecordNumber"]))
    vendor_rows.sort(key=lambda r: int(r["VendorRecordNumber"]))

    write_csv(customer_sync, customer_fields, customer_rows)
    write_csv(item_sync, item_fields, item_rows)
    write_csv(vendors_sync, vendor_fields, vendor_rows)

    print(f"OK: customers updated with Odoo IDs: {updated_customers}")
    print(f"OK: products updated with Odoo IDs: {updated_items}")
    print(f"OK: vendors updated with Odoo IDs: {updated_vendors}")

    # Build FAILS for Odoo customers/items not found in Sage
    customer_fails_path = os.path.join(os.path.dirname(customer_sync), "_customer_FAILS.csv")
    item_fails_path = os.path.join(os.path.dirname(item_sync), "_product_FAILS.csv")

    sage_customer_ids = set()
    for r in customer_rows:
        cid = (r.get("CustomerID") or "").strip()
        if cid:
            sage_customer_ids.add(cid)

    odoo_customers_fail = []
    for r in odoo_cust_rows:
        ref = (r.get("OdooRef") or "").strip()
        if not ref or ref not in sage_customer_ids:
            row = dict(r)
            row["Reason"] = "missing_ref" if not ref else "ref_not_in_sage"
            odoo_customers_fail.append(row)

    # Build Sage index for customer name suggestions
    cust_index = defaultdict(list)
    for r in customer_rows:
        name = r.get("Customer_Bill_Name", "")
        norm = normalize_name(name)
        if norm:
            cust_index[norm[0]].append({
                "CustomerRecordNumber": r.get("CustomerRecordNumber", ""),
                "CustomerID": r.get("CustomerID", ""),
                "Customer_Bill_Name": name,
                "NormName": norm,
            })

    def best_match_customer(norm_name_value: str):
        candidates = cust_index.get(norm_name_value[0], [])
        if not candidates:
            return None, 0.0
        best = None
        best_score = 0.0
        for entry in candidates:
            ratio = SequenceMatcher(a=norm_name_value, b=entry["NormName"]).ratio()
            if ratio > best_score:
                best_score = ratio
                best = entry
        return best, best_score

    customer_fail_fields = ["OdooId", "OdooName", "OdooRef", "Active", "Reason"]
    extra_cust_cols = [
        "SuggestedSageCustomerRecordNumber",
        "SuggestedSageCustomerID",
        "SuggestedSageCustomerName",
        "SuggestedSageScore",
    ]
    for c in extra_cust_cols:
        if c not in customer_fail_fields:
            customer_fail_fields.append(c)

    with open(customer_fails_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=customer_fail_fields, delimiter=DELIMITER)
        writer.writeheader()
        for r in odoo_customers_fail:
            name = r.get("OdooName", "")
            norm = normalize_name(name)
            if norm:
                match, score = best_match_customer(norm)
                if match and score >= 0.75:
                    r["SuggestedSageCustomerRecordNumber"] = match.get("CustomerRecordNumber", "")
                    r["SuggestedSageCustomerID"] = match.get("CustomerID", "")
                    r["SuggestedSageCustomerName"] = match.get("Customer_Bill_Name", "")
                    r["SuggestedSageScore"] = f"{score:.3f}"
            writer.writerow({k: r.get(k, "") for k in customer_fail_fields})

    # Items FAILS (Odoo products not found in Sage)
    sage_item_ids = set()
    for r in item_rows:
        iid = (r.get("ItemID") or "").strip()
        if iid:
            sage_item_ids.add(iid)

    odoo_items_fail = []
    for r in odoo_item_rows:
        code = (r.get("OdooItemCode") or "").strip()
        if not code or code not in sage_item_ids:
            row = dict(r)
            row["Reason"] = "missing_code" if not code else "code_not_in_sage"
            odoo_items_fail.append(row)

    item_index = defaultdict(list)
    for r in item_rows:
        name = r.get("ItemDescription", "")
        norm = normalize_name(name)
        if norm:
            item_index[norm[0]].append({
                "ItemRecordNumber": r.get("ItemRecordNumber", ""),
                "ItemID": r.get("ItemID", ""),
                "ItemDescription": name,
                "NormName": norm,
            })

    def best_match_item(norm_name_value: str):
        candidates = item_index.get(norm_name_value[0], [])
        if not candidates:
            return None, 0.0
        best = None
        best_score = 0.0
        for entry in candidates:
            ratio = SequenceMatcher(a=norm_name_value, b=entry["NormName"]).ratio()
            if ratio > best_score:
                best_score = ratio
                best = entry
        return best, best_score

    item_fail_fields = ["OdooVariantId", "OdooName", "OdooColor", "OdooItemCode", "Active", "Reason"]
    extra_item_cols = [
        "SuggestedSageItemRecordNumber",
        "SuggestedSageItemID",
        "SuggestedSageItemDescription",
        "SuggestedSageScore",
    ]
    for c in extra_item_cols:
        if c not in item_fail_fields:
            item_fail_fields.append(c)

    with open(item_fails_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=item_fail_fields, delimiter=DELIMITER)
        writer.writeheader()
        for r in odoo_items_fail:
            name = r.get("OdooVariantName", "") or r.get("OdooName", "")
            norm = normalize_name(name)
            if norm:
                match, score = best_match_item(norm)
                if match and score >= 0.75:
                    r["SuggestedSageItemRecordNumber"] = match.get("ItemRecordNumber", "")
                    r["SuggestedSageItemID"] = match.get("ItemID", "")
                    r["SuggestedSageItemDescription"] = match.get("ItemDescription", "")
                    r["SuggestedSageScore"] = f"{score:.3f}"
            writer.writerow({k: r.get(k, "") for k in item_fail_fields})

    print(f"OK: customer FAILS -> {customer_fails_path}")
    print(f"OK: product FAILS -> {item_fails_path}")
    return 0


def build_employees_sync(args: argparse.Namespace) -> int:
    root = args.root_dir
    master_sage = os.path.join(root, "_master_sage")
    master_odoo = os.path.join(root, "_master_odoo")
    master = os.path.join(root, "_master")
    os.makedirs(master, exist_ok=True)

    employees_path = os.path.join(master_sage, "employees.csv")
    users_path = os.path.join(master_odoo, "users_odoo.csv")

    if not os.path.exists(employees_path):
        print(f"ERROR: missing {employees_path}")
        return 2

    months = [m.strip() for m in (args.months or "").split(",") if m.strip()]
    if not months:
        months = ["2026_02", "2026_03", "2026_04"]

    # Collect EmpRecordNumber from sales orders headers
    emp_ids_orders = set()
    # The main data lives under ENZO-Sage50/13_2026
    year_root = os.path.join(root, "13_2026")
    if not os.path.isdir(year_root):
        print(f"ERROR: missing {year_root}")
        return 2

    for m in months:
        header_name = f"{m}_sales_orders_headers.csv"
        found = False
        for sub in os.listdir(year_root):
            subdir = os.path.join(year_root, sub)
            if not os.path.isdir(subdir):
                continue
            candidate = os.path.join(subdir, header_name)
            if os.path.exists(candidate):
                found = True
                with open(candidate, newline="", encoding="utf-8") as f:
                    reader = csv.DictReader(f, delimiter=DELIMITER)
                    for row in reader:
                        emp = (row.get("EmpRecordNumber") or "").strip()
                        if emp:
                            emp_ids_orders.add(emp)
                break
        if not found:
            print(f"WARNING: sales order headers not found for {m}")

    # Collect EmpRecordNumber from invoices (JrnlHdr master) within months
    emp_ids_invoices = set()
    jrnlhdr_path = os.path.join(master_sage, "jrnlhdr.csv")
    if os.path.exists(jrnlhdr_path):
        # Build date ranges for months
        month_ranges = []
        for m in months:
            try:
                year, month = m.split("_")
                start = f"{year}-{month}-01"
                # naive month end: handle 12 -> next year
                y = int(year)
                mo = int(month)
                if mo == 12:
                    end = f"{y+1}-01-01"
                else:
                    end = f"{y}-{mo+1:02d}-01"
                month_ranges.append((start, end))
            except Exception:
                continue
        def in_ranges(date_str: str) -> bool:
            if not date_str:
                return False
            for start, end in month_ranges:
                if start <= date_str < end:
                    return True
            return False

        with open(jrnlhdr_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=DELIMITER)
            for row in reader:
                if row.get("Module") != "R":
                    continue
                # Invoices: JournalEx 8 (per existing export process)
                if row.get("JournalEx") != "8":
                    continue
                if not in_ranges(row.get("TransactionDate", "")):
                    continue
                emp = (row.get("EmpRecordNumber") or "").strip()
                if emp:
                    emp_ids_invoices.add(emp)

    emp_ids_2026 = emp_ids_orders.union(emp_ids_invoices)

    # Load employees (include all)
    employees = []
    with open(employees_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=DELIMITER)
        for row in reader:
            employees.append(row)

    # Load Odoo users for optional matching
    users = []
    if os.path.exists(users_path):
        with open(users_path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=DELIMITER)
            users = list(reader)

    def normalize(value: str) -> str:
        return (value or "").strip().lower()

    user_by_name = {normalize(u.get("OdooName", "")): u for u in users if u.get("OdooName")}
    user_by_login = {normalize(u.get("OdooLogin", "")): u for u in users if u.get("OdooLogin")}

    out_path = os.path.join(master, "employees_sync.csv")
    out_new_path = os.path.join(master, "employees_NEW.csv")
    fields = [
        "EmpRecordNumber",
        "EmployeeID",
        "EmployeeName",
        "Employee_FirstName",
        "Employee_LastName",
        "IsSalesRep",
        "EmployeeIsInactive",
        "JobTitle",
        "Email",
        "PhoneNumber",
        "PhoneWork",
        "PhoneMobile",
        "Address1",
        "Address2",
        "City",
        "State",
        "ZIP",
        "Country",
        "Invoiced2026",
        "OdooUserId",
        "OdooUserName",
        "OdooUserLogin",
        "OdooUserActive",
        "MatchReason",
    ]
    fields_new = [
        "EmpRecordNumber",
        "EmployeeID",
        "EmployeeName",
        "Employee_FirstName",
        "Employee_LastName",
        "IsSalesRep",
        "EmployeeIsInactive",
        "JobTitle",
        "Email",
        "PhoneNumber",
        "PhoneWork",
        "PhoneMobile",
        "Address1",
        "Address2",
        "City",
        "State",
        "ZIP",
        "Country",
        "Invoiced2026",
    ]

    rows_out = []
    rows_new = []
    for row in employees:
        emp_name = row.get("EmployeeName", "") or ""
        emp_id = row.get("EmployeeID", "") or ""
        match = None
        reason = ""
        if normalize(emp_name) in user_by_name:
            match = user_by_name[normalize(emp_name)]
            reason = "name"
        elif normalize(emp_id) in user_by_login:
            match = user_by_login[normalize(emp_id)]
            reason = "login"
        elif normalize(emp_id) in user_by_name:
            match = user_by_name[normalize(emp_id)]
            reason = "name_empid"

        out_row = {
            "EmpRecordNumber": row.get("EmpRecordNumber", "") or "",
            "EmployeeID": emp_id,
            "EmployeeName": emp_name,
            "Employee_FirstName": row.get("Employee_FirstName", "") or "",
            "Employee_LastName": row.get("Employee_LastName", "") or "",
            "IsSalesRep": row.get("IsSalesRep", "") or "",
            "EmployeeIsInactive": row.get("EmployeeIsInactive", "") or "",
            "JobTitle": row.get("JobTitle", "") or "",
            "Email": row.get("Email", "") or "",
            "PhoneNumber": row.get("PhoneNumber", "") or "",
            "PhoneWork": row.get("PhoneWork", "") or "",
            "PhoneMobile": row.get("PhoneMobile", "") or "",
            "Address1": row.get("Address1", "") or "",
            "Address2": row.get("Address2", "") or "",
            "City": row.get("City", "") or "",
            "State": row.get("State", "") or "",
            "ZIP": row.get("ZIP", "") or "",
            "Country": row.get("Country", "") or "",
            "Invoiced2026": "X" if (row.get("EmpRecordNumber") or "").strip() in emp_ids_2026 else "",
            "OdooUserId": match.get("OdooId", "") if match else "",
            "OdooUserName": match.get("OdooName", "") if match else "",
            "OdooUserLogin": match.get("OdooLogin", "") if match else "",
            "OdooUserActive": match.get("Active", "") if match else "",
            "MatchReason": reason,
        }
        rows_out.append(out_row)
        if not match:
            rows_new.append(out_row)

    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields, delimiter=DELIMITER)
        writer.writeheader()
        for r in rows_out:
            writer.writerow(r)

    with open(out_new_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields_new, delimiter=DELIMITER)
        writer.writeheader()
        for r in rows_new:
            writer.writerow({k: r.get(k, "") for k in fields_new})

    print(f"OK: employees sync -> {out_path} ({len(rows_out)} rows)")
    print(f"OK: employees NEW -> {out_new_path} ({len(rows_new)} rows)")
    return 0


def build_pricelist_parity(args: argparse.Namespace) -> int:
    root = args.root_dir
    master = os.path.join(root, "_master")
    master_odoo = os.path.join(root, "_master_odoo")
    os.makedirs(master, exist_ok=True)

    pricelists_path = os.path.join(master_odoo, "pricelists_odoo.csv")
    currencies_path = os.path.join(master_odoo, "currencies_odoo.csv")
    parity_currency_out = os.path.join(master, "_parity_currency.csv")
    parity_pricelist_out = os.path.join(master, "_parity_pricelist.csv")

    # Build currency parity from Odoo currencies
    currency_rows = []
    if os.path.exists(currencies_path):
        _, currency_rows = read_csv(currencies_path)
    with open(parity_currency_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["currency_code", "odoo_currency_id", "odoo_currency_name", "symbol", "match_type"],
            delimiter=DELIMITER,
        )
        writer.writeheader()
        for r in currency_rows:
            writer.writerow({
                "currency_code": (r.get("Code") or "").strip(),
                "odoo_currency_id": (r.get("OdooId") or "").strip(),
                "odoo_currency_name": (r.get("OdooName") or "").strip(),
                "symbol": (r.get("Symbol") or "").strip(),
                "match_type": "odoo",
            })

    # Build pricelist parity template (levels 1..10)
    pricelist_rows = []
    if os.path.exists(pricelists_path):
        _, pricelist_rows = read_csv(pricelists_path)
    with open(parity_pricelist_out, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "sage_price_level",
                "odoo_pricelist_id",
                "odoo_pricelist_name",
                "odoo_currency_id",
                "odoo_currency_code",
                "notes",
            ],
            delimiter=DELIMITER,
        )
        writer.writeheader()
        # default empty mappings for levels 1..10
        for level in range(1, 11):
            writer.writerow({
                "sage_price_level": str(level),
                "odoo_pricelist_id": "",
                "odoo_pricelist_name": "",
                "odoo_currency_id": "",
                "odoo_currency_code": "",
                "notes": "",
            })
        # Append available pricelists as reference rows
        for r in pricelist_rows:
            writer.writerow({
                "sage_price_level": "",
                "odoo_pricelist_id": (r.get("OdooId") or "").strip(),
                "odoo_pricelist_name": (r.get("OdooName") or "").strip(),
                "odoo_currency_id": (r.get("CurrencyId") or "").strip(),
                "odoo_currency_code": (r.get("CurrencyName") or "").strip(),
                "notes": "odoo_pricelist_reference",
            })

    print(f"OK: parity currency -> {parity_currency_out}")
    print(f"OK: parity pricelist -> {parity_pricelist_out}")
    return 0


def build_pricelist_lines(args: argparse.Namespace) -> int:
    root = args.root_dir
    master = os.path.join(root, "_master")
    master_sage = os.path.join(root, "_master_sage")
    master_odoo = os.path.join(root, "_master_odoo")
    os.makedirs(master, exist_ok=True)

    items_path = args.items_master
    odoo_items_path = args.items_odoo
    pricelist_items_path = args.pricelist_items_odoo
    parity_pricelist_path = args.parity_pricelist

    if not os.path.exists(items_path):
        print(f"ERROR: missing {items_path}")
        return 2
    if not os.path.exists(odoo_items_path):
        print(f"ERROR: missing {odoo_items_path}")
        return 2
    if not os.path.exists(parity_pricelist_path):
        print(f"ERROR: missing {parity_pricelist_path}")
        return 2

    # Map Sage ItemID -> OdooTemplateId
    item_to_template = {}
    _, odoo_item_rows = read_csv(odoo_items_path)
    for r in odoo_item_rows:
        code = (r.get("OdooItemCode") or "").strip()
        tmpl_id = (r.get("OdooTemplateId") or "").strip()
        if code and tmpl_id:
            item_to_template[code] = tmpl_id

    # Load parity mapping: price level -> pricelist info
    parity_map = {}
    _, parity_rows = read_csv(parity_pricelist_path)
    for r in parity_rows:
        level = (r.get("sage_price_level") or r.get("\ufeffsage_price_level") or "").strip()
        if not level:
            continue
        parity_map[level] = r

    def _to_float(raw: str):
        value = (raw or "").strip()
        if not value:
            return None
        # Sage exports usually use comma decimal separator (e.g. 79,95)
        if "," in value and "." not in value:
            value = value.replace(",", ".")
        else:
            value = value.replace(",", "")
        try:
            return float(value)
        except Exception:
            return None

    def _fmt_price(number: float) -> str:
        return f"{number:.2f}"

    price_cols = [f"PriceLevel{i}Amount" for i in range(1, 11)]
    # Consolidate by pricelist + product template (one row per product, not per variant)
    grouped_prices: Dict[tuple, set] = defaultdict(set)
    grouped_meta: Dict[tuple, Dict[str, str]] = {}
    _, items_rows = read_csv(items_path)
    for r in items_rows:
        item_id = (r.get("ItemID") or "").strip()
        tmpl_id = item_to_template.get(item_id)
        if not tmpl_id:
            continue
        for i, col in enumerate(price_cols, start=1):
            price_num = _to_float(r.get(col) or "")
            if price_num is None or abs(price_num) < 0.0000001:
                continue
            parity = parity_map.get(str(i), {})
            pricelist_id = (parity.get("odoo_pricelist_id") or "").strip()
            pricelist_name = (parity.get("odoo_pricelist_name") or "").strip()
            currency_id = (parity.get("odoo_currency_id") or "").strip()
            if not pricelist_id:
                continue
            key = (pricelist_id, tmpl_id)
            grouped_prices[key].add(price_num)
            if key not in grouped_meta:
                grouped_meta[key] = {
                    "OdooId": "",
                    "PricelistId": pricelist_id,
                    "PricelistName": pricelist_name,
                    "AppliedOn": "1_product",
                    "ProductTemplateId": tmpl_id,
                    "ProductId": "",
                    "MinQuantity": "0.0",
                    "PercentPrice": "0.0",
                    "DateStart": "",
                    "DateEnd": "",
                    "ComputePrice": "fixed",
                    "Base": "list_price",
                    "BasePricelistId": "",
                    "CurrencyId": currency_id,
                }

    lines = []
    conflicts = []
    for key, prices in grouped_prices.items():
        meta = grouped_meta[key]
        if len(prices) == 1:
            price_num = next(iter(prices))
            row = dict(meta)
            row["FixedPrice"] = _fmt_price(price_num)
            lines.append(row)
        else:
            # Keep conflicts out of import; export to a dedicated review file.
            conflicts.append({
                "PricelistId": meta["PricelistId"],
                "PricelistName": meta["PricelistName"],
                "ProductTemplateId": meta["ProductTemplateId"],
                "CurrencyId": meta["CurrencyId"],
                "PricesFound": " | ".join([_fmt_price(p) for p in sorted(prices)]),
                "Reason": "multiple_variant_prices_for_same_template",
            })

    out_path = os.path.join(master, "pricelist_lines.csv")
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "OdooId",
                "PricelistId",
                "PricelistName",
                "AppliedOn",
                "ProductTemplateId",
                "ProductId",
                "MinQuantity",
                "FixedPrice",
                "PercentPrice",
                "DateStart",
                "DateEnd",
                "ComputePrice",
                "Base",
                "BasePricelistId",
                "CurrencyId",
            ],
            delimiter=DELIMITER,
        )
        writer.writeheader()
        writer.writerows(lines)

    conflicts_path = os.path.join(master, "pricelist_lines_CONFLICTS.csv")
    with open(conflicts_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "PricelistId",
                "PricelistName",
                "ProductTemplateId",
                "CurrencyId",
                "PricesFound",
                "Reason",
            ],
            delimiter=DELIMITER,
        )
        writer.writeheader()
        writer.writerows(conflicts)

    # Build NEW by excluding existing Odoo pricelist items
    existing_keys = set()
    if os.path.exists(pricelist_items_path):
        _, existing_rows = read_csv(pricelist_items_path)
        for r in existing_rows:
            key = (
                (r.get("PricelistId") or "").strip(),
                (r.get("AppliedOn") or "").strip(),
                (r.get("ProductTemplateId") or "").strip(),
                (r.get("MinQuantity") or "").strip(),
            )
            existing_keys.add(key)

    new_lines = []
    for r in lines:
        key = (
            (r.get("PricelistId") or "").strip(),
            (r.get("AppliedOn") or "").strip(),
            (r.get("ProductTemplateId") or "").strip(),
            (r.get("MinQuantity") or "").strip(),
        )
        if key in existing_keys:
            continue
        new_lines.append(r)

    out_new_path = os.path.join(master, "pricelist_lines_NEW.csv")
    with open(out_new_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "OdooId",
                "PricelistId",
                "PricelistName",
                "AppliedOn",
                "ProductTemplateId",
                "ProductId",
                "MinQuantity",
                "FixedPrice",
                "PercentPrice",
                "DateStart",
                "DateEnd",
                "ComputePrice",
                "Base",
                "BasePricelistId",
                "CurrencyId",
            ],
            delimiter=DELIMITER,
        )
        writer.writeheader()
        writer.writerows(new_lines)

    print(f"OK: pricelist lines -> {out_path} ({len(lines)} rows)")
    print(f"OK: pricelist lines NEW -> {out_new_path} ({len(new_lines)} rows)")
    print(f"OK: pricelist conflicts -> {conflicts_path} ({len(conflicts)} rows)")
    return 0


def build_pricelist_import(args: argparse.Namespace) -> int:
    root = args.root_dir
    master_odoo = os.path.join(root, "_master_odoo")
    template_path = args.template_path
    sync_path = args.sync_path
    out_path = args.out_path
    items_odoo_path = os.path.join(master_odoo, "items_odoo.csv")
    pricelists_path = os.path.join(master_odoo, "pricelists_odoo.csv")

    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2
    if not os.path.exists(sync_path):
        print(f"ERROR: sync not found: {sync_path}")
        return 2
    if not os.path.exists(items_odoo_path):
        print(f"ERROR: odoo items not found: {items_odoo_path}")
        return 2
    if not os.path.exists(pricelists_path):
        print(f"ERROR: odoo pricelists not found: {pricelists_path}")
        return 2

    # Load template headers from CSV
    with open(template_path, newline="", encoding="utf-8") as f:
        reader = csv.reader(f, delimiter=DELIMITER)
        headers = next(reader)

    if "{date}" in out_path:
        out_path = out_path.format(date=datetime.now().strftime("%Y%m%d"))
    if not out_path.lower().endswith(".csv"):
        out_path = os.path.splitext(out_path)[0] + ".csv"

    tmpl_to_external = {}
    _, odoo_item_rows = read_csv(items_odoo_path)
    for row in odoo_item_rows:
        tmpl_id = (row.get("OdooTemplateId") or "").strip()
        external_id = (row.get("OdooTemplateExternalId") or "").strip()
        if tmpl_id and external_id and tmpl_id not in tmpl_to_external:
            tmpl_to_external[tmpl_id] = external_id.split(".", 1)[1] if "." in external_id else external_id

    pricelist_map = {}
    _, pricelist_rows = read_csv(pricelists_path)
    for row in pricelist_rows:
        pricelist_id = (row.get("OdooId") or "").strip()
        if pricelist_id:
            pricelist_map[pricelist_id] = {
                "name": (row.get("OdooName") or "").strip(),
                "currency": (row.get("CurrencyName") or "").strip(),
            }

    def normalize_pricelist(name: str, currency: str):
        raw_name = (name or "").strip()
        upper_name = raw_name.upper()
        cur = (currency or "").strip().upper()
        if upper_name in {"USA", "USA (USD)"}:
            return "USA", "USA", "USD"
        if upper_name in {"EU", "EU (EUR)"}:
            return "EU", "EU", "EUR"
        if upper_name in {"UK", "UK (GBP)"}:
            return "UK", "UK", "GBP"
        if upper_name in {"CAD", "CAD (CAD)"}:
            return "CAD", "CAD", "CAD"
        if upper_name in {"AUD", "AUD (AUD)"}:
            return "AUD", "AUD", "AUD"
        if upper_name in {"PRICE LEVEL 2", "PRICE LEVEL 2 (USD)"}:
            return "PRICE_LEVEL_2_USD", "Price Level 2", "USD"
        if upper_name in {"PRICE LEVEL 3", "PRICE LEVEL 3 (USD)"}:
            return "PRICE_LEVEL_3_USD", "Price Level 3", "USD"
        if upper_name in {"PRICE LEVEL 4", "PRICE LEVEL 4 (USD)"}:
            return "PRICE_LEVEL_4_USD", "Price Level 4", "USD"
        if upper_name in {"PRICE LEVEL 5", "PRICE LEVEL 5 (USD)"}:
            return "PRICE_LEVEL_5_USD", "Price Level 5", "USD"
        if upper_name in {"DISTRIBUTOR US", "DISTRIBUTOR US (USD)"}:
            return "DISTRIBUTOR_US", "Distributor US", "USD"
        plain_name = upper_name.split("(", 1)[0].strip() if upper_name else ""
        safe_id = "".join(ch if ch.isalnum() or ch == "_" else "_" for ch in plain_name)
        while "__" in safe_id:
            safe_id = safe_id.replace("__", "_")
        safe_id = safe_id.strip("_")
        return safe_id, (raw_name or plain_name), cur

    def normalize_price(value: str) -> str:
        raw = (value or "").strip()
        if not raw:
            return ""
        try:
            return f"{float(raw.replace(',', '.')):.2f}"
        except Exception:
            return raw.replace(",", ".")

    headers_sync, rows = read_csv(sync_path)
    out_rows = []
    missing_template_external = 0
    for row in rows:
        pricelist_id = (row.get("PricelistId") or "").strip()
        tmpl_id = (row.get("ProductTemplateId") or "").strip()
        fixed_price = (row.get("FixedPrice") or "").strip()
        tmpl_external = tmpl_to_external.get(tmpl_id, "")
        if tmpl_id and not tmpl_external:
            missing_template_external += 1
        if not pricelist_id or not tmpl_external:
            continue
        plist = pricelist_map.get(pricelist_id, {})
        out_id, out_name, out_currency = normalize_pricelist(
            plist.get("name", row.get("PricelistName", "")),
            plist.get("currency", row.get("CurrencyId", "")),
        )
        out_rows.append({
            "id": out_id,
            "name": out_name,
            "country_group_ids/id": "",
            "currency_id": out_currency,
            "item_ids/applied_on": "Product",
            "item_ids/base": "Sales Price",
            "item_ids/compute_price": "Fixed Price",
            "item_ids/fixed_price": normalize_price(fixed_price),
            "item_ids/product_tmpl_id/id": tmpl_external,
        })

    seen = set()
    unique_rows = []
    for row in out_rows:
        key = (row.get("id", ""), row.get("item_ids/product_tmpl_id/id", ""))
        if key in seen:
            continue
        seen.add(key)
        unique_rows.append(row)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter=DELIMITER)
        writer.writeheader()
        for row in unique_rows:
            writer.writerow({h: row.get(h, "") for h in headers})

    print(f"OK: pricelist import -> {out_path} ({len(unique_rows)} rows)")
    if missing_template_external:
        print(f"WARNING: missing product template external IDs: {missing_template_external}")
    return 0


def build_pricelist_update(args: argparse.Namespace) -> int:
    root = args.root_dir
    master = os.path.join(root, "_master")
    master_odoo = os.path.join(root, "_master_odoo")
    template_path = args.template_path
    lines_path = args.lines_path
    existing_path = args.pricelist_items_odoo
    items_odoo_path = args.items_odoo
    pricelists_path = args.pricelists_odoo
    out_path = args.out_path

    required = [template_path, lines_path, existing_path, items_odoo_path, pricelists_path]
    for path in required:
        if not os.path.exists(path):
            print(f"ERROR: missing {path}")
            return 2

    with open(template_path, newline="", encoding="utf-8") as f:
        headers = next(csv.reader(f, delimiter=DELIMITER))
    if "item_ids/id" in headers:
        headers = ["item_ids/.id" if h == "item_ids/id" else h for h in headers]
    if "item_ids/.id" not in headers:
        headers = ["item_ids/.id"] + headers

    if "{date}" in out_path:
        out_path = out_path.format(date=datetime.now().strftime("%Y%m%d"))

    tmpl_to_external = {}
    _, item_rows = read_csv(items_odoo_path)
    for row in item_rows:
        tmpl_id = (row.get("OdooTemplateId") or "").strip()
        external_id = (row.get("OdooTemplateExternalId") or "").strip()
        if tmpl_id and external_id and tmpl_id not in tmpl_to_external:
            tmpl_to_external[tmpl_id] = external_id.split(".", 1)[1] if "." in external_id else external_id

    pricelist_map = {}
    _, pricelist_rows = read_csv(pricelists_path)
    for row in pricelist_rows:
        pricelist_id = (row.get("OdooId") or "").strip()
        if pricelist_id:
            pricelist_map[pricelist_id] = {
                "name": (row.get("OdooName") or "").strip(),
                "currency": (row.get("CurrencyName") or "").strip(),
            }

    def normalize_price(value: str):
        raw = (value or "").strip()
        if not raw:
            return None
        try:
            return float(raw.replace(",", "."))
        except Exception:
            return None

    def format_price(value: float) -> str:
        return f"{value:.2f}"

    def normalize_pricelist(name: str, currency: str):
        raw_name = (name or "").strip()
        upper_name = raw_name.upper()
        cur = (currency or "").strip().upper()
        if upper_name in {"USA", "USA (USD)"}:
            return "USA", "USA", "USD"
        if upper_name in {"EU", "EU (EUR)"}:
            return "EU", "EU", "EUR"
        if upper_name in {"UK", "UK (GBP)"}:
            return "UK", "UK", "GBP"
        if upper_name in {"CAD", "CAD (CAD)"}:
            return "CAD", "CAD", "CAD"
        if upper_name in {"AUD", "AUD (AUD)"}:
            return "AUD", "AUD", "AUD"
        if upper_name in {"PRICE LEVEL 2", "PRICE LEVEL 2 (USD)"}:
            return "PRICE_LEVEL_2_USD", "Price Level 2", "USD"
        if upper_name in {"PRICE LEVEL 3", "PRICE LEVEL 3 (USD)"}:
            return "PRICE_LEVEL_3_USD", "Price Level 3", "USD"
        if upper_name in {"PRICE LEVEL 4", "PRICE LEVEL 4 (USD)"}:
            return "PRICE_LEVEL_4_USD", "Price Level 4", "USD"
        if upper_name in {"PRICE LEVEL 5", "PRICE LEVEL 5 (USD)"}:
            return "PRICE_LEVEL_5_USD", "Price Level 5", "USD"
        if upper_name in {"DISTRIBUTOR US", "DISTRIBUTOR US (USD)"}:
            return "DISTRIBUTOR_US", "Distributor US", "USD"
        plain_name = upper_name.split("(", 1)[0].strip() if upper_name else ""
        safe_id = "".join(ch if ch.isalnum() or ch == "_" else "_" for ch in plain_name)
        while "__" in safe_id:
            safe_id = safe_id.replace("__", "_")
        safe_id = safe_id.strip("_")
        return safe_id, (raw_name or plain_name), cur

    existing_by_key = {}
    _, existing_rows = read_csv(existing_path)
    for row in existing_rows:
        key = (
            (row.get("PricelistId") or "").strip(),
            (row.get("AppliedOn") or "").strip(),
            (row.get("ProductTemplateId") or "").strip(),
            (row.get("MinQuantity") or "").strip(),
        )
        existing_by_key[key] = row

    _, line_rows = read_csv(lines_path)
    update_rows = []
    for row in line_rows:
        key = (
            (row.get("PricelistId") or "").strip(),
            (row.get("AppliedOn") or "").strip(),
            (row.get("ProductTemplateId") or "").strip(),
            (row.get("MinQuantity") or "").strip(),
        )
        existing = existing_by_key.get(key)
        if not existing:
            continue
        new_price = normalize_price(row.get("FixedPrice") or "")
        old_price = normalize_price(existing.get("FixedPrice") or "")
        if new_price is None or old_price is None:
            continue
        if abs(new_price - old_price) < 0.0001:
            continue

        pricelist_id = key[0]
        tmpl_id = key[2]
        tmpl_external = tmpl_to_external.get(tmpl_id, "")
        if not tmpl_external:
            continue
        plist = pricelist_map.get(pricelist_id, {})
        out_id, out_name, out_currency = normalize_pricelist(
            plist.get("name", row.get("PricelistName", "")),
            plist.get("currency", row.get("CurrencyId", "")),
        )
        update_rows.append({
            "item_ids/.id": existing.get("OdooId", ""),
            "id": out_id,
            "name": out_name,
            "country_group_ids/id": "",
            "currency_id": out_currency,
            "item_ids/applied_on": "Product",
            "item_ids/base": "Sales Price",
            "item_ids/compute_price": "Fixed Price",
            "item_ids/fixed_price": format_price(new_price),
            "item_ids/product_tmpl_id/id": tmpl_external,
        })

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, delimiter=DELIMITER)
        writer.writeheader()
        for row in update_rows:
            writer.writerow({h: row.get(h, "") for h in headers})

    print(f"OK: pricelist update -> {out_path} ({len(update_rows)} rows)")
    return 0


def build_customers_update(args: argparse.Namespace) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        print("ERROR: openpyxl not available")
        return 2

    sync_path = args.customers_sync
    customers_master_path = args.customers_master
    template_path = args.template_path
    master_out = args.master_out
    out_path = args.out_path

    if not os.path.exists(sync_path):
        print(f"ERROR: customers sync not found: {sync_path}")
        return 2
    if not os.path.exists(customers_master_path):
        print(f"ERROR: customers master not found: {customers_master_path}")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2
    if "{date}" in out_path:
        out_path = out_path.format(date=datetime.now().strftime("%Y%m%d"))

    state_parity_path = os.path.join(os.path.dirname(sync_path), "_parity_state.csv")
    state_parity = load_state_parity(state_parity_path)

    def _odoo_state(raw_state: str) -> str:
        raw = (raw_state or "").strip()
        if not raw:
            return ""
        info = state_parity.get(raw, {})
        state_name = info.get("state_name") or raw
        country_name = info.get("country_name") or ""
        if state_name and country_name == "United States":
            return f"{state_name} (US)"
        if state_name and country_name == "Canada":
            return f"{state_name} (CA)"
        return state_name

    headers_sync, rows = read_csv(sync_path)
    _, customers_master_rows = read_csv(customers_master_path)
    credit_msg_by_record: Dict[str, str] = {}
    default_credit_msg = "You have requested to be notified when a transaction is created for this customer."
    for r in customers_master_rows:
        rec = (r.get("CustomerRecordNumber") or "").strip()
        if not rec:
            continue
        msg = (r.get("CreditStatusMsg") or "").strip()
        if msg and msg != default_credit_msg:
            credit_msg_by_record[rec] = msg
    update_rows = [
        r for r in rows
        if not truthy(r.get("Exclude"))
        and (r.get("CustomerIsInactive") or "").strip() != "1"
    ]

    skipped_blank_name = 0
    skipped_blank_external_id = 0
    prepared_rows = []
    for r in update_rows:
        name = (r.get("Customer_Bill_Name") or "").strip()
        if not name:
            skipped_blank_name += 1
            continue
        odoo_id = (r.get("OdooId") or "").strip()
        if not odoo_id:
            skipped_blank_external_id += 1
            continue
        ext_id = (r.get("OdooExternalId") or "").strip()
        if ext_id.startswith("__import__."):
            ext_id = ext_id.split(".", 1)[1]
        if not ext_id:
            ext_id = sanitize_external_id(r.get("CustomerID", ""))
        if not ext_id:
            ext_id = sanitize_external_id(r.get("CustomerRecordNumber", ""))
        if not ext_id:
            skipped_blank_external_id += 1
            continue
        rec = (r.get("CustomerRecordNumber") or "").strip()
        credit_msg = credit_msg_by_record.get(rec, "")
        sync_status = (r.get("CustomerSyncStatus") or "").strip().upper()
        # Include normal UPDATE mismatches plus rows that need Notes backfill from CreditStatusMsg.
        if sync_status != "UPDATE" and not credit_msg:
            continue
        prepared_rows.append((r, name, ext_id, odoo_id))

    def build_workbook(target_path: str) -> None:
        wb = load_workbook(template_path)
        ws = wb["Partners"] if "Partners" in wb.sheetnames else wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if "Notes" not in headers:
            headers.append("Notes")
        for c, h in enumerate(headers, start=1):
            ws.cell(row=1, column=c).value = h
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for r, name, ext_id, odoo_id in prepared_rows:
            out = {
                "Database ID": odoo_id,
                "External_ID": ext_id,
                "name": name,
                "is_company": "TRUE",
                "company_name": "",
                "country_id": r.get("Country", ""),
                "state_id": _odoo_state(r.get("State", "")),
                "zip": r.get("Zip", ""),
                "city": r.get("City", ""),
                "street": r.get("Street", ""),
                "street2": r.get("Street2", ""),
                "phone": r.get("Phone", ""),
                "email": r.get("Email", ""),
                "Reference": r.get("CustomerID", ""),
                "Pricelist": r.get("ExpectedOdooPricelist", ""),
                "Notes": credit_msg_by_record.get((r.get("CustomerRecordNumber") or "").strip(), ""),
                "Language": "en_US",
            }
            ws.append([out.get(h, "") for h in headers])
        os.makedirs(os.path.dirname(target_path), exist_ok=True)
        wb.save(target_path)

    build_workbook(master_out)
    build_workbook(out_path)
    print(f"OK: customers UPDATE master -> {master_out} ({len(prepared_rows)} rows)")
    print(f"OK: customers UPDATE import -> {out_path} ({len(prepared_rows)} rows)")
    if skipped_blank_name or skipped_blank_external_id:
        print(
            "WARN: skipped customers UPDATE rows "
            f"(blank name={skipped_blank_name}, blank OdooId/external id={skipped_blank_external_id})"
        )
    return 0


def _safe_float(raw: str) -> float:
    value = (raw or "").strip()
    if not value:
        return 0.0
    value = value.replace(",", ".")
    try:
        return float(value)
    except ValueError:
        return 0.0


def _short_external_id(raw: str) -> str:
    value = (raw or "").strip()
    if value.startswith("__import__."):
        return value.split(".", 1)[1]
    if "." in value:
        return value.split(".", 1)[1]
    return value


def _template_external_id_for_import(raw: str) -> str:
    value = (raw or "").strip()
    if not value:
        return ""
    return value


def build_vendor_pricelist_sync(args: argparse.Namespace) -> int:
    jrnlhdr_path = args.jrnlhdr_master
    jrnlrow_path = args.jrnlrow_master
    vendors_sync_path = args.vendors_sync
    products_sync_path = args.products_sync
    odoo_vendor_pricelist_path = args.odoo_vendor_pricelist
    out_sync = args.out_sync
    out_new = args.out_new
    out_update = args.out_update
    out_conflicts = args.out_conflicts

    required = [
        jrnlhdr_path,
        jrnlrow_path,
        vendors_sync_path,
        products_sync_path,
        odoo_vendor_pricelist_path,
    ]
    for p in required:
        if not os.path.exists(p):
            print(f"ERROR: missing {p}")
            return 2

    hdr_by_post: Dict[str, Dict[str, str]] = {}
    with open(jrnlhdr_path, newline="", encoding="utf-8-sig") as f_hdr:
        reader = csv.DictReader(f_hdr, delimiter=DELIMITER)
        for h in reader:
            post = (h.get("PostOrder") or "").strip()
            if not post:
                continue
            hdr_by_post[post] = {
                "TransactionDate": (h.get("TransactionDate") or "").strip(),
                "JrnlKey_Journal": (h.get("JrnlKey_Journal") or "").strip(),
                "Module": (h.get("Module") or "").strip(),
            }

    latest_cost_by_vendor_item: Dict[tuple, Dict[str, str]] = {}
    with open(jrnlrow_path, newline="", encoding="utf-8-sig") as f_row:
        reader = csv.DictReader(f_row, delimiter=DELIMITER)
        for r in reader:
            vendor_record = (r.get("VendorRecordNumber") or "").strip()
            item_record = (r.get("ItemRecordNumber") or "").strip()
            if not vendor_record or vendor_record == "0" or not item_record or item_record == "0":
                continue
            qty = _safe_float(r.get("Quantity") or "")
            unit_cost = _safe_float(r.get("UnitCost") or "")
            if unit_cost <= 0:
                amount = abs(_safe_float(r.get("Amount") or ""))
                if qty > 0 and amount > 0:
                    unit_cost = amount / qty
            if unit_cost <= 0:
                continue
            post = (r.get("PostOrder") or "").strip()
            hdr = hdr_by_post.get(post, {})
            tx_date = (hdr.get("TransactionDate") or "").strip()
            row_number = int((r.get("RowNumber") or "0").strip() or 0)
            key = (vendor_record, item_record)
            current = latest_cost_by_vendor_item.get(key)
            candidate = {
                "VendorRecordNumber": vendor_record,
                "ItemRecordNumber": item_record,
                "SageUnitPrice": f"{unit_cost:.4f}",
                "SageLastPurchaseDate": tx_date,
                "SagePostOrder": post,
                "SageRowNumber": str(row_number),
                "SageJournal": (r.get("Journal") or "").strip(),
                "SageModule": (hdr.get("Module") or "").strip(),
            }
            if current is None:
                latest_cost_by_vendor_item[key] = candidate
            else:
                cur_date = current.get("SageLastPurchaseDate", "")
                if (tx_date, int(candidate["SagePostOrder"] or "0"), row_number) > (
                    cur_date,
                    int(current.get("SagePostOrder") or "0"),
                    int(current.get("SageRowNumber") or "0"),
                ):
                    latest_cost_by_vendor_item[key] = candidate

    _, vendors_rows = read_csv(vendors_sync_path)
    vendor_by_record = {str(v.get("VendorRecordNumber") or "").strip(): v for v in vendors_rows}

    _, product_rows = read_csv(products_sync_path)
    product_by_record = {str(p.get("ItemRecordNumber") or "").strip(): p for p in product_rows}

    _, supplier_rows = read_csv(odoo_vendor_pricelist_path)
    supplier_by_key: Dict[tuple, List[Dict[str, str]]] = defaultdict(list)
    for s in supplier_rows:
        vendor_id = str(s.get("OdooVendorId") or "").strip()
        tmpl_id = str(s.get("OdooTemplateId") or "").strip()
        if vendor_id and tmpl_id:
            supplier_by_key[(vendor_id, tmpl_id)].append(s)

    def pick_supplier(rows: List[Dict[str, str]]) -> Dict[str, str]:
        if not rows:
            return {}
        def _key(r: Dict[str, str]):
            min_qty = _safe_float(r.get("MinQty") or "")
            oid = int((r.get("OdooId") or "0").strip() or 0)
            return (min_qty, oid)
        return sorted(rows, key=_key)[0]

    sync_fields = [
        "VendorRecordNumber",
        "VendorID",
        "VendorName",
        "ItemRecordNumber",
        "ItemID",
        "ItemDescription",
        "SageUnitPrice",
        "SageLastPurchaseDate",
        "SagePostOrder",
        "SageRowNumber",
        "SageJournal",
        "SageModule",
        "OdooVendorId",
        "OdooVendorExternalId",
        "OdooVendorName",
        "OdooTemplateId",
        "OdooTemplateExternalId",
        "OdooVariantId",
        "OdooVariantExternalId",
        "OdooSupplierinfoId",
        "OdooSupplierinfoExternalId",
        "OdooUnitPrice",
        "Currency",
        "MinQty",
        "Delay",
        "VendorPricelistSyncStatus",
        "VendorPricelistMismatchFields",
        "LastLookupAt",
    ]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sync_rows: List[Dict[str, str]] = []
    new_rows: List[Dict[str, str]] = []
    update_rows: List[Dict[str, str]] = []
    conflict_rows: List[Dict[str, str]] = []

    for (_, _), src in latest_cost_by_vendor_item.items():
        vendor_record = src["VendorRecordNumber"]
        item_record = src["ItemRecordNumber"]
        v = vendor_by_record.get(vendor_record, {})
        p = product_by_record.get(item_record, {})
        odoo_vendor_id = (v.get("OdooId") or "").strip()
        odoo_tmpl_id = (p.get("OdooTemplateId") or "").strip()
        supplier_matches = supplier_by_key.get((odoo_vendor_id, odoo_tmpl_id), []) if odoo_vendor_id and odoo_tmpl_id else []
        supplier = pick_supplier(supplier_matches)
        odoo_price = _safe_float(supplier.get("OdooUnitPrice") or "")
        sage_price = _safe_float(src.get("SageUnitPrice") or "")
        mismatches: List[str] = []
        status = "NEW"
        if not odoo_vendor_id:
            mismatches.append("missing_odoo_vendor")
        if not odoo_tmpl_id:
            mismatches.append("missing_odoo_template")
        if supplier:
            if abs(odoo_price - sage_price) > 0.009:
                status = "UPDATE"
                mismatches.append("price")
            else:
                status = "MATCH"
        elif mismatches:
            status = "CONFLICT"
        else:
            status = "NEW"

        row = {
            "VendorRecordNumber": vendor_record,
            "VendorID": (v.get("VendorID") or "").strip(),
            "VendorName": (v.get("Name") or "").strip(),
            "ItemRecordNumber": item_record,
            "ItemID": (p.get("ItemID") or "").strip(),
            "ItemDescription": (p.get("ItemDescription") or "").strip(),
            "SageUnitPrice": f"{sage_price:.4f}",
            "SageLastPurchaseDate": src.get("SageLastPurchaseDate", ""),
            "SagePostOrder": src.get("SagePostOrder", ""),
            "SageRowNumber": src.get("SageRowNumber", ""),
            "SageJournal": src.get("SageJournal", ""),
            "SageModule": src.get("SageModule", ""),
            "OdooVendorId": odoo_vendor_id,
            "OdooVendorExternalId": (v.get("OdooExternalId") or "").strip(),
            "OdooVendorName": (v.get("OdooName") or "").strip(),
            "OdooTemplateId": odoo_tmpl_id,
            "OdooTemplateExternalId": (p.get("OdooTemplateExternalId") or "").strip(),
            "OdooVariantId": (p.get("OdooVariantId") or "").strip(),
            "OdooVariantExternalId": (p.get("OdooVariantExternalId") or "").strip(),
            "OdooSupplierinfoId": (supplier.get("OdooId") or "").strip(),
            "OdooSupplierinfoExternalId": (supplier.get("OdooExternalId") or "").strip(),
            "OdooUnitPrice": f"{odoo_price:.4f}" if supplier else "",
            "Currency": (supplier.get("Currency") or "USD").strip(),
            "MinQty": (supplier.get("MinQty") or "0").strip(),
            "Delay": (supplier.get("Delay") or "0").strip(),
            "VendorPricelistSyncStatus": status,
            "VendorPricelistMismatchFields": "|".join(mismatches),
            "LastLookupAt": now,
        }
        sync_rows.append(row)
        if status == "NEW":
            new_rows.append(row)
        elif status == "UPDATE":
            update_rows.append(row)
        elif status == "CONFLICT":
            conflict_rows.append(row)

    sync_rows.sort(key=lambda r: (int(r.get("VendorRecordNumber") or "0"), int(r.get("ItemRecordNumber") or "0")))
    new_rows.sort(key=lambda r: (int(r.get("VendorRecordNumber") or "0"), int(r.get("ItemRecordNumber") or "0")))
    update_rows.sort(key=lambda r: (int(r.get("VendorRecordNumber") or "0"), int(r.get("ItemRecordNumber") or "0")))

    os.makedirs(os.path.dirname(out_sync), exist_ok=True)
    write_csv(out_sync, sync_fields, sync_rows)
    write_csv(out_new, sync_fields, new_rows)
    write_csv(out_update, sync_fields, update_rows)
    write_csv(out_conflicts, sync_fields, conflict_rows)
    print(f"OK: vendor pricelist sync -> {out_sync} ({len(sync_rows)} rows)")
    print(f"OK: vendor pricelist NEW -> {out_new} ({len(new_rows)} rows)")
    print(f"OK: vendor pricelist UPDATE -> {out_update} ({len(update_rows)} rows)")
    print(f"OK: vendor pricelist CONFLICTS -> {out_conflicts} ({len(conflict_rows)} rows)")
    return 0


def _build_vendor_pricelist_xlsx(sync_rows: List[Dict[str, str]], template_path: str, out_path: str, is_update: bool) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        print("ERROR: openpyxl not available")
        return 0
    wb = load_workbook(template_path)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    count = 0

    has_vendor_db = "Vendor/Database ID" in headers
    has_vendor_ext = "Vendor/External ID" in headers
    has_variant_db = "Product Variant/Database ID" in headers
    has_variant_ext = "Product Variant/External ID" in headers

    for r in sync_rows:
        vendor_name = (r.get("OdooVendorName") or "").strip()
        vendor_external = (r.get("OdooVendorExternalId") or "").strip()
        vendor_db = (r.get("OdooVendorId") or "").strip()

        if has_vendor_db:
            vendor_value = vendor_db
        elif has_vendor_ext:
            vendor_value = vendor_external
        else:
            # Plain "Vendor" column in Odoo imports resolves by display name.
            vendor_value = vendor_name

        tmpl_external = _template_external_id_for_import(r.get("OdooTemplateExternalId") or "")
        variant_external = _short_external_id(r.get("OdooVariantExternalId") or "")
        variant_db = (r.get("OdooVariantId") or "").strip()
        if not vendor_value or not tmpl_external:
            continue
        if not is_update:
            currency = "USD"
            min_qty = "1"
            delay = "120"
        else:
            currency = (r.get("Currency") or "USD").strip()
            min_qty = (r.get("MinQty") or "1").strip()
            delay = (r.get("Delay") or "120").strip()
        out = {
            "Vendor": vendor_value,
            "Vendor/Database ID": vendor_db,
            "Vendor/External ID": vendor_external,
            "Product Template/External ID": tmpl_external,
            "Product Variant/External ID": (variant_external if has_variant_ext else ""),
            "Product Variant/Database ID": (variant_db if has_variant_db else ""),
            "Vendor Product Code": (r.get("ItemID") or "").strip(),
            "Vendor Product Name": (r.get("ItemDescription") or "").strip(),
            "Unit Price": (r.get("SageUnitPrice") or "").strip(),
            "Currency": currency,
            "min_qty": min_qty,
            "product_uom_id": "Units",
            "delay": delay,
        }
        if is_update:
            ext = _short_external_id(r.get("OdooSupplierinfoExternalId") or "")
            if not ext:
                continue
            out["EsternalId"] = ext
        ws.append([out.get(h, "") for h in headers])
        count += 1
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return count


def build_vendor_pricelist_import(args: argparse.Namespace) -> int:
    sync_path = args.sync_new
    template_path = args.template_path
    master_out = args.master_out
    out_path = args.out_path
    if "{date}" in out_path:
        out_path = out_path.format(date=datetime.now().strftime("%Y%m%d"))
    if not os.path.exists(sync_path):
        print(f"ERROR: sync NEW not found: {sync_path}")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2
    _, rows = read_csv(sync_path)
    c1 = _build_vendor_pricelist_xlsx(rows, template_path, master_out, is_update=False)
    c2 = _build_vendor_pricelist_xlsx(rows, template_path, out_path, is_update=False)
    print(f"OK: vendor pricelist NEW master -> {master_out} ({c1} rows)")
    print(f"OK: vendor pricelist NEW import -> {out_path} ({c2} rows)")
    return 0


def build_vendor_pricelist_update(args: argparse.Namespace) -> int:
    sync_path = args.sync_update
    template_path = args.template_path
    master_out = args.master_out
    out_path = args.out_path
    if "{date}" in out_path:
        out_path = out_path.format(date=datetime.now().strftime("%Y%m%d"))
    if not os.path.exists(sync_path):
        print(f"ERROR: sync UPDATE not found: {sync_path}")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2
    _, rows = read_csv(sync_path)
    c1 = _build_vendor_pricelist_xlsx(rows, template_path, master_out, is_update=True)
    c2 = _build_vendor_pricelist_xlsx(rows, template_path, out_path, is_update=True)
    print(f"OK: vendor pricelist UPDATE master -> {master_out} ({c1} rows)")
    print(f"OK: vendor pricelist UPDATE import -> {out_path} ({c2} rows)")
    return 0


def _vendor_city_state_zip(row: Dict[str, str]) -> str:
    city = (row.get("City") or "").strip()
    state = (row.get("State") or "").strip()
    zip_code = (row.get("Zip") or "").strip()
    parts = [p for p in [city, state, zip_code] if p]
    return " ".join(parts)


def build_vendors(args: argparse.Namespace) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        print("ERROR: openpyxl not available")
        return 2

    sync_path = args.vendors_sync
    template_path = args.template_path
    master_out = args.master_out
    out_path = args.out_path

    if not os.path.exists(sync_path):
        print(f"ERROR: vendors sync not found: {sync_path}")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2
    if "{date}" in out_path:
        out_path = out_path.format(date=datetime.now().strftime("%Y%m%d"))

    headers_sync, rows = read_csv(sync_path)
    selected = [
        r for r in rows
        if (r.get("VendorSyncStatus") or "").strip().upper() == "NEW"
        and not truthy(r.get("IsInactive"))
    ]

    # Keep an explicit CSV snapshot for review before import.
    csv_out = os.path.join(os.path.dirname(sync_path), "vendors_sync_NEW.csv")
    write_csv(csv_out, headers_sync, selected)

    def build_workbook(target_path: str) -> int:
        wb = load_workbook(template_path)
        ws = wb["9_VENDOR"] if "9_VENDOR" in wb.sheetnames else wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        count = 0
        for r in selected:
            vendor_name = (r.get("Name") or "").strip()
            if not vendor_name:
                continue
            out = {
                "Vendor_ID": (r.get("VendorID") or "").strip(),
                "name": vendor_name,
                "company_name": "",
                "Reference": (r.get("VendorID") or "").strip(),
                "is_company": "TRUE",
                "phone": (r.get("Phone") or "").strip(),
                "country": (r.get("Country") or "").strip(),
                "street": (r.get("Street") or "").strip(),
                "street2": (r.get("Street2") or "").strip(),
                "City ST ZIP": _vendor_city_state_zip(r),
                "Mail to City": (r.get("MailToCity") or "").strip(),
                "Mail to Zip": (r.get("MailToZip") or "").strip(),
                "Mail to Country": (r.get("MailToCountry") or "").strip(),
            }
            ws.append([out.get(h, "") for h in headers])
            count += 1
        os.makedirs(os.path.dirname(target_path), exist_ok=True)
        wb.save(target_path)
        return count

    count_master = build_workbook(master_out)
    count_import = build_workbook(out_path)
    print(f"OK: vendors NEW csv -> {csv_out} ({len(selected)} rows)")
    print(f"OK: vendors NEW master -> {master_out} ({count_master} rows)")
    print(f"OK: vendors NEW import -> {out_path} ({count_import} rows)")
    return 0


def build_vendors_update(args: argparse.Namespace) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        print("ERROR: openpyxl not available")
        return 2

    sync_path = args.vendors_sync
    template_path = args.template_path
    master_out = args.master_out
    out_path = args.out_path

    if not os.path.exists(sync_path):
        print(f"ERROR: vendors sync not found: {sync_path}")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2
    if "{date}" in out_path:
        out_path = out_path.format(date=datetime.now().strftime("%Y%m%d"))

    headers_sync, rows = read_csv(sync_path)
    selected = [
        r for r in rows
        if (r.get("VendorSyncStatus") or "").strip().upper() == "UPDATE"
        and not truthy(r.get("IsInactive"))
    ]
    csv_out = os.path.join(os.path.dirname(sync_path), "vendors_sync_UPDATE.csv")
    write_csv(csv_out, headers_sync, selected)

    def _safe_ext_id(raw: str, row: Dict[str, str]) -> str:
        ext = (raw or "").strip()
        if ext.startswith("__import__."):
            ext = ext.split(".", 1)[1]
        if not ext:
            ext = sanitize_external_id(row.get("VendorID", ""))
        if not ext:
            ext = sanitize_external_id(row.get("VendorRecordNumber", ""))
        return ext

    def build_workbook(target_path: str) -> int:
        wb = load_workbook(template_path)
        ws = wb["9_VENDOR"] if "9_VENDOR" in wb.sheetnames else wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        count = 0
        for r in selected:
            vendor_name = (r.get("Name") or "").strip()
            if not vendor_name:
                continue
            ext_id = _safe_ext_id(r.get("OdooExternalId", ""), r)
            if not ext_id:
                continue
            out = {
                "ExternalId": ext_id,
                "Vendor_ID": (r.get("VendorID") or "").strip(),
                "name": vendor_name,
                "company_name": "",
                "Reference": (r.get("VendorID") or "").strip(),
                "is_company": "TRUE",
                "phone": (r.get("Phone") or "").strip(),
                "country": (r.get("Country") or "").strip(),
                "street": (r.get("Street") or "").strip(),
                "street2": (r.get("Street2") or "").strip(),
                "City ST ZIP": _vendor_city_state_zip(r),
                "Mail to City": (r.get("MailToCity") or "").strip(),
                "Mail to Zip": (r.get("MailToZip") or "").strip(),
                "Mail to Country": (r.get("MailToCountry") or "").strip(),
            }
            ws.append([out.get(h, "") for h in headers])
            count += 1
        os.makedirs(os.path.dirname(target_path), exist_ok=True)
        wb.save(target_path)
        return count

    count_master = build_workbook(master_out)
    count_update = build_workbook(out_path)
    print(f"OK: vendors UPDATE csv -> {csv_out} ({len(selected)} rows)")
    print(f"OK: vendors UPDATE master -> {master_out} ({count_master} rows)")
    print(f"OK: vendors UPDATE import -> {out_path} ({count_update} rows)")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Sage ↔ Odoo parity tables")
    sub = parser.add_subparsers(dest="command", required=True)

    p1 = sub.add_parser("refresh_sage", help="Build sync tables from Sage master exports")
    p1.add_argument(
        "--customers-master",
        default=r"ENZO-Sage50\_master_sage\customers.csv",
    )
    p1.add_argument(
        "--items-master",
        default=r"ENZO-Sage50\_master_sage\items.csv",
    )
    p1.add_argument(
        "--vendors-master",
        default=r"ENZO-Sage50\_master_sage\vendors.csv",
    )
    p1.add_argument(
        "--customers-out",
        default=r"ENZO-Sage50\_master\customers_sync.csv",
    )
    p1.add_argument(
        "--items-out",
        default=r"ENZO-Sage50\_master\products_sync.csv",
    )
    p1.add_argument(
        "--vendors-out",
        default=r"ENZO-Sage50\_master\vendors_sync.csv",
    )
    p1.set_defaults(func=refresh_sage)

    p2 = sub.add_parser("refresh_odoo", help="Export Odoo masters to local CSVs")
    p2.add_argument(
        "--customers-out",
        default=r"ENZO-Sage50\_master_odoo\customers_odoo.csv",
    )
    p2.add_argument(
        "--items-out",
        default=r"ENZO-Sage50\_master_odoo\items_odoo.csv",
    )
    p2.add_argument(
        "--vendors-out",
        default=r"ENZO-Sage50\_master_odoo\vendors_odoo.csv",
    )
    p2.add_argument(
        "--chart-out",
        default=r"ENZO-Sage50\_master_odoo\chart_of_accounts_odoo.csv",
    )
    p2.add_argument(
        "--env-file",
        default=".env",
    )
    p2.add_argument(
        "--batch-size",
        type=int,
        default=1000,
        help="Batch size for Odoo export",
    )
    p2.add_argument(
        "--only",
        default="",
        help="Optional subset export. Supported: items_odoo",
    )
    p2.add_argument(
        "only_pos",
        nargs="?",
        default="",
        help=argparse.SUPPRESS,
    )
    p2.set_defaults(func=refresh_odoo)

    p3 = sub.add_parser("sync", help="Match Sage sync tables with Odoo masters")
    p3.add_argument(
        "--customers-sync",
        default=r"ENZO-Sage50\_master\customers_sync.csv",
    )
    p3.add_argument(
        "--items-sync",
        default=r"ENZO-Sage50\_master\products_sync.csv",
    )
    p3.add_argument(
        "--odoo-customers",
        default=r"ENZO-Sage50\_master_odoo\customers_odoo.csv",
    )
    p3.add_argument(
        "--odoo-items",
        default=r"ENZO-Sage50\_master_odoo\items_odoo.csv",
    )
    p3.add_argument(
        "--vendors-sync",
        default=r"ENZO-Sage50\_master\vendors_sync.csv",
    )
    p3.add_argument(
        "--odoo-vendors",
        default=r"ENZO-Sage50\_master_odoo\vendors_odoo.csv",
    )
    p3.add_argument(
        "--customer-match-name",
        action="store_true",
        help="Allow exact name match if ref match fails",
    )
    p3.add_argument(
        "--vendor-match-name",
        action="store_true",
        help="Allow exact name match for vendors if ref match fails",
    )
    p3.set_defaults(func=sync_local)

    p4 = sub.add_parser("build_contacts_sync", help="Build contacts sync file using Odoo data")
    p4.add_argument(
        "--customers-sync",
        default=r"ENZO-Sage50\_master\customers_sync.csv",
    )
    p4.add_argument(
        "--customers-master",
        default=r"ENZO-Sage50\_master_sage\customers.csv",
    )
    p4.add_argument(
        "--contacts-sync",
        default=r"ENZO-Sage50\_master\customer_contacts_sync.csv",
    )
    p4.add_argument(
        "--odoo-contacts",
        default=r"ENZO-Sage50\_master_odoo\customers_contacts.csv",
    )
    p4.set_defaults(func=build_contacts_sync)

    p5 = sub.add_parser("build_contacts", help="Build contacts import file from contacts sync")
    p5.add_argument(
        "--contacts-sync",
        default=r"ENZO-Sage50\_master\customer_contacts_sync.csv",
    )
    p5.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\NEW_customer_contacts.xlsx",
    )
    p5.set_defaults(func=build_contacts_import)

    p5b = sub.add_parser("build_addresses_sync", help="Build delivery addresses sync file from Sage contacts + addresses")
    p5b.add_argument(
        "--contacts-master",
        default=r"ENZO-Sage50\_master_sage\contacts.csv",
    )
    p5b.add_argument(
        "--address-master",
        default=r"ENZO-Sage50\_master_sage\address.csv",
    )
    p5b.add_argument(
        "--customers-sync",
        default=r"ENZO-Sage50\_master\customers_sync.csv",
    )
    p5b.add_argument(
        "--odoo-delivery",
        default=r"ENZO-Sage50\_master_odoo\customers_delivery_addresses.csv",
    )
    p5b.add_argument(
        "--country-parity",
        default=r"ENZO-Sage50\_master\_parity_country.csv",
    )
    p5b.add_argument(
        "--state-parity",
        default=r"ENZO-Sage50\_master\_parity_state.csv",
    )
    p5b.add_argument(
        "--countries-odoo",
        default=r"ENZO-Sage50\_master_odoo\countries_odoo.csv",
    )
    p5b.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\customer_delivery_addresses_sync.csv",
    )
    p5b.set_defaults(func=build_addresses_sync)

    p5c = sub.add_parser("build_delivery_addresses", help="Build delivery address import XLSX from sync file")
    p5c.add_argument(
        "--sync-path",
        default=r"ENZO-Sage50\_master\customer_delivery_addresses_sync.csv",
    )
    p5c.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\NEW_customer_delivery_address.xlsx",
    )
    p5c.set_defaults(func=build_delivery_import)

    p5cu = sub.add_parser("build_delivery_addresses_update", help="Build delivery address UPDATE XLSX from sync mismatches")
    p5cu.add_argument(
        "--sync-path",
        default=r"ENZO-Sage50\_master\customer_delivery_addresses_sync.csv",
    )
    p5cu.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\UPDATE_customers_delivery_address.xlsx",
    )
    p5cu.set_defaults(func=build_delivery_update)

    p5d = sub.add_parser("build_billto_sync", help="Build bill-to sync file from primary contacts + addresses")
    p5d.add_argument(
        "--contacts-master",
        default=r"ENZO-Sage50\_master_sage\contacts.csv",
    )
    p5d.add_argument(
        "--address-master",
        default=r"ENZO-Sage50\_master_sage\address.csv",
    )
    p5d.add_argument(
        "--customers-sync",
        default=r"ENZO-Sage50\_master\customers_sync.csv",
    )
    p5d.add_argument(
        "--country-parity",
        default=r"ENZO-Sage50\_master\_parity_country.csv",
    )
    p5d.add_argument(
        "--state-parity",
        default=r"ENZO-Sage50\_master\_parity_state.csv",
    )
    p5d.add_argument(
        "--countries-odoo",
        default=r"ENZO-Sage50\_master_odoo\countries_odoo.csv",
    )
    p5d.add_argument(
        "--odoo-children",
        default=r"ENZO-Sage50\_master_odoo\customers_child_partners_all.csv",
    )
    p5d.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\customers_billto_sync.csv",
    )
    p5d.set_defaults(func=build_billto_sync)

    p5e = sub.add_parser("build_billto", help="Build bill-to import XLSX from sync file")
    p5e.add_argument(
        "--sync-path",
        default=r"ENZO-Sage50\_master\customers_billto_sync.csv",
    )
    p5e.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\NEW_customer_billto.xlsx",
    )
    p5e.set_defaults(func=build_billto_import)

    p5eu = sub.add_parser("build_billto_update", help="Build bill-to UPDATE XLSX from sync mismatches")
    p5eu.add_argument(
        "--sync-path",
        default=r"ENZO-Sage50\_master\customers_billto_sync.csv",
    )
    p5eu.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\UPDATE_customers_billto.xlsx",
    )
    p5eu.set_defaults(func=build_billto_update)

    p5f = sub.add_parser("build_employees_sync", help="Build employees sync from Sage employees + sales orders")
    p5f.add_argument(
        "--root-dir",
        default=r"ENZO-Sage50",
        help="Root ENZO-Sage50 directory",
    )
    p5f.add_argument(
        "--months",
        default="2026_02,2026_03,2026_04",
        help="Comma-separated list of months (YYYY_MM) to include",
    )
    p5f.set_defaults(func=build_employees_sync)

    p5g = sub.add_parser("build_customers_update", help="Build customers_UPDATE from customers_sync mismatches")
    p5g.add_argument(
        "--customers-sync",
        default=r"ENZO-Sage50\_master\customers_sync.csv",
    )
    p5g.add_argument(
        "--customers-master",
        default=r"ENZO-Sage50\_master_sage\customers.csv",
    )
    p5g.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\UPDATE_customers.xlsx",
    )
    p5g.add_argument(
        "--master-out",
        default=r"ENZO-Sage50\_master\customers_UPDATE.xlsx",
    )
    p5g.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\odoo_UPDATE\{date}_customers_UPDATE.xlsx",
    )
    p5g.set_defaults(func=build_customers_update)

    p5h = sub.add_parser("build_vendors", help="Build vendors_NEW from vendors_sync")
    p5h.add_argument(
        "--vendors-sync",
        default=r"ENZO-Sage50\_master\vendors_sync.csv",
    )
    p5h.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\NEW_vendors.xlsx",
    )
    p5h.add_argument(
        "--master-out",
        default=r"ENZO-Sage50\_master\vendors_NEW.xlsx",
    )
    p5h.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\odoo_imports\{date}_vendors_NEW.xlsx",
    )
    p5h.set_defaults(func=build_vendors)

    p5i = sub.add_parser("build_vendors_update", help="Build vendors_UPDATE from vendors_sync")
    p5i.add_argument(
        "--vendors-sync",
        default=r"ENZO-Sage50\_master\vendors_sync.csv",
    )
    p5i.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\UPDATE_vendors.xlsx",
    )
    p5i.add_argument(
        "--master-out",
        default=r"ENZO-Sage50\_master\vendors_UPDATE.xlsx",
    )
    p5i.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\odoo_UPDATE\{date}_vendors_UPDATE.xlsx",
    )
    p5i.set_defaults(func=build_vendors_update)

    p5j = sub.add_parser("build_vendor_pricelist_sync", help="Build vendor_pricelist sync/NEW/UPDATE from Sage purchase history")
    p5j.add_argument(
        "--jrnlhdr-master",
        default=r"ENZO-Sage50\_master_sage\JrnlHdr.csv",
    )
    p5j.add_argument(
        "--jrnlrow-master",
        default=r"ENZO-Sage50\_master_sage\JrnlRow.csv",
    )
    p5j.add_argument(
        "--vendors-sync",
        default=r"ENZO-Sage50\_master\vendors_sync.csv",
    )
    p5j.add_argument(
        "--products-sync",
        default=r"ENZO-Sage50\_master\products_sync.csv",
    )
    p5j.add_argument(
        "--odoo-vendor-pricelist",
        default=r"ENZO-Sage50\_master_odoo\vendor_pricelist_odoo.csv",
    )
    p5j.add_argument(
        "--out-sync",
        default=r"ENZO-Sage50\_master\vendor_pricelist_sync.csv",
    )
    p5j.add_argument(
        "--out-new",
        default=r"ENZO-Sage50\_master\vendor_pricelist_sync_NEW.csv",
    )
    p5j.add_argument(
        "--out-update",
        default=r"ENZO-Sage50\_master\vendor_pricelist_sync_UPDATE.csv",
    )
    p5j.add_argument(
        "--out-conflicts",
        default=r"ENZO-Sage50\_master\vendor_pricelist_sync_CONFLICTS.csv",
    )
    p5j.set_defaults(func=build_vendor_pricelist_sync)

    p5k = sub.add_parser("build_vendor_pricelist_import", help="Build vendor pricelist NEW XLSX from vendor_pricelist_sync_NEW")
    p5k.add_argument(
        "--sync-new",
        default=r"ENZO-Sage50\_master\vendor_pricelist_sync_NEW.csv",
    )
    p5k.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\NEW_vendor_pricelist.xlsx",
    )
    p5k.add_argument(
        "--master-out",
        default=r"ENZO-Sage50\_master\vendor_pricelist_NEW.xlsx",
    )
    p5k.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\odoo_imports\{date}_vendor_pricelist_NEW.xlsx",
    )
    p5k.set_defaults(func=build_vendor_pricelist_import)

    p5l = sub.add_parser("build_vendor_pricelist_update", help="Build vendor pricelist UPDATE XLSX from vendor_pricelist_sync_UPDATE")
    p5l.add_argument(
        "--sync-update",
        default=r"ENZO-Sage50\_master\vendor_pricelist_sync_UPDATE.csv",
    )
    p5l.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\UPDATE_vendor_pricelist.xlsx",
    )
    p5l.add_argument(
        "--master-out",
        default=r"ENZO-Sage50\_master\vendor_pricelist_UPDATE.xlsx",
    )
    p5l.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\odoo_UPDATE\{date}_vendor_pricelist_UPDATE.xlsx",
    )
    p5l.set_defaults(func=build_vendor_pricelist_update)

    p6 = sub.add_parser("build_product_sync", help="Build product sync for a given YYYY_MM using invoice + credit note lines")
    p6.add_argument(
        "--year-month",
        required=True,
        help="Year and month in YYYY_MM format (e.g. 2026_02)",
    )
    p6.add_argument(
        "--base-dir",
        default=r"ENZO-Sage50",
        help="Base directory to search for invoice/credit note lines",
    )
    p6.add_argument(
        "--items-master",
        default=r"ENZO-Sage50\_master_sage\items.csv",
    )
    p6.add_argument(
        "--items-sync",
        default=r"ENZO-Sage50\_master\products_sync.csv",
    )
    p6.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\{year_month}_products_sync.csv",
        help="Output path (supports {year_month} placeholder)",
    )
    p6.set_defaults(func=build_product_sync)

    p7 = sub.add_parser("build_items_sync_new", help="Build products_sync_NEW with filters (no Odoo ID, active, barcode)")
    p7.add_argument(
        "--items-sync",
        default=r"ENZO-Sage50\_master\products_sync.csv",
    )
    p7.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\products_sync_NEW.csv",
    )
    p7.add_argument(
        "--invoice-base-dir",
        default=r"ENZO-Sage50",
        help="Base directory to search for 2026_02/2026_03 invoice lines",
    )
    p7.add_argument(
        "--barcode-digits",
        type=int,
        default=12,
        help="Require barcode to have exactly N digits (default: 12). Use 0 to disable.",
    )
    p7.set_defaults(func=build_items_sync_new)

    p7b = sub.add_parser("build_products_sync_nobarcode_new", help="Build products_sync_nobarcode_NEW (no Odoo ID, empty/short barcode)")
    p7b.add_argument(
        "--items-sync",
        default=r"ENZO-Sage50\_master\products_sync.csv",
    )
    p7b.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\products_sync_nobarcode_NEW.csv",
    )
    p7b.add_argument(
        "--barcode-digits",
        type=int,
        default=12,
        help="Require barcode to have at least N digits (default: 12). Use 0 to disable.",
    )
    p7b.add_argument(
        "--invoice-base-dir",
        default=r"ENZO-Sage50",
        help="Base directory to search for 2026_02/2026_03 invoice lines",
    )
    p7b.set_defaults(func=build_products_sync_nobarcode_new)

    p7c = sub.add_parser("build_products_import", help="Build products import XLSX from products_sync_NEW")
    p7c.add_argument(
        "--sync-path",
        default=r"ENZO-Sage50\_master\products_sync_NEW.csv",
    )
    p7c.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\NEW_products.xlsx",
    )
    p7c.set_defaults(func=build_products_import)

    p7d = sub.add_parser("build_products_nobarcode_import", help="Build products import XLSX from products_sync_nobarcode_NEW")
    p7d.add_argument(
        "--sync-path",
        default=r"ENZO-Sage50\_master\products_sync_nobarcode_NEW.csv",
    )
    p7d.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\NEW_products.xlsx",
    )
    p7d.set_defaults(func=build_products_nobarcode_import)

    p7e = sub.add_parser("build_pricelist_parity", help="Build parity files for pricelists/currencies")
    p7e.add_argument(
        "--root-dir",
        default=r"ENZO-Sage50",
    )
    p7e.set_defaults(func=build_pricelist_parity)

    p7f = sub.add_parser("build_pricelist_lines", help="Build pricelist_lines and pricelist_lines_NEW from Sage items")
    p7f.add_argument(
        "--root-dir",
        default=r"ENZO-Sage50",
    )
    p7f.add_argument(
        "--items-master",
        default=r"ENZO-Sage50\_master_sage\items.csv",
    )
    p7f.add_argument(
        "--items-odoo",
        default=r"ENZO-Sage50\_master_odoo\items_odoo.csv",
    )
    p7f.add_argument(
        "--pricelist-items-odoo",
        default=r"ENZO-Sage50\_master_odoo\pricelist_items_odoo.csv",
    )
    p7f.add_argument(
        "--parity-pricelist",
        default=r"ENZO-Sage50\_master\_parity_pricelist.csv",
    )
    p7f.set_defaults(func=build_pricelist_lines)

    p7g = sub.add_parser("build_pricelist_import", help="Build pricelist import CSV from pricelist_lines_NEW")
    p7g.add_argument(
        "--root-dir",
        default=r"ENZO-Sage50",
    )
    p7g.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\NEW_pricelist.csv",
    )
    p7g.add_argument(
        "--template-xlsx",
        default="",
        help="Optional XLSX template (if provided, first sheet is reused)",
    )
    p7g.add_argument(
        "--sync-path",
        default=r"ENZO-Sage50\_master\pricelist_lines_NEW.csv",
    )
    p7g.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\odoo_imports\{date}_pricelist.csv",
    )
    p7g.set_defaults(func=build_pricelist_import)

    p7h = sub.add_parser("build_pricelist_update", help="Build pricelist update CSV for changed existing lines")
    p7h.add_argument(
        "--root-dir",
        default=r"ENZO-Sage50",
    )
    p7h.add_argument(
        "--template-path",
        default=r"ENZO-Sage50\_master\odoo_templates\NEW_pricelist.csv",
    )
    p7h.add_argument(
        "--lines-path",
        default=r"ENZO-Sage50\_master\pricelist_lines.csv",
    )
    p7h.add_argument(
        "--pricelist-items-odoo",
        default=r"ENZO-Sage50\_master_odoo\pricelist_items_odoo.csv",
    )
    p7h.add_argument(
        "--items-odoo",
        default=r"ENZO-Sage50\_master_odoo\items_odoo.csv",
    )
    p7h.add_argument(
        "--pricelists-odoo",
        default=r"ENZO-Sage50\_master_odoo\pricelists_odoo.csv",
    )
    p7h.add_argument(
        "--out-path",
        default=r"ENZO-Sage50\_master\odoo_UPDATE\{date}_pricelist_UPDATE.csv",
    )
    p7h.set_defaults(func=build_pricelist_update)

    p8 = sub.add_parser("export_countries", help="Export Odoo countries + build Sage parity table (address only)")
    p8.add_argument(
        "--customers-sync",
        default=r"ENZO-Sage50\_master\customers_sync.csv",
    )
    p8.add_argument(
        "--customers-master",
        default=r"ENZO-Sage50\_master_sage\customers.csv",
    )
    p8.add_argument(
        "--odoo-customers",
        default=r"ENZO-Sage50\_master_odoo\customers_odoo.csv",
    )
    p8.add_argument(
        "--env-file",
        default=".env",
    )
    p8.add_argument(
        "--batch-size",
        type=int,
        default=1000,
        help="Batch size for Odoo export",
    )
    p8.set_defaults(func=export_countries)

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
