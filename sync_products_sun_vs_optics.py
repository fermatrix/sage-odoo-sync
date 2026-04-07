import argparse
import os
import re
import csv
from datetime import datetime
from typing import Dict, List

from sync_customers import load_env_file, get_env_value

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except Exception:
    openpyxl = None

try:
    import xmlrpc.client as xmlrpc_client
except Exception:
    xmlrpc_client = None


def _connect_odoo(env_file: str):
    if xmlrpc_client is None:
        raise RuntimeError("xmlrpc.client unavailable")
    env = load_env_file(env_file)
    url = get_env_value(env, "ODOO_STUDIOOPTYX_URL")
    db = get_env_value(env, "ODOO_STUDIOOPTYX_DB")
    user = get_env_value(env, "ODOO_STUDIOOPTYX_USER")
    apikey = get_env_value(env, "ODOO_STUDIOOPTYX_APIKEY")
    if not (url and db and user and apikey):
        raise RuntimeError("Missing Odoo credentials (URL/DB/USER/APIKEY)")
    url = url.rstrip("/")
    common = xmlrpc_client.ServerProxy(f"{url}/xmlrpc/2/common")
    uid = common.authenticate(db, user, apikey, {})
    if not uid:
        raise RuntimeError("Odoo authentication failed")
    models = xmlrpc_client.ServerProxy(f"{url}/xmlrpc/2/object")
    return db, uid, apikey, models


def _read_excel_rows(xlsx_path: str) -> List[Dict[str, str]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl not available")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows: List[Dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str):
            v = v.strip().upper()
        if v in {"E", "F"}:
            row_data = {headers[i - 1]: ws.cell(r, i).value for i in range(1, ws.max_column + 1)}
            row_data["_row"] = r
            rows.append(row_data)
    return rows


def _extract_barcode_conflict(msg: str) -> str:
    if not msg:
        return ""
    left = msg.split(" and ")[0]
    m = re.search(r"\[(.*?)\]\s*([^\n]+)", left)
    if not m:
        return ""
    return f"[{m.group(1)}] {m.group(2).strip()}"


def process_sun_vs_optics(args: argparse.Namespace) -> int:
    xlsx_path = args.xlsx_path
    log_path = args.log_path
    start = args.start
    limit = args.limit
    env_file = args.env_file

    if not os.path.exists(xlsx_path):
        print(f"ERROR: XLSX not found: {xlsx_path}")
        return 2

    rows = _read_excel_rows(xlsx_path)
    if not rows:
        print("ERROR: no rows with A=E/F found")
        return 2

    start_idx = max(0, start - 1)
    rows = rows[start_idx:start_idx + limit]
    if not rows:
        print("ERROR: slice produced no rows")
        return 2

    db, uid, apikey, models = _connect_odoo(env_file)

    log_fields = [
        "timestamp",
        "row",
        "id",
        "Item Description",
        "color_color_code",
        "product_code",
        "sku",
        "barcode",
        "APIPath",
        "APICalls",
        "status",
        "detail",
        "barcode_error",
    ]

    if openpyxl is None:
        print("ERROR: openpyxl not available for XLSX export")
        return 2

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LOG"
    for col_idx, name in enumerate(log_fields, start=1):
        ws.cell(row=1, column=col_idx, value=name)
    log_row = 2

    for row_data in rows:
            product_code = str(row_data.get("product_code_odoo") or "").strip()
            brand_model = str(row_data.get("brand_model") or "").strip()
            category_name = str(row_data.get("category") or "").strip()
            color_value = str(row_data.get("color_color_code") or row_data.get("color_code") or row_data.get("color") or "").strip()
            sku = str(row_data.get("id") or "").strip()
            barcode = str(row_data.get("barcode") or "").strip()
            barcode_log = f"'{barcode}" if barcode else ""
            item_desc = str(row_data.get("Item Description") or "").strip()
            color_code_col = str(row_data.get("color_color_code") or "").strip()

            api_calls = 0

            def call(model, method, *args, **kwargs):
                nonlocal api_calls
                api_calls += 1
                return models.execute_kw(db, uid, apikey, model, method, *args, **kwargs)

            path_label = ""

            ok_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

            def write_row(status: str, detail: str, barcode_error: str = "") -> None:
                nonlocal log_row
                values = [
                    datetime.now().isoformat(timespec="seconds"),
                    row_data.get("_row"),
                    sku,
                    item_desc,
                    color_code_col,
                    product_code,
                    sku,
                    barcode_log,
                    path_label,
                    api_calls,
                    status,
                    detail,
                    barcode_error,
                ]
                for col_idx, val in enumerate(values, start=1):
                    ws.cell(row=log_row, column=col_idx, value=val)
                if status == "OK":
                    for col_idx in range(1, len(values) + 1):
                        ws.cell(row=log_row, column=col_idx).fill = ok_fill
                log_row += 1

            if not product_code or not color_value:
                write_row("SKIP", "Missing product_code_odoo or color")
                continue

            # 0) If variant already exists by SKU, validate color + barcode and skip/update
            sku_match = call(
                "product.product",
                "search_read",
                [[("default_code", "=", sku)]],
                {"fields": ["id", "barcode", "is_storable", "product_template_attribute_value_ids"]},
            )
            if sku_match:
                path_label = "SHORT"
                v = sku_match[0]
                variant_id = v.get("id")
                current_barcode = (v.get("barcode") or "").strip()
                current_store = v.get("is_storable")
                ptav_ids = v.get("product_template_attribute_value_ids") or []

                if args.verify_color:
                    # Resolve color value name on variant
                    color_name = ""
                    if ptav_ids:
                        ptav_rows = call(
                            "product.template.attribute.value",
                            "read",
                            [ptav_ids],
                            {"fields": ["name", "attribute_id"]},
                        )
                        for pr in ptav_rows:
                            attr = pr.get("attribute_id") or []
                            attr_name = (attr[1] if isinstance(attr, list) and len(attr) > 1 else "") or ""
                            if attr_name.lower() == "color":
                                color_name = (pr.get("name") or "").strip()
                                break

                    if not color_name:
                        write_row("ERROR", "SKU exists but no color on variant")
                        continue
                    if color_name != color_value:
                        write_row("ERROR", f"SKU exists but color mismatch ({color_name})")
                        continue

                if current_barcode == barcode and bool(current_store) is True:
                    write_row("OK", "Already up to date")
                    continue

                # Update is_storable only if needed
                if not bool(current_store):
                    try:
                        call("product.product", "write", [[variant_id], {"is_storable": True}])
                    except Exception as e:
                        write_row("WARN", f"is_storable failed: {e}")

                if args.update_barcode and barcode:
                    try:
                        call("product.product", "write", [[variant_id], {"barcode": barcode}])
                        write_row("OK", "Barcode updated")
                    except Exception as e:
                        msg = str(e)
                        left = msg.split(" and ")[0]
                        conflict = _extract_barcode_conflict(left)
                        write_row("ERROR", "Barcode update failed", conflict)
                else:
                    if barcode:
                        if current_barcode and current_barcode != barcode:
                            detail = "Barcode mismatch (update disabled)"
                        elif current_barcode:
                            detail = "Barcode already set (update disabled)"
                        else:
                            detail = "Barcode missing (update disabled)"
                    else:
                        detail = "No barcode provided (update disabled)"
                    write_row("SKIP", detail)
                continue

            # Find product.template by external id __import__.product_code
            path_label = "LONG"
            model_data = call(
                "ir.model.data",
                "search_read",
                [[("model", "=", "product.template"), ("module", "=", "__import__"), ("name", "=", product_code)]],
                {"fields": ["res_id", "module", "name"]},
            )
            tmpl_id = None
            created = False
            if model_data:
                tmpl_id = model_data[0]["res_id"]
            else:
                # Create product.template
                categ_id = None
                if category_name:
                    c_ids = call(
                        "product.category",
                        "search",
                        [[("name", "=", category_name)]],
                        {"limit": 1},
                    )
                    if c_ids:
                        categ_id = c_ids[0]
                vals = {"name": brand_model}
                if categ_id:
                    vals["categ_id"] = categ_id
                tmpl_id = call("product.template", "create", [vals])
                call(
                    "ir.model.data",
                    "create",
                    [{
                        "name": product_code,
                        "module": "__import__",
                        "model": "product.template",
                        "res_id": tmpl_id,
                        "noupdate": True,
                    }],
                )
                created = True

            # Ensure color attribute value exists
            color_attrs = call(
                "product.attribute",
                "search_read",
                [[("name", "ilike", "color")]],
                {"fields": ["id", "name"]},
            )
            color_attr_ids = [a["id"] for a in color_attrs]
            if not color_attr_ids:
                write_row("ERROR", "No color attribute found")
                continue

            val_ids = call(
                "product.attribute.value",
                "search",
                [[("name", "=", color_value), ("attribute_id", "in", color_attr_ids)]],
            )
            if val_ids:
                color_val_id = val_ids[0]
            else:
                color_val_id = call(
                    "product.attribute.value",
                    "create",
                    [{"name": color_value, "attribute_id": color_attr_ids[0]}],
                )

            # Add color to product.template (attribute line)
            line_ids = call(
                "product.template",
                "read",
                [[tmpl_id]],
                {"fields": ["attribute_line_ids"]},
            )[0].get("attribute_line_ids") or []

            color_line_id = None
            if line_ids:
                lines = call(
                    "product.template.attribute.line",
                    "read",
                    [line_ids],
                    {"fields": ["attribute_id", "value_ids"]},
                )
                for line in lines:
                    attr = line.get("attribute_id") or []
                    if attr and attr[0] in color_attr_ids:
                        color_line_id = line["id"]
                        if color_val_id not in (line.get("value_ids") or []):
                            call(
                                "product.template.attribute.line",
                                "write",
                                [[color_line_id], {"value_ids": [(4, color_val_id)]}],
                            )
                        break

            if color_line_id is None:
                call(
                    "product.template",
                    "write",
                    [[tmpl_id], {"attribute_line_ids": [(0, 0, {"attribute_id": color_attr_ids[0], "value_ids": [(6, 0, [color_val_id])]} )]}],
                )

            # Find variant
            ptav_ids = call(
                "product.template.attribute.value",
                "search",
                [[("product_tmpl_id", "=", tmpl_id), ("product_attribute_value_id", "=", color_val_id)]],
            )
            if not ptav_ids:
                write_row("ERROR", "No product.template.attribute.value found")
                continue

            variant_ids = call(
                "product.product",
                "search",
                [[("product_tmpl_id", "=", tmpl_id), ("product_template_attribute_value_ids", "in", ptav_ids)]],
            )
            if not variant_ids:
                write_row("ERROR", "No variant found")
                continue

            variant_id = variant_ids[0]

            # Check current values to decide SKIP
            current = call(
                "product.product",
                "read",
                [[variant_id]],
                {"fields": ["default_code", "barcode", "is_storable"]},
            )[0]
            current_sku = (current.get("default_code") or "").strip()
            current_barcode = (current.get("barcode") or "").strip()
            current_store = current.get("is_storable")

            if current_sku == sku and current_barcode == barcode and bool(current_store) is True:
                detail = "Already up to date"
                if created:
                    detail = "Created product; already up to date"
                write_row("OK", detail)
                continue

            # Update SKU first
            try:
                call("product.product", "write", [[variant_id], {"default_code": sku}])
            except Exception as e:
                write_row("ERROR", f"SKU update failed: {e}")
                continue

            # Update is_storable only if needed
            if not bool(current_store):
                try:
                    call("product.product", "write", [[variant_id], {"is_storable": True}])
                except Exception as e:
                    write_row("WARN", f"is_storable failed: {e}")

            # Update barcode separately (always for created/long path)
            if barcode:
                try:
                    call("product.product", "write", [[variant_id], {"barcode": barcode}])
                    detail = "SKU + barcode updated"
                    if created:
                        detail = "Created product; SKU + barcode updated"
                    write_row("OK", detail)
                except Exception as e:
                    msg = str(e)
                    left = msg.split(" and ")[0]
                    conflict = _extract_barcode_conflict(left)
                    write_row("ERROR", "Barcode update failed", conflict)
            else:
                detail = "SKU updated; no barcode provided"
                if created:
                    detail = "Created product; SKU updated; no barcode provided"
                write_row("OK", detail)

    os.makedirs(os.path.dirname(log_path), exist_ok=True)
    wb.save(log_path)
    print(f"OK: processed {len(rows)} rows -> {log_path}")
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Process sun_vs_optics_USA variants (create product if needed)")
    parser.add_argument(
        "--xlsx-path",
        default=r"ENZO-Sage50\_master\odoo_imports\20260507_sun_vs_optics_USA.xlsx",
        help="Input XLSX with E/F rows",
    )
    parser.add_argument(
        "--log-path",
        default=r"ENZO-Sage50\_master\odoo_imports\20260507_sun_vs_optics_USA_LOG.xlsx",
        help="Output log XLSX",
    )
    parser.add_argument(
        "--start",
        type=int,
        default=1,
        help="1-based index among filtered E/F rows",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=100,
        help="Number of rows to process",
    )
    parser.add_argument(
        "--env-file",
        default=".env",
        help="Env file with Odoo credentials",
    )
    parser.add_argument(
        "--update-barcode",
        action="store_true",
        help="Update barcode on variant (default: disabled)",
    )
    parser.add_argument(
        "--verify-color",
        action="store_true",
        help="Verify variant color in step 0 (default: disabled)",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    return process_sun_vs_optics(args)


if __name__ == "__main__":
    raise SystemExit(main())
