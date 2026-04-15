import argparse
import os
from datetime import datetime
from typing import Dict, List

from parity_utils import (
    load_country_name_to_code,
    load_country_parity,
    load_state_parity,
    normalize_country,
)
from sync_customers import normalize_name, sanitize_external_id, read_csv, write_csv


def _norm_text(value: str) -> str:
    return normalize_name(value or "")


def _norm_text(value: str) -> str:
    return normalize_name(value or "")


def _norm_addr(street: str, street2: str, city: str, state: str, zip_code: str) -> str:
    parts = [street, street2, city, state, zip_code]
    return " ".join([_norm_text(p) for p in parts if p]).strip()


def _norm_email(value: str) -> str:
    return (value or "").strip().lower()


def _norm_phone(value: str) -> str:
    return "".join([c for c in (value or "") if c.isdigit()])


def _load_customers_by_record(customers_sync: str) -> Dict[str, Dict[str, str]]:
    if not os.path.exists(customers_sync):
        return {}
    _, rows = read_csv(customers_sync)
    by_record: Dict[str, Dict[str, str]] = {}
    for r in rows:
        crn = (r.get("CustomerRecordNumber") or "").strip()
        if crn:
            by_record[crn] = r
    return by_record


def _load_addresses_by_record(address_master: str) -> Dict[str, Dict[str, str]]:
    _, rows = read_csv(address_master)
    by_record: Dict[str, Dict[str, str]] = {}
    for r in rows:
        arn = (r.get("AddressRecordNumber") or "").strip()
        if arn:
            by_record[arn] = r
    return by_record


def _load_odoo_invoice_by_parent(odoo_children_csv: str) -> Dict[str, List[Dict[str, str]]]:
    existing: Dict[str, List[Dict[str, str]]] = {}
    if not os.path.exists(odoo_children_csv):
        return existing
    _, rows = read_csv(odoo_children_csv)
    for r in rows:
        if (r.get("Type") or "").strip().lower() != "invoice":
            continue
        parent_id = (r.get("ParentId") or "").strip()
        if not parent_id:
            continue
        existing.setdefault(parent_id, []).append(r)
    return existing


def _match_invoice(existing_rows: List[Dict[str, str]], ref: str) -> str:
    if not existing_rows:
        return ""
    n_ref = (ref or "").strip()
    if not n_ref:
        return ""
    for r in existing_rows:
        r_ref = (r.get("OdooRef", "") or "").strip()
        if r_ref and r_ref == n_ref:
            return str(r.get("OdooId", ""))
    return ""


def _match_invoice_row(existing_rows: List[Dict[str, str]], ref: str) -> Dict[str, str]:
    if not existing_rows:
        return {}
    n_ref = (ref or "").strip()
    if not n_ref:
        return {}
    for r in existing_rows:
        r_ref = (r.get("OdooRef", "") or "").strip()
        if r_ref and r_ref == n_ref:
            return r
    return {}


def _normalize_country(
    raw_country: str,
    country_parity: Dict[str, str],
    country_name_to_code: Dict[str, str],
) -> str:
    return normalize_country(raw_country, country_parity, country_name_to_code)


def build_billto_sync(args: argparse.Namespace) -> int:
    contacts_master = args.contacts_master
    address_master = args.address_master
    customers_sync = args.customers_sync
    out_path = args.out_path
    country_parity_path = args.country_parity
    state_parity_path = args.state_parity
    countries_odoo_path = args.countries_odoo
    odoo_children = args.odoo_children

    if not os.path.exists(contacts_master):
        print(f"ERROR: contacts master not found: {contacts_master}")
        return 2
    if not os.path.exists(address_master):
        print(f"ERROR: address master not found: {address_master}")
        return 2

    customers_by_record = _load_customers_by_record(customers_sync)
    addresses_by_record = _load_addresses_by_record(address_master)
    country_parity = load_country_parity(country_parity_path)
    country_name_to_code = load_country_name_to_code(countries_odoo_path)
    state_parity = load_state_parity(state_parity_path)
    existing_by_parent = _load_odoo_invoice_by_parent(odoo_children)

    _, contact_rows = read_csv(contacts_master)

    fieldnames = [
        "ContactRecordNumber",
        "CustomerRecordNumber",
        "CustomerID",
        "CustomerName",
        "AddressRecordNumber",
        "DeliveryName",
        "AddressTypeNumber",
        "AddressTypeDesc",
        "FirstName",
        "LastName",
        "Title",
        "Email",
        "Phone",
        "Street",
        "Street2",
        "City",
        "State",
        "Zip",
        "Country",
        "OdooState",
        "OdooCountry",
        "Notes",
        "OdooContactId",
        "OdooContactExternalId",
        "OdooParentId",
        "BilltoSyncStatus",
        "BilltoMismatchFields",
        "LastLookupAt",
    ]

    rows_out: List[Dict[str, str]] = []
    now_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for c in contact_rows:
        if (c.get("IsPrimaryContact") or "").strip() != "1":
            continue
        customer_record = (c.get("CustomerRecord") or "").strip()
        address_record = (c.get("AddressRecordNumber") or "").strip()
        if not customer_record or not address_record:
            continue

        delivery_name = (c.get("CompanyName") or "").strip()
        if not delivery_name:
            delivery_name = (customers_by_record.get(customer_record, {}).get("Customer_Bill_Name") or "").strip()
        first_name = (c.get("FirstName") or "").strip()
        last_name = (c.get("LastName") or "").strip()
        full_name = " ".join([p for p in [first_name, last_name] if p]).strip()
        match_name = full_name if full_name else delivery_name

        addr = addresses_by_record.get(address_record, {})
        cust = customers_by_record.get(customer_record, {})

        raw_country = (addr.get("Country") or "").strip()
        mapped_country = _normalize_country(raw_country, country_parity, country_name_to_code)

        raw_state = (addr.get("State") or "").strip()
        state_info = state_parity.get(raw_state, {})
        mapped_state = state_info.get("state_name") or raw_state
        if mapped_state and state_info.get("country_name"):
            inferred = country_name_to_code.get(state_info.get("country_name", ""), "")
            if inferred:
                mapped_state = f"{mapped_state} ({inferred})"
                if not mapped_country:
                    mapped_country = inferred

        odoo_parent_id = (cust.get("OdooId") or "").strip() if cust else ""
        match_row: Dict[str, str] = {}
        if odoo_parent_id:
            match_row = _match_invoice_row(
                existing_by_parent.get(odoo_parent_id, []),
                (c.get("RecordNumber") or "").strip(),
            )
        match_id = (match_row.get("OdooId") or "").strip()

        mismatches: List[str] = []
        if match_row:
            checks = [
                ("name", match_name, match_row.get("OdooName", "")),
                ("email", (c.get("Email") or "").strip(), match_row.get("OdooEmail", "")),
                ("street", (addr.get("AddressLine1") or "").strip(), match_row.get("Street", "")),
                ("street2", (addr.get("AddressLine2") or "").strip(), match_row.get("Street2", "")),
                ("city", (addr.get("City") or "").strip(), match_row.get("City", "")),
                ("zip", (addr.get("Zip") or "").strip(), match_row.get("Zip", "")),
                ("state", mapped_state, match_row.get("State", "")),
            ]
            mismatches = [name for name, left, right in checks if _norm_text(left) != _norm_text(right)]
            if _norm_phone((c.get("Telephone1") or "").strip()) != _norm_phone(match_row.get("OdooPhone", "")):
                mismatches.append("phone")

        rows_out.append({
            "ContactRecordNumber": (c.get("RecordNumber") or "").strip(),
            "CustomerRecordNumber": customer_record,
            "CustomerID": (cust.get("CustomerID") or "").strip(),
            "CustomerName": (cust.get("Customer_Bill_Name") or "").strip(),
            "AddressRecordNumber": address_record,
            "DeliveryName": delivery_name,
            "AddressTypeNumber": (addr.get("AddressTypeNumber") or "").strip(),
            "AddressTypeDesc": (addr.get("AddressTypeDesc") or "").strip(),
            "FirstName": first_name,
            "LastName": last_name,
            "Title": (c.get("Title") or "").strip(),
            "Email": (c.get("Email") or "").strip(),
            "Phone": (c.get("Telephone1") or "").strip(),
            "Street": (addr.get("AddressLine1") or "").strip(),
            "Street2": (addr.get("AddressLine2") or "").strip(),
            "City": (addr.get("City") or "").strip(),
            "State": raw_state,
            "Zip": (addr.get("Zip") or "").strip(),
            "Country": raw_country,
            "OdooState": mapped_state,
            "OdooCountry": mapped_country,
            "Notes": (c.get("Notes") or "").strip(),
            "OdooContactId": match_id,
            "OdooContactExternalId": (match_row.get("OdooExternalId") or "").strip(),
            "OdooParentId": odoo_parent_id,
            "BilltoSyncStatus": "UPDATE" if (match_id and mismatches) else ("MATCH" if match_id else "NEW"),
            "BilltoMismatchFields": "|".join(mismatches),
            "LastLookupAt": now_stamp,
        })

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    write_csv(out_path, fieldnames, rows_out)
    print(f"OK: billto sync rows: {len(rows_out)} -> {out_path}")
    return 0


def build_billto_import(args: argparse.Namespace) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        load_workbook = None

    if load_workbook is None:
        print("ERROR: openpyxl not available for XLSX export")
        return 2

    sync_path = args.sync_path
    template_path = args.template_path

    if not os.path.exists(sync_path):
        print(f"ERROR: billto sync file not found: {sync_path}")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2

    master_root = os.path.dirname(sync_path)
    odoo_imports = os.path.join(master_root, "odoo_imports")
    os.makedirs(odoo_imports, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d")
    out_xlsx = os.path.join(odoo_imports, f"{stamp}_customers_billto_NEW.xlsx")
    out_csv = os.path.join(master_root, "customers_billto_sync_NEW.csv")

    _, sync_rows = read_csv(sync_path)

    wb = load_workbook(template_path)
    ws = wb["Partners"] if "Partners" in wb.sheetnames else wb.active
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header:
            header_map[str(header).strip()] = col_idx

    def set_cell(col_name: str, value: str) -> None:
        col_idx = header_map.get(col_name)
        if col_idx:
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.delete_rows(2, ws.max_row)
    row_idx = 2
    emitted = 0
    new_rows: List[Dict[str, str]] = []
    for r in sync_rows:
        if (r.get("BilltoSyncStatus") or "").strip().upper() != "NEW":
            continue
        parent_id = (r.get("OdooParentId") or "").strip()
        if not parent_id:
            continue
        if (r.get("OdooContactId") or "").strip():
            continue

        first = (r.get("FirstName") or "").strip()
        last = (r.get("LastName") or "").strip()
        full_name = " ".join([p for p in [first, last] if p]).strip()
        company_name = (r.get("DeliveryName") or "").strip()
        name = full_name if full_name else company_name

        customer_id = (r.get("CustomerID") or "").strip()
        contact_record = (r.get("ContactRecordNumber") or "").strip()
        raw_ext = f"{customer_id}_{contact_record}" if customer_id and contact_record else ""
        ext_id = sanitize_external_id(raw_ext)

        new_rows.append(r)

        set_cell("External_ID", ext_id)
        set_cell("Parent/Database ID", parent_id)
        set_cell("Reference", contact_record)
        set_cell("is_company", 0)
        set_cell("type", "invoice")
        set_cell("Name", name)
        set_cell("Email", (r.get("Email") or "").strip())
        set_cell("Phone", (r.get("Phone") or "").strip())
        set_cell("Job Position", (r.get("Title") or "").strip())
        set_cell("Street", (r.get("Street") or "").strip())
        set_cell("Street2", (r.get("Street2") or "").strip())
        set_cell("City", (r.get("City") or "").strip())
        set_cell("State", (r.get("OdooState") or "").strip() or (r.get("State") or "").strip())
        set_cell("ZIP", (r.get("Zip") or "").strip())
        set_cell("Country", (r.get("OdooCountry") or "").strip() or (r.get("Country") or "").strip())
        set_cell("Notes", (r.get("Notes") or "").strip())
        set_cell("Language", "English (US)")
        row_idx += 1
        emitted += 1

    wb.save(out_xlsx)
    fieldnames = list(new_rows[0].keys()) if new_rows else []
    write_csv(out_csv, fieldnames, new_rows)
    print(f"OK: billto new rows: {emitted} -> {out_xlsx}")
    print(f"OK: billto NEW sync -> {out_csv}")
    return 0


def build_billto_update(args: argparse.Namespace) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        load_workbook = None

    if load_workbook is None:
        print("ERROR: openpyxl not available for XLSX export")
        return 2

    sync_path = args.sync_path
    template_path = args.template_path

    if not os.path.exists(sync_path):
        print(f"ERROR: billto sync file not found: {sync_path}")
        return 2
    if not os.path.exists(template_path):
        print(f"ERROR: template not found: {template_path}")
        return 2

    master_root = os.path.dirname(sync_path)
    odoo_update = os.path.join(master_root, "odoo_UPDATE")
    os.makedirs(odoo_update, exist_ok=True)

    stamp = datetime.now().strftime("%Y%m%d")
    out_xlsx = os.path.join(odoo_update, f"{stamp}_customers_billto_UPDATE.xlsx")
    out_csv = os.path.join(master_root, "customers_billto_sync_UPDATE.csv")

    _, sync_rows = read_csv(sync_path)
    update_rows = [
        r for r in sync_rows
        if (r.get("BilltoSyncStatus") or "").strip().upper() == "UPDATE"
        and (r.get("OdooParentId") or "").strip()
        and (r.get("OdooContactExternalId") or "").strip()
    ]

    wb = load_workbook(template_path)
    ws = wb["Partners"] if "Partners" in wb.sheetnames else wb.active
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header:
            header_map[str(header).strip()] = col_idx

    def set_cell(col_name: str, value: str) -> None:
        col_idx = header_map.get(col_name)
        if col_idx:
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.delete_rows(2, ws.max_row)
    row_idx = 2
    emitted = 0
    for r in update_rows:
        ext_id = (r.get("OdooContactExternalId") or "").strip()
        if ext_id.startswith("__import__."):
            ext_id = ext_id.split(".", 1)[1]
        parent_id = (r.get("OdooParentId") or "").strip()
        first = (r.get("FirstName") or "").strip()
        last = (r.get("LastName") or "").strip()
        full_name = " ".join([p for p in [first, last] if p]).strip()
        company_name = (r.get("DeliveryName") or "").strip()
        name = full_name if full_name else company_name
        if not ext_id or not parent_id or not name:
            continue

        set_cell("External_ID", ext_id)
        set_cell("Parent/Database ID", parent_id)
        set_cell("Reference", (r.get("ContactRecordNumber") or "").strip())
        set_cell("is_company", 0)
        set_cell("type", "invoice")
        set_cell("Name", name)
        set_cell("Email", (r.get("Email") or "").strip())
        set_cell("Phone", (r.get("Phone") or "").strip())
        set_cell("Job Position", (r.get("Title") or "").strip())
        set_cell("Street", (r.get("Street") or "").strip())
        set_cell("Street2", (r.get("Street2") or "").strip())
        set_cell("City", (r.get("City") or "").strip())
        set_cell("State", (r.get("OdooState") or "").strip() or (r.get("State") or "").strip())
        set_cell("ZIP", (r.get("Zip") or "").strip())
        set_cell("Notes", (r.get("Notes") or "").strip())
        set_cell("Language", "English (US)")
        row_idx += 1
        emitted += 1

    wb.save(out_xlsx)
    fieldnames = list(update_rows[0].keys()) if update_rows else []
    write_csv(out_csv, fieldnames, update_rows)
    print(f"OK: billto update rows: {emitted} -> {out_xlsx}")
    print(f"OK: billto UPDATE sync -> {out_csv}")
    return 0
